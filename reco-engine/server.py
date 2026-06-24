from __future__ import annotations

import logging
logging.basicConfig(level=logging.INFO, format="%(name)s %(levelname)s: %(message)s")

from email.parser import BytesParser
from email.policy import default
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
import json
import os
from pathlib import Path
import sys
import re
import threading
from urllib.parse import parse_qs, urlparse
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers as xl_numbers
)
from openpyxl.formatting.rule import CellIsRule, FormulaRule

ROOT = Path(__file__).resolve().parent
STATIC = ROOT / "static"
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from recon.core import MatchResult, reconcile, summarize
from recon.gstr_2a_2b_books import read_three_way_uploads, reconcile_three_way, summarize_three_way
from recon.gstr_3b_vs_2b import read_3b_2b_uploads, reconcile_3b_vs_2b, summarize_3b_vs_2b, build_month_pivot
from recon.gstr_1_vs_books import (
    read_octa_excel, read_tally_sales_raw, read_credit_note_raw,
    parse_gstr1_pdf_monthly,
    extract_gstr3b_monthly, aggregate_gstr1_monthly, aggregate_books_monthly,
    build_monthly_comparison, reconcile_b2b_new, reconcile_b2c_new,
    df_to_records, build_summary,
)
from recon.parsers import read_upload, read_excel_rows, normalize_rows
from recon.bank_reco import process_bank_statement
from recon.gstr_2b_books import reconcile_gstr2b_vs_books
from recon.gstr_2b_books_multistate import (
    reconcile_gstr2b_vs_books_multistate,
    build_gstr2b_books_multistate_workbook,
)
from recon.gstr_3b_tally_entry import (
    parse_gstr3b,
    build_tally_entries,
    build_gstr3b_tally_workbook,
    process_multi as gstr3b_process_multi,
    _load_coa,
    _load_coa_from_list,
    _load_voucher_types,
    _load_voucher_types_from_list,
)


JOBS: dict[str, dict] = {}

# Limit simultaneous reconciliation jobs to prevent OOM under heavy load.
# Each job can hold large DataFrames in memory; 8 concurrent runs is safe
# for a server with 8+ GB RAM. Override via MAX_CONCURRENT_RECO env var.
_MAX_RECO = int(os.environ.get("MAX_CONCURRENT_RECO", "8"))
_RECO_SEMAPHORE = threading.Semaphore(_MAX_RECO)


class ReconciliationHandler(BaseHTTPRequestHandler):
    server_version = "CARecon/0.1"

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path
        if path == "/":
            self.serve_static("index.html")
            return
        if path.startswith("/static/"):
            self.serve_static(path.removeprefix("/static/"))
            return
        if path.startswith("/api/jobs/") and path.endswith("/export.xlsx"):
            job_id = path.split("/")[3]
            self.export_job(job_id)
            return
        if path.startswith("/api/jobs/"):
            job_id = path.split("/")[3]
            self.write_json(JOBS.get(job_id) or {"error": "Job not found"}, 200 if job_id in JOBS else 404)
            return
        self.write_json({"error": "Not found"}, 404)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path != "/api/reconcile":
            self.write_json({"error": "Not found"}, 404)
            return
        # Reject immediately if the server is at capacity — avoids OOM
        if not _RECO_SEMAPHORE.acquire(blocking=False):
            self.write_json(
                {"error": "Server busy. Too many reconciliations running. Retry in 30 seconds."},
                503,
            )
            return
        try:
            fields, files = self.read_multipart()
            tolerance = float(fields.get("tolerance", "1") or 1)
            reco_type = fields.get("reco_type", "gst_2b_purchase")

            if reco_type == "gstr_3b_vs_2b":
                gstr2b_records, gstr3b_records = read_3b_2b_uploads(files)
                results = reconcile_3b_vs_2b(gstr2b_records, gstr3b_records, tolerance=tolerance)
                pivot = build_month_pivot(results)
                job_id = uuid4().hex
                payload = {
                    "job_id": job_id,
                    "reco_type": reco_type,
                    "summary": summarize_3b_vs_2b(results),
                    "pivot": pivot,
                    "counts": {
                        "gstr2b_records": len(gstr2b_records),
                        "gstr3b_records": len(gstr3b_records),
                        "result_rows": len(results),
                    },
                    "results": [result.as_dict() for result in results],
                }
                JOBS[job_id] = payload
                self.write_json(payload)
                return

            if reco_type == "gstr_2a_2b_books":
                gstr2a_records, gstr2b_records, books_records = read_three_way_uploads(files)
                results = reconcile_three_way(gstr2a_records, gstr2b_records, books_records, tolerance=tolerance)
                job_id = uuid4().hex
                payload = {
                    "job_id": job_id,
                    "reco_type": reco_type,
                    "summary": summarize_three_way(results),
                    "counts": {
                        "gstr2a_records": len(gstr2a_records),
                        "gstr2b_records": len(gstr2b_records),
                        "books_records": len(books_records),
                        "result_rows": len(results),
                    },
                    "results": [result.as_dict() for result in results],
                }
                JOBS[job_id] = payload
                self.write_json(payload)
                return

            if reco_type == "gstr_1_vs_books":
                octa_file   = files.get("gstr1_octa") or files.get("gstr1")
                tally_file  = files.get("tally_sales")
                pdf_file    = files.get("gstr1_pdf")
                cn_file     = files.get("credit_note")
                if not octa_file:
                    self.write_json({"error": "Upload the GSTR-1 OCTA Report file (gstr1_octa)."}, 400)
                    return
                if not tally_file:
                    self.write_json({"error": "Upload the Tally Sales Register file (tally_sales)."}, 400)
                    return

                # Read inputs
                gstr1_df, gstr3b_df, gstr2b_df = read_octa_excel(octa_file)
                tally_df = read_tally_sales_raw(tally_file)
                cn_df = read_credit_note_raw(cn_file) if cn_file else None

                # Step 0 (optional): GSTR-1 Pivot
                pdf_monthly = None
                if pdf_file and pdf_file.get("content"):
                    pdf_monthly = parse_gstr1_pdf_monthly(pdf_file["content"])

                # Steps 1–4: monthly comparison sections
                gstr1_monthly   = aggregate_gstr1_monthly(gstr1_df)
                gstr3b_monthly  = extract_gstr3b_monthly(gstr3b_df)
                books_all       = aggregate_books_monthly(tally_df)
                books_b2b       = aggregate_books_monthly(tally_df, category="B2B")
                books_b2c       = aggregate_books_monthly(tally_df, category="B2C")

                # Filter GSTR-1 B2B/B2C monthly totals
                import re as _re
                gstr1_b2b_monthly = {}
                gstr1_b2c_monthly = {}
                if not gstr1_df.empty:
                    from recon.gstr_1_vs_books import _find_col, _f, _norm_month, _FY_MONTHS, _zero_amounts, _add_amounts
                    period_col  = _find_col(gstr1_df, ["Tax Period", "Period"])
                    gstin_col   = _find_col(gstr1_df, ["Customer GSTIN", "GSTIN of Recipient"])
                    taxable_col = _find_col(gstr1_df, ["Item Taxable Value", "Taxable Value"])
                    igst_col    = _find_col(gstr1_df, ["IGST", "Integrated Tax"])
                    cgst_col    = _find_col(gstr1_df, ["CGST", "Central Tax"])
                    sgst_col    = _find_col(gstr1_df, ["SGST", "State Tax"])
                    from collections import defaultdict as _dd
                    _b2b = _dd(_zero_amounts)
                    _b2c = _dd(_zero_amounts)
                    for _, row in gstr1_df.iterrows():
                        from recon.gstr_1_vs_books import _col_val
                        month = _norm_month(_col_val(row, period_col))
                        if month not in _FY_MONTHS:
                            continue
                        gstin = str(_col_val(row, gstin_col, "")).strip()
                        is_b2b = bool(gstin) and len(_re.sub(r"[^A-Z0-9]","",gstin.upper())) == 15
                        bucket = _b2b[month] if is_b2b else _b2c[month]
                        _add_amounts(bucket,
                            taxable=_f(_col_val(row, taxable_col, 0)),
                            igst=_f(_col_val(row, igst_col, 0)),
                            cgst=_f(_col_val(row, cgst_col, 0)),
                            sgst=_f(_col_val(row, sgst_col, 0)))
                    gstr1_b2b_monthly = dict(_b2b)
                    gstr1_b2c_monthly = dict(_b2c)

                gst_reco_sections = {
                    "gstr1_vs_gstr3b":    build_monthly_comparison(gstr1_monthly,  gstr3b_monthly,  "gstr1", "gstr3b"),
                    "books_all_vs_gstr1": build_monthly_comparison(books_all,       gstr1_monthly,   "books", "gstr1"),
                    "books_b2b_vs_gstr1": build_monthly_comparison(books_b2b,       gstr1_b2b_monthly, "books", "gstr1"),
                    "books_b2c_vs_gstr1": build_monthly_comparison(books_b2c,       gstr1_b2c_monthly, "books", "gstr1"),
                }

                # Step 5: B2B Reco
                b2b_rows = reconcile_b2b_new(tally_df, gstr1_df, tolerance)

                # Step 6: B2C Reco
                b2c_rows = reconcile_b2c_new(tally_df, gstr1_df, tolerance)

                # GSTR-1 Pivot (Step 0)
                pivot_rows = None
                if pdf_monthly and gstr1_monthly:
                    pivot_rows = build_monthly_comparison(gstr1_monthly, pdf_monthly, "excel", "pdf")

                # Slim b2b rows for UI display (key fields only — avoids sending 52-col Tally data)
                from recon.gstr_1_vs_books import _find_col as _fc2
                _inv_k  = _fc2(tally_df, ["Voucher No.", "Voucher No", "Invoice No", "Doc No", "Bill No"])
                _date_k = _fc2(tally_df, ["Date", "Invoice Date", "Voucher Date"])
                _part_k = _fc2(tally_df, ["Particulars", "Party Name", "Buyer", "Ledger Name"])
                _gst_k  = _fc2(tally_df, ["GSTIN", "GSTIN/UIN", "Buyer GSTIN"])
                _tax_k  = _fc2(tally_df, ["Total Sales", "Taxable Value", "Taxable Amount"])
                _igst_k = _fc2(tally_df, ["Total IGST", "IGST"])
                _cgst_k = _fc2(tally_df, ["Total CGST", "CGST"])
                _sgst_k = _fc2(tally_df, ["Total SGST", "SGST"])
                b2b_ui_rows = [{
                    "date":         _r.get(_date_k),
                    "inv_no":       _r.get(_inv_k),
                    "party":        _r.get(_part_k),
                    "gstin":        _r.get(_gst_k),
                    "t_taxable":    _r.get(_tax_k, 0),
                    "t_igst":       _r.get(_igst_k, 0),
                    "t_cgst":       _r.get(_cgst_k, 0),
                    "t_sgst":       _r.get(_sgst_k, 0),
                    "g1_inv":       _r.get("_gstr1_inv_no"),
                    "g1_taxable":   _r.get("_gstr1_taxable", 0),
                    "g1_igst":      _r.get("_gstr1_igst", 0),
                    "g1_cgst":      _r.get("_gstr1_cgst", 0),
                    "g1_sgst":      _r.get("_gstr1_sgst", 0),
                    "diff_taxable": _r.get("_diff_taxable", 0),
                    "diff_igst":    _r.get("_diff_igst", 0),
                    "diff_cgst":    _r.get("_diff_cgst", 0),
                    "diff_sgst":    _r.get("_diff_sgst", 0),
                    "remark":       _r.get("_remark"),
                } for _r in b2b_rows]

                import base64
                from io import BytesIO as _BytesIO
                job_id = uuid4().hex
                payload = {
                    "job_id": job_id,
                    "reco_type": reco_type,
                    "summary": build_summary(b2b_rows, b2c_rows),
                    "counts": {
                        "tally_rows":    len(tally_df),
                        "gstr1_rows":    len(gstr1_df),
                        "b2b_reco_rows": len(b2b_rows),
                        "b2c_reco_rows": len(b2c_rows),
                        "total_records": len(b2b_rows) + len(b2c_rows),
                    },
                    "results": [],  # not used by frontend for this agent
                    # Private: reco data used by workbook builder
                    "_gst_reco_sections": gst_reco_sections,
                    "_b2b_reco_rows": b2b_rows,
                    "_b2c_reco_rows": b2c_rows,
                    "_pivot_rows": pivot_rows,
                    "_tally_cols": list(tally_df.columns),
                    # Raw DataFrames as records for passthrough sheets
                    "_raw_gstr1":  df_to_records(gstr1_df),
                    "_raw_gstr2b": df_to_records(gstr2b_df),
                    "_raw_gstr3b": df_to_records(gstr3b_df),
                    "_raw_tally":  df_to_records(tally_df),
                    "_raw_cn":     df_to_records(cn_df) if cn_df is not None else None,
                }
                # Pre-build workbook now (during upload) so download is instant
                try:
                    _wb = build_gstr1_workbook(
                        [], monthly_summary=[], summary=payload["summary"],
                        counts=payload["counts"], payload=payload,
                    )
                    _buf = _BytesIO()
                    _wb.save(_buf)
                    payload["_xlsx_bytes"] = _buf.getvalue()
                except Exception as _e:
                    import logging as _log
                    _log.getLogger(__name__).error("Pre-build workbook failed: %s", _e)
                    payload["_xlsx_bytes"] = None
                JOBS[job_id] = payload
                # Return only public fields to frontend (plus UI display data)
                public = {k: v for k, v in payload.items() if not k.startswith("_")}
                public["gst_reco_sections"] = gst_reco_sections
                public["b2b_ui_rows"] = b2b_ui_rows
                public["b2c_rows"] = b2c_rows
                self.write_json(public)
                return

            if reco_type == "bank_reco":
                bank_file = files.get("bank_statement")
                if not bank_file:
                    self.write_json({"error": "Upload Bank Statement file."}, 400)
                    return
                
                payload = process_bank_statement(bank_file["content"])
                JOBS[payload["job_id"]] = payload
                self.write_json(payload)
                return

            # GSTR-2B vs Books (Purchase Register + Debit Note Register)
            if reco_type == "gstr_2b_books":
                gstr2b_file = files.get("gstr2b")
                purchase_file = files.get("purchase")
                debit_file = files.get("debit")
                if not gstr2b_file or not purchase_file or not debit_file:
                    self.write_json({"error": "Upload GSTR-2B, Purchase Register, and Debit Note Register files."}, 400)
                    return
                from recon.gstr_2b_books import _ensure_xlsx
                gstr2b_bytes = _ensure_xlsx(gstr2b_file["content"])
                purchase_bytes = _ensure_xlsx(purchase_file["content"])
                debit_bytes = _ensure_xlsx(debit_file["content"])
                gstr2b_records, books_records, results = reconcile_gstr2b_vs_books(
                    gstr2b_bytes,
                    purchase_bytes,
                    debit_bytes,
                    tolerance=tolerance,
                )
                job_id = uuid4().hex
                import base64
                payload = {
                    "job_id": job_id,
                    "reco_type": reco_type,
                    "summary": summarize(results),
                    "counts": {
                        "gstr2b_records": len(gstr2b_records),
                        "books_records": len(books_records),
                        "result_rows": len(results),
                    },
                    "results": [result.as_dict() for result in results],
                    "_gstr2b_b64": base64.b64encode(gstr2b_bytes).decode("utf-8"),
                    "_purchase_b64": base64.b64encode(purchase_bytes).decode("utf-8"),
                    "_debit_b64": base64.b64encode(debit_bytes).decode("utf-8"),
                }
                JOBS[job_id] = payload
                self.write_json(payload)
                return

            # GSTR-2B vs Books — Multi-State (N files per input type)
            if reco_type == "gstr_2b_books_multistate":
                def _file_list(name):
                    """Normalise single dict or list-of-dicts into list of content bytes."""
                    val = files.get(name)
                    if val is None:
                        return []
                    items = val if isinstance(val, list) else [val]
                    return [item["content"] for item in items if item.get("content")]

                def _file_items(name):
                    """Return raw list of file dicts (keeps filename alongside content)."""
                    val = files.get(name)
                    if val is None:
                        return []
                    items = val if isinstance(val, list) else [val]
                    return [item for item in items if item.get("content")]

                _GSTIN_IN_FILENAME = re.compile(
                    r'[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][0-9A-Z]Z[0-9A-Z]',
                    re.IGNORECASE,
                )

                def _entity_gstin_from_filename(fn: str) -> str:
                    """Extract entity GSTIN from GSTN portal filename (e.g. 102024_29AAECF7751Q1ZS_GSTR2B_...)."""
                    m = _GSTIN_IN_FILENAME.search(fn.upper())
                    return m.group(0).upper() if m else ""

                gstr2b_items  = _file_items("gstr2b")
                gstr2b_list   = [it["content"] for it in gstr2b_items]
                entity_gstins = [_entity_gstin_from_filename(it.get("filename", "")) for it in gstr2b_items]

                purchase_list = _file_list("purchase")
                debit_list    = _file_list("debit")

                if not gstr2b_list or not purchase_list:
                    self.write_json({"error": "Upload at least one GSTR-2B and one Purchase Register file."}, 400)
                    return

                gstr2b_recs, books_recs, results = reconcile_gstr2b_vs_books_multistate(
                    gstr2b_list, purchase_list, debit_list or [b""] * len(purchase_list),
                    tolerance=tolerance,
                    entity_gstins=entity_gstins,
                )

                import base64
                job_id  = uuid4().hex
                payload = {
                    "job_id":    job_id,
                    "reco_type": reco_type,
                    "summary":   summarize(results),
                    "counts": {
                        "gstr2b_records": len(gstr2b_recs),
                        "books_records":  len(books_recs),
                        "result_rows":    len(results),
                        "file_count":     len(gstr2b_list),
                    },
                    "results": [result.as_dict() for result in results],
                    # First file of each type (for base workbook source sheets — state 1)
                    "_gstr2b_b64":   base64.b64encode(gstr2b_list[0]).decode("utf-8"),
                    "_purchase_b64": base64.b64encode(purchase_list[0]).decode("utf-8"),
                    "_debit_b64":    base64.b64encode(debit_list[0]).decode("utf-8") if debit_list else "",
                    # All state files (for adding per-state source sheets to the workbook)
                    "_all_gstr2b_b64":   [base64.b64encode(f).decode("utf-8") for f in gstr2b_list],
                    "_all_purchase_b64": [base64.b64encode(f).decode("utf-8") for f in purchase_list if f],
                    "_all_debit_b64":    [base64.b64encode(f).decode("utf-8") for f in debit_list   if f],
                    # Stash MatchResult objects so export_job can rebuild Remark 3
                    "_results_obj":  results,
                }
                JOBS[job_id] = payload
                self.write_json({k: v for k, v in payload.items() if not k.startswith("_")})
                return

            if reco_type == "gstr_3b_tally_entry":
                # Accept 1–15 gstr3b files (repeated field name) + optional coa file
                def _gstr3b_file_items(name):
                    val = files.get(name)
                    if val is None:
                        return []
                    items = val if isinstance(val, list) else [val]
                    return [item for item in items if item.get("content")]

                gstr3b_files = _gstr3b_file_items("gstr3b")
                if not gstr3b_files:
                    self.write_json({"error": "Upload at least one GSTR-3B file."}, 400)
                    return

                coa_file = files.get("coa")
                if isinstance(coa_file, list):
                    coa_file = coa_file[0]

                vt_file = files.get("vouchertype")
                if isinstance(vt_file, list):
                    vt_file = vt_file[0]

                # COA resolution: uploaded file → DB JSON list → None
                if coa_file:
                    coa = _load_coa(coa_file["content"], coa_file["filename"])
                    coa_parsed_list = list(coa.values())
                elif fields.get("coa_ledgers"):
                    try:
                        coa = _load_coa_from_list(json.loads(fields["coa_ledgers"]))
                    except Exception:
                        coa = None
                    coa_parsed_list = []
                else:
                    coa = None
                    coa_parsed_list = []

                # Voucher Type resolution: uploaded file → DB JSON list → None
                if vt_file:
                    vt_master = _load_voucher_types(vt_file["content"], vt_file["filename"])
                    vt_parsed_list = list(vt_master.values())
                elif fields.get("vt_ledgers"):
                    try:
                        vt_master = _load_voucher_types_from_list(json.loads(fields["vt_ledgers"]))
                    except Exception:
                        vt_master = None
                    vt_parsed_list = []
                else:
                    vt_master = None
                    vt_parsed_list = []

                all_entries, monthly_data, state_summary = gstr3b_process_multi(
                    gstr3b_files, coa=coa, vt_master=vt_master
                )
                job_id = uuid4().hex
                payload = {
                    "job_id": job_id,
                    "reco_type": reco_type,
                    "monthly_data": monthly_data,
                    "state_summary": state_summary,
                    "coa_ledgers_parsed": coa_parsed_list,
                    "vt_ledgers_parsed": vt_parsed_list,
                    "summary": {
                        "months": len(monthly_data),
                        "gstin":  monthly_data[0]["gstin"]  if monthly_data else "",
                        "state":  monthly_data[0]["state"]  if monthly_data else "",
                        "period": monthly_data[0]["period"] if monthly_data else "",
                    },
                    "counts": {"entry_rows": len(all_entries), "months": len(monthly_data)},
                    "results": all_entries,
                    "_monthly_data":   monthly_data,
                    "_state_summary":  state_summary,
                }
                JOBS[job_id] = payload
                self.write_json({k: v for k, v in payload.items() if not k.startswith("_")})
                return

            if reco_type == "pdf_bank_extract":
                from recon.pdf_bank_extractor import extract_bank_statement, build_pdf_bank_excel
                pdf_file = files.get("bank_pdf")
                if not pdf_file:
                    self.write_json({"error": "Upload a bank statement PDF (field name: bank_pdf)."}, 400)
                    return
                content = pdf_file["content"] if isinstance(pdf_file, dict) else pdf_file[0]["content"]
                data = extract_bank_statement(content)
                excel_bytes = build_pdf_bank_excel(data)
                job_id = uuid4().hex
                payload = {
                    "job_id":            job_id,
                    "reco_type":         reco_type,
                    "bank_name":         data.get("bank_name", ""),
                    "account_no":        data.get("account_no", ""),
                    "account_name":      data.get("account_name", ""),
                    "period_from":       data.get("period_from", ""),
                    "period_to":         data.get("period_to", ""),
                    "transaction_count": data.get("transaction_count", 0),
                    "validation":        data.get("validation", {}),
                    "preview_rows":      data.get("preview_rows", []),
                    "summary":           {"total": data.get("transaction_count", 0)},
                    "counts":            {"transaction_rows": data.get("transaction_count", 0)},
                    "results":           [],  # not used — download is via export endpoint
                    "_xlsx_bytes":       excel_bytes,
                }
                JOBS[job_id] = payload
                self.write_json({k: v for k, v in payload.items() if not k.startswith("_")})
                return

            # Default two-file reconciliation (gst_2b_purchase)
            gstr2b_file = files.get("gstr2b")
            purchase_file = files.get("purchase")
            if not gstr2b_file or not purchase_file:
                self.write_json({"error": "Upload both GSTR-2B and Purchase Register files."}, 400)
                return

            gstr2b_records = read_upload(gstr2b_file["filename"], gstr2b_file["content"], "GSTR-2B")
            purchase_records = read_upload(purchase_file["filename"], purchase_file["content"], "Purchase Register")
            results = reconcile(gstr2b_records, purchase_records, tolerance=tolerance)

            job_id = uuid4().hex
            payload = {
                "job_id": job_id,
                "reco_type": reco_type,
                "summary": summarize(results),
                "counts": {
                    "gstr2b_records": len(gstr2b_records),
                    "purchase_records": len(purchase_records),
                    "result_rows": len(results),
                },
                "results": [result.as_dict() for result in results],
            }
            JOBS[job_id] = payload
            self.write_json(payload)
        except Exception as exc:
            self.write_json({"error": str(exc)}, 500)
        finally:
            _RECO_SEMAPHORE.release()

    def read_multipart(self) -> tuple[dict[str, str], dict[str, dict]]:
        content_type = self.headers.get("Content-Type", "")
        content_length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(content_length)
        logging.info("Multipart: Content-Type=%s, body length=%d", content_type, len(body))
        message = BytesParser(policy=default).parsebytes(
            f"Content-Type: {content_type}\r\nMIME-Version: 1.0\r\n\r\n".encode() + body
        )
        fields: dict[str, str] = {}
        files: dict[str, dict] = {}
        for part in message.iter_parts():
            disposition = part.get_content_disposition()
            name = part.get_param("name", header="content-disposition")
            filename = part.get_filename()
            content = part.get_payload(decode=True) or b""
            logging.info("  Part: disposition=%s, name=%s, filename=%s, size=%d", disposition, name, filename, len(content))
            if disposition != "form-data":
                continue
            if filename:
                entry = {"filename": filename, "content": content}
                if name in files:
                    existing = files[name]
                    if not isinstance(existing, list):
                        files[name] = [existing]
                    files[name].append(entry)
                else:
                    files[name] = entry
            else:
                fields[name] = content.decode("utf-8", errors="replace")
        def _log_file(v):
            if isinstance(v, list):
                return [{'fn': i['filename'], 'sz': len(i['content'])} for i in v]
            return {'fn': v['filename'], 'sz': len(v['content'])}
        logging.info("Parsed fields: %s, files: %s", list(fields.keys()), {k: _log_file(v) for k, v in files.items()})
        return fields, files

    def serve_static(self, relative_path: str) -> None:
        target = (STATIC / relative_path).resolve()
        if not str(target).startswith(str(STATIC.resolve())) or not target.exists():
            self.write_json({"error": "Not found"}, 404)
            return
        content_type = "text/plain"
        if target.suffix == ".html":
            content_type = "text/html; charset=utf-8"
        elif target.suffix == ".css":
            content_type = "text/css; charset=utf-8"
        elif target.suffix == ".js":
            content_type = "application/javascript; charset=utf-8"
        data = target.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def write_json(self, payload: dict, status: int = 200) -> None:
        data = json.dumps(payload, indent=2).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def export_job(self, job_id: str) -> None:
        payload = JOBS.get(job_id)
        if not payload:
            self.write_json({"error": "Job not found"}, 404)
            return
        # Use pre-built bytes when available (avoids timeout for large workbooks)
        if payload.get("_xlsx_bytes"):
            data = payload["_xlsx_bytes"]
        else:
            workbook = build_workbook(
                payload["results"],
                payload["summary"],
                payload["counts"],
                payload.get("reco_type", "gst_2b_purchase"),
                pivot=payload.get("pivot"),
                payload=payload,
            )
            from io import BytesIO
            buffer = BytesIO()
            workbook.save(buffer)
            data = buffer.getvalue()
        self.send_response(200)
        reco_type = payload.get("reco_type", "gst_2b_purchase")
        if reco_type == "bank_reco":
            filename_prefix = "bank_statement"
        elif reco_type == "gstr_1_vs_books":
            filename_prefix = "gstr1_vs_books"
        elif reco_type == "gstr_2b_books_multistate":
            filename_prefix = "2b_vs_books_multistate"
        elif reco_type == "gstr_3b_tally_entry":
            filename_prefix = "gstr3b_tally_entry"
        elif reco_type == "pdf_bank_extract":
            acct = payload.get("account_no", "")
            filename_prefix = f"bank_statement_{acct}" if acct else "bank_statement_pdf"
        else:
            filename_prefix = "reconciliation"
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="{filename_prefix}-{job_id}.xlsx"')
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, format: str, *args) -> None:
        print(f"{self.address_string()} - {format % args}")


def build_workbook(results: list[dict], summary: dict[str, int], counts: dict[str, int], reco_type: str = "gst_2b_purchase", pivot: list[dict] | None = None, payload: dict | None = None) -> Workbook:
    if reco_type == "gstr_2b_books":
        from recon.gstr_2b_books import build_gstr2b_books_workbook
        return build_gstr2b_books_workbook(results, payload=payload)
    if reco_type == "gstr_2b_books_multistate":
        # Use MatchResult objects (with suggested_action_3) if available in the job payload
        result_objs = (payload or {}).get("_results_obj") or results
        return build_gstr2b_books_multistate_workbook(result_objs, payload=payload)
    if reco_type == "gstr_2a_2b_books":
        return build_three_way_workbook(results, summary, counts)
    if reco_type == "gstr_3b_vs_2b":
        return build_3b_vs_2b_workbook(results, summary, counts, pivot or [])
    if reco_type == "bank_reco":
        return build_bank_reco_workbook(results, summary, counts)
    if reco_type == "gstr_1_vs_books":
        return build_gstr1_workbook(
            results, monthly_summary=[], summary=summary, counts=counts, payload=payload
        )
    if reco_type == "gstr_3b_tally_entry":
        monthly_data  = (payload or {}).get("_monthly_data",  [])
        state_summary = (payload or {}).get("_state_summary", [])
        return build_gstr3b_tally_workbook(monthly_data, results, state_summary=state_summary)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append(["GSTR-2B records", counts["gstr2b_records"]])
    summary_sheet.append(["Purchase records", counts["purchase_records"]])
    summary_sheet.append(["Result rows", counts["result_rows"]])
    summary_sheet.append([])
    summary_sheet.append(["Category", "Count"])
    for category, count in summary.items():
        summary_sheet.append([category, count])
    style_header(summary_sheet)

    categories = sorted({result["category"] for result in results})
    for category in categories:
        sheet = workbook.create_sheet(safe_sheet_title(category))
        write_result_sheet(sheet, [result for result in results if result["category"] == category])
    return workbook


def safe_sheet_title(title: str) -> str:
    cleaned = re.sub(r"[\[\]\*:/\\?]", "-", title).strip()
    return (cleaned or "Sheet")[:31]


def write_result_sheet(sheet, rows: list[dict]) -> None:
    headers = [
        "Category",
        "Confidence",
        "2B GSTIN",
        "PR GSTIN",
        "2B Doc No",
        "PR Doc No",
        "2B Date",
        "PR Date",
        "2B Taxable",
        "PR Taxable",
        "2B Tax",
        "PR Tax",
        "Mismatches",
        "Suggested Action",
        "Explanation",
    ]
    sheet.append(headers)
    for result in rows:
        gstr2b = result.get("gstr2b") or {}
        purchase = result.get("purchase") or {}
        sheet.append(
            [
                result["category"],
                result["confidence"],
                gstr2b.get("supplier_gstin", ""),
                purchase.get("supplier_gstin", ""),
                gstr2b.get("doc_no", ""),
                purchase.get("doc_no", ""),
                gstr2b.get("doc_date", ""),
                purchase.get("doc_date", ""),
                gstr2b.get("taxable_value", ""),
                purchase.get("taxable_value", ""),
                gstr2b.get("total_tax", ""),
                purchase.get("total_tax", ""),
                ", ".join(result.get("mismatch_fields", [])),
                result["suggested_action"],
                result["explanation"],
            ]
        )
    style_header(sheet)


def build_three_way_workbook(results: list[dict], summary: dict[str, int], counts: dict[str, int]) -> Workbook:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append(["GSTR-2A records", counts.get("gstr2a_records", 0)])
    summary_sheet.append(["GSTR-2B records", counts.get("gstr2b_records", 0)])
    summary_sheet.append(["Books records", counts.get("books_records", 0)])
    summary_sheet.append(["Result rows", counts.get("result_rows", 0)])
    summary_sheet.append([])
    summary_sheet.append(["Category", "Count", "Amount"])
    for category, count in summary.items():
        amount = sum(record_amount(result) for result in results if result["category"] == category)
        summary_sheet.append([category, count, amount])
    style_header(summary_sheet)

    line_sheet = workbook.create_sheet("Invoice-level Matching")
    write_three_way_sheet(line_sheet, results)

    pan_sheet = workbook.create_sheet("PAN Summary")
    pan_totals: dict[str, dict[str, float]] = {}
    for result in results:
        pan = first_value(result, "pan") or "Unknown"
        bucket = pan_totals.setdefault(pan, {"2A": 0.0, "2B": 0.0, "Books": 0.0})
        if result.get("gstr2a"):
            bucket["2A"] += result["gstr2a"].get("doc_value", 0) or 0
        if result.get("gstr2b"):
            bucket["2B"] += result["gstr2b"].get("doc_value", 0) or 0
        if result.get("books"):
            bucket["Books"] += result["books"].get("doc_value", 0) or 0
    pan_sheet.append(["PAN", "As per 2A", "As per 2B", "As per Books", "2A - Books", "2B - Books"])
    for pan, totals in sorted(pan_totals.items()):
        pan_sheet.append([pan, totals["2A"], totals["2B"], totals["Books"], totals["2A"] - totals["Books"], totals["2B"] - totals["Books"]])
    style_header(pan_sheet)

    for category in sorted({result["category"] for result in results}):
        sheet = workbook.create_sheet(safe_sheet_title(category))
        write_three_way_sheet(sheet, [result for result in results if result["category"] == category])
    return workbook


def write_three_way_sheet(sheet, rows: list[dict]) -> None:
    headers = [
        "Category",
        "Confidence",
        "PAN",
        "Supplier",
        "Doc No",
        "2A Value",
        "2B Value",
        "Books Value",
        "Difference",
        "2A Month",
        "2B Month",
        "Books Month",
        "Mismatches",
        "Suggested Action",
        "Explanation",
    ]
    sheet.append(headers)
    for result in rows:
        value_2a = (result.get("gstr2a") or {}).get("doc_value", 0) or 0
        value_2b = (result.get("gstr2b") or {}).get("doc_value", 0) or 0
        value_books = (result.get("books") or {}).get("doc_value", 0) or 0
        portal_value = value_2b if result.get("gstr2b") else value_2a
        sheet.append(
            [
                result["category"],
                result["confidence"],
                first_value(result, "pan"),
                first_value(result, "supplier_name"),
                first_value(result, "doc_no"),
                value_2a,
                value_2b,
                value_books,
                portal_value - value_books,
                (result.get("gstr2a") or {}).get("month", ""),
                (result.get("gstr2b") or {}).get("month", ""),
                (result.get("books") or {}).get("month", ""),
                ", ".join(result.get("mismatch_fields", [])),
                result["suggested_action"],
                result["explanation"],
            ]
        )
    style_header(sheet)


def first_value(result: dict, field: str) -> str:
    for key in ("gstr2a", "gstr2b", "books"):
        record = result.get(key) or {}
        if record.get(field):
            return record[field]
    return ""


def record_amount(result: dict) -> float:
    for key in ("books", "gstr2b", "gstr2a"):
        record = result.get(key)
        if record:
            return record.get("doc_value", 0) or 0
    return 0.0


def style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="123C69")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    for column_cells in sheet.columns:
        max_length = max((len(str(cell.value or "")) for cell in column_cells), default=8)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 48)


def build_3b_vs_2b_workbook(results: list[dict], summary: dict[str, int], counts: dict[str, int], pivot: list[dict]) -> Workbook:
    """Build the Excel workbook matching the accountant's expected format.

    Pro-accountant sign convention:
      B2B:  2B row = POSITIVE values, 3B mirror row = NEGATIVE (negate 2B)
      CDNR: 2B row = NEGATIVE values (credit notes reduce ITC), 3B mirror row = POSITIVE

    The 3B mirror row ALWAYS uses 2B values (negated) — NOT the raw 3B PR values.
    This ensures every matched pair nets to EXACTLY zero.

    Output ordering:
      1. Matched CDNRs (sorted by date)
      2. Not Claimed in 3B (CDNR entries)
      3. Matched B2B (sorted by date)
      4. Not Claimed in 3B (B2B entries)
      5. Not in 2B entries
    """
    workbook = Workbook()

    # --- Primary sheet: 2B vs 3B (matches accountant's expected output) ---
    ws = workbook.active
    ws.title = "2B vs 3B"
    headers = [
        "Month", "Date", "Invoice No", "Party Name",
        "Value", "CGST", "SGST", "IGST",
        "Source", "Remarks", "Remarks 2 (Detail Explanation)", "Taxable",
        "Effect in Books", "Effect in 9-9C",
    ]
    ws.append(headers)

    def _is_cdnr(r: dict) -> bool:
        """Check if a result originates from a CDNR source."""
        rec_2b = r.get("rec_2b")
        if rec_2b and "CDNR" in (rec_2b.get("source", "") or "").upper():
            return True
        return False

    def _sort_key(r: dict) -> tuple:
        """Sort key: CDNRs first, then B2B; within each group, sort by date."""
        is_cdnr = _is_cdnr(r)
        category = r.get("category", "")
        # Category order: Matched/Amount Mismatch first, then Not Claimed, then Not in 2B
        cat_order = 0 if category in ("Matched", "Amount Mismatch") else 1 if category == "Not Claimed in 3B" else 2
        rec = r.get("rec_2b") or r.get("rec_3b") or {}
        date_str = rec.get("date", "") or ""
        return (0 if is_cdnr else 1, cat_order, date_str)

    sorted_results = sorted(results, key=_sort_key)

    for r in sorted_results:
        category = r.get("category", "")
        rec_2b = r.get("rec_2b")
        rec_3b = r.get("rec_3b")
        is_cdnr = _is_cdnr(r)

        if category in ("Matched", "Amount Mismatch"):
            if rec_2b:
                # Get raw 2B values from portal (always positive in portal)
                value_2b = rec_2b.get("value", 0) or 0
                cgst_2b = rec_2b.get("cgst", 0) or 0
                sgst_2b = rec_2b.get("sgst", 0) or 0
                igst_2b = rec_2b.get("igst", 0) or 0
                taxable_2b = rec_2b.get("taxable", 0) or 0

                if is_cdnr:
                    # CDNR: Credit notes REDUCE ITC → 2B row is NEGATIVE
                    sign_2b = -1
                else:
                    # B2B: Regular invoices → 2B row is POSITIVE
                    sign_2b = 1

                # Row 1: 2B side
                ws.append([
                    rec_2b.get("month", ""),
                    rec_2b.get("date", ""),
                    rec_2b.get("invoice_no", ""),
                    rec_2b.get("party_name", ""),
                    sign_2b * abs(value_2b),
                    sign_2b * abs(cgst_2b) if cgst_2b else 0,
                    sign_2b * abs(sgst_2b) if sgst_2b else 0,
                    sign_2b * abs(igst_2b) if igst_2b else 0,
                    rec_2b.get("source", "2B"),
                    "Matched",
                    None,
                    sign_2b * abs(taxable_2b),
                    None,
                    None,
                ])

                # Row 2: 3B Working mirror row (NEGATE the 2B row → nets to zero)
                # Use 2B values with opposite sign, and 2B party name for consistency
                ws.append([
                    rec_2b.get("month", ""),
                    rec_2b.get("date", ""),
                    rec_2b.get("invoice_no", ""),
                    rec_2b.get("party_name", ""),
                    -sign_2b * abs(value_2b),
                    -sign_2b * abs(cgst_2b) if cgst_2b else 0,
                    -sign_2b * abs(sgst_2b) if sgst_2b else 0,
                    -sign_2b * abs(igst_2b) if igst_2b else 0,
                    "3B Working",
                    "Matched",
                    None,
                    -sign_2b * abs(taxable_2b),
                    None,
                    None,
                ])

        elif category == "Not in 2B":
            # Single row — 3B Working entry not in 2B
            if rec_3b:
                ws.append([
                    rec_3b.get("month", ""),
                    rec_3b.get("date", ""),
                    rec_3b.get("invoice_no", ""),
                    rec_3b.get("party_name", ""),
                    rec_3b.get("value", 0),
                    rec_3b.get("cgst", 0) or 0,
                    rec_3b.get("sgst", 0) or 0,
                    rec_3b.get("igst", 0) or 0,
                    "3B Working",
                    "Not in 2B",
                    rec_3b.get("remarks2", "") or "ITC Reversal",
                    rec_3b.get("taxable", 0),
                    "No effect",
                    "No effect",
                ])

        elif category == "Not Claimed in 3B":
            # Single row — 2B entry not claimed in 3B
            if rec_2b:
                value_2b = rec_2b.get("value", 0) or 0
                cgst_2b = rec_2b.get("cgst", 0) or 0
                sgst_2b = rec_2b.get("sgst", 0) or 0
                igst_2b = rec_2b.get("igst", 0) or 0
                taxable_2b = rec_2b.get("taxable", 0) or 0

                ws.append([
                    rec_2b.get("month", ""),
                    rec_2b.get("date", ""),
                    rec_2b.get("invoice_no", ""),
                    rec_2b.get("party_name", ""),
                    value_2b,
                    cgst_2b if cgst_2b else 0,
                    sgst_2b if sgst_2b else 0,
                    igst_2b if igst_2b else 0,
                    rec_2b.get("source", "2B"),
                    "Not claimed in 3B",
                    rec_2b.get("remarks2", ""),
                    taxable_2b,
                    "No effect",
                    "No effect",
                ])

    style_header(ws)

    # --- Summary sheet ---
    summary_ws = workbook.create_sheet("Summary")
    summary_ws.append(["GSTR-3B vs GSTR-2B Reconciliation"])
    summary_ws.append([])
    summary_ws.append(["Metric", "Value"])
    summary_ws.append(["GSTR-2B records", counts.get("gstr2b_records", 0)])
    summary_ws.append(["GSTR-3B Working records", counts.get("gstr3b_records", 0)])
    summary_ws.append(["Total result rows", counts.get("result_rows", 0)])
    summary_ws.append([])
    summary_ws.append(["Category", "Count"])
    for cat, cnt in summary.items():
        summary_ws.append([cat, cnt])
    style_header_row(summary_ws, 3)
    style_header_row(summary_ws, 8)

    # --- Month Pivot sheet ---
    pv = workbook.create_sheet("Month Pivot")
    pv.append([
        "Month",
        "2B IGST", "2B CGST", "2B SGST",
        "3B IGST", "3B CGST", "3B SGST",
        "IGST Diff", "CGST Diff", "SGST Diff",
        "2B Value", "3B Value", "Value Diff",
    ])
    for row in pivot:
        pv.append([
            row["month"],
            row["igst_2b"], row["cgst_2b"], row["sgst_2b"],
            row["igst_3b"], row["cgst_3b"], row["sgst_3b"],
            row["igst_diff"], row["cgst_diff"], row["sgst_diff"],
            row["value_2b"], row["value_3b"], row["value_diff"],
        ])
    style_header(pv)

    return workbook


def style_header_row(sheet, row_num: int) -> None:
    """Style a specific row as a header (used for multi-section summary sheets)."""
    fill = PatternFill("solid", fgColor="123C69")
    for cell in sheet[row_num]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill


def build_gstr1_workbook(
    results: list[dict],
    monthly_summary: list[dict],
    summary: dict[str, int],
    counts: dict[str, int],
    payload: dict | None = None,
) -> Workbook:
    """
    Build the GSTR-1 vs Books output workbook (9 sheets).
    Sheet order:
      1. GST Reco         — 4-section master summary
      2. GSTR-1 Pivot     — (only when PDF uploaded)
      3. B2B Reco         — invoice-level, all Tally columns
      4. B2C Reco         — state+rate aggregation
      5. Final GSTR-1     — raw GSTR-1 data
      6. GSTR2B           — raw GSTR-2B data
      7. GSTR3B           — raw GSTR-3B data
      8. Sales Register   — raw Tally data
      9. Credit Note      — (only when credit note uploaded)
    """
    p = payload or {}
    sections      = p.get("_gst_reco_sections") or {}
    b2b_rows      = p.get("_b2b_reco_rows") or []
    b2c_rows      = p.get("_b2c_reco_rows") or []
    pivot_rows    = p.get("_pivot_rows")
    tally_cols    = p.get("_tally_cols") or []
    raw_gstr1     = p.get("_raw_gstr1") or []
    raw_gstr2b    = p.get("_raw_gstr2b") or []
    raw_gstr3b    = p.get("_raw_gstr3b") or []
    raw_tally     = p.get("_raw_tally") or []
    raw_cn        = p.get("_raw_cn")

    wb = Workbook()

    # -----------------------------------------------------------------------
    # Sheet 1: GST Reco (master summary — 4 sections)
    # -----------------------------------------------------------------------
    ws = wb.active
    ws.title = "GST Reco"
    ws.sheet_view.showGridLines = False

    _section_configs = [
        ("Sales Reco",                "gstr1_vs_gstr3b",    "GSTR-1",        "GSTR-3B"),
        ("As per books (All sales)",  "books_all_vs_gstr1", "As per books",   "GSTR-1 (All sales)"),
        ("As per books B2B",          "books_b2b_vs_gstr1", "As per books B2B", "GSTR-1 B2B"),
        ("As per books B2C",          "books_b2c_vs_gstr1", "As per books B2C", "GSTR-1 B2C"),
    ]

    current_row = 1
    for section_title, section_key, left_label, right_label in _section_configs:
        rows = sections.get(section_key) or []
        current_row = _write_gst_reco_section(
            ws, current_row, section_title, left_label, right_label, rows, section_key
        )
        current_row += 3  # gap between sections

    _auto_col_width(ws, min_width=12, max_width=22)
    ws.freeze_panes = "B6"

    # -----------------------------------------------------------------------
    # Sheet 2: GSTR-1 Pivot (only when PDF data present)
    # -----------------------------------------------------------------------
    if pivot_rows:
        ws_pivot = wb.create_sheet("GSTR-1 Pivot")
        ws_pivot.sheet_view.showGridLines = False
        _write_pivot_sheet(ws_pivot, pivot_rows)

    # -----------------------------------------------------------------------
    # Sheet 3: B2B Reco
    # -----------------------------------------------------------------------
    ws_b2b = wb.create_sheet("B2B Reco")
    ws_b2b.sheet_view.showGridLines = False
    _write_b2b_reco_sheet(ws_b2b, b2b_rows, tally_cols)

    # -----------------------------------------------------------------------
    # Sheet 4: B2C Reco
    # -----------------------------------------------------------------------
    ws_b2c = wb.create_sheet("B2C Reco")
    ws_b2c.sheet_view.showGridLines = False
    _write_b2c_reco_sheet(ws_b2c, b2c_rows)

    # -----------------------------------------------------------------------
    # Sheets 5–9: Raw passthrough data
    # -----------------------------------------------------------------------
    _write_raw_sheet(wb, "Final GSTR-1", raw_gstr1)
    _write_raw_sheet(wb, "GSTR2B",       raw_gstr2b)
    _write_raw_sheet(wb, "GSTR3B",       raw_gstr3b)
    _write_raw_sheet(wb, "Sales Register", raw_tally)
    if raw_cn is not None:
        _write_raw_sheet(wb, "Credit Note", raw_cn)

    return wb


# ---------------------------------------------------------------------------
# GST Reco section writer
# ---------------------------------------------------------------------------

_NAVY      = "1F3864"
_MED_BLUE  = "2E75B6"
_LIGHT_BLUE = "EBF3FB"
_GREY_TOTAL = "D9D9D9"
_GREEN_FG  = "276221";  _GREEN_BG  = "C6EFCE"
_RED_FG    = "9C0006";  _RED_BG    = "FFC7CE"
_ORANGE_FG = "9C5700";  _ORANGE_BG = "FFEB9C"
_BLUE_FG   = "1F4E79";  _BLUE_BG   = "DDEBF7"

_NUM_FMT = '#,##0.00'
_THIN = Side(style="thin")
_MED  = Side(style="medium")


def _cell_style(cell, bold=False, bg=None, fg="000000", num_fmt=None,
               border=None, align="left", wrap=False):
    cell.font = Font(bold=bold, color=fg)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        cell.number_format = num_fmt
    if border:
        cell.border = border
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)


def _header_border():
    return Border(left=_THIN, right=_THIN, top=_MED, bottom=_MED)


def _data_border():
    return Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _write_gst_reco_section(ws, start_row, title, left_label, right_label, rows, section_key):
    """Write a 4-column-group section in the GST Reco sheet."""
    # Title row
    title_cell = ws.cell(row=start_row, column=1, value=title)
    _cell_style(title_cell, bold=True, bg=_NAVY, fg="FFFFFF", align="center")
    start_row += 2

    # Group headers row
    group_headers = [left_label, None, None, None, None, None, None, right_label, None, None, None, None, None, None, "Difference"]
    for ci, val in enumerate(group_headers, 1):
        cell = ws.cell(row=start_row, column=ci, value=val)
        if val:
            _cell_style(cell, bold=True, bg=_MED_BLUE, fg="FFFFFF", align="center")
    start_row += 1

    # Column headers
    COL_HEADERS = ["Month", "Taxable amount", "IGST", "CGST", "SGST", "Total", "",
                   "Month", "Taxable amount", "IGST", "CGST", "SGST", "Total", "",
                   "Month", "Taxable amount", "IGST", "CGST", "SGST", "Total"]
    for ci, h in enumerate(COL_HEADERS, 1):
        cell = ws.cell(row=start_row, column=ci, value=h if h else None)
        _cell_style(cell, bold=True, bg=_MED_BLUE, fg="FFFFFF",
                   border=_header_border(), align="center")
    start_row += 1

    # Map row keys based on section_key
    if "gstr3b" in section_key:
        left_pfx, right_pfx = "gstr1", "gstr3b"
    else:
        left_pfx, right_pfx = ("books", "gstr1") if rows and "books_taxable" in rows[0] else ("gstr1", "gstr3b")
    # Detect actual prefixes from first data row
    if rows:
        keys = list(rows[0].keys())
        left_pfx  = next((k.split("_")[0] for k in keys if "_taxable" in k), left_pfx)
        # second prefix
        matches = [k.split("_")[0] for k in keys if "_taxable" in k]
        right_pfx = matches[1] if len(matches) > 1 else right_pfx

    data_rows = [r for r in rows if r.get("month") != "Total"]
    total_row = next((r for r in rows if r.get("month") == "Total"), None)
    first_data_row = start_row

    for ri, row in enumerate(data_rows):
        is_alt = (ri % 2 == 1)
        bg = _LIGHT_BLUE if is_alt else None
        month = row.get("month", "")

        def g(key, pfx=None):
            for p in ([pfx] if pfx else [left_pfx, right_pfx, ""]):
                v = row.get(f"{p}_{key}" if p else key)
                if v is not None:
                    return v
            return 0

        def lv(k):   return row.get(f"{left_pfx}_{k}", 0) or 0
        def rv(k):   return row.get(f"{right_pfx}_{k}", 0) or 0
        def dv(k):   return row.get(f"diff_{k}", 0) or 0

        vals = [
            month,
            lv("taxable"), lv("igst"), lv("cgst"), lv("sgst"),
            lv("taxable") + lv("igst") + lv("cgst") + lv("sgst"),
            None,
            month,
            rv("taxable"), rv("igst"), rv("cgst"), rv("sgst"),
            rv("taxable") + rv("igst") + rv("cgst") + rv("sgst"),
            None,
            month,
            dv("taxable"), dv("igst"), dv("cgst"), dv("sgst"),
            dv("taxable") + dv("igst") + dv("cgst") + dv("sgst"),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=start_row, column=ci, value=v)
            is_num = isinstance(v, (int, float)) and v is not None
            is_diff = ci >= 15
            if is_diff and is_num and v < 0:
                _cell_style(cell, bg=bg, fg=_RED_FG, num_fmt=_NUM_FMT, border=_data_border())
            elif is_diff and is_num and v > 0:
                _cell_style(cell, bg=bg, fg="006100", num_fmt=_NUM_FMT, border=_data_border())
            elif is_num:
                _cell_style(cell, bg=bg, num_fmt=_NUM_FMT, border=_data_border(), align="right")
            else:
                _cell_style(cell, bg=bg, bold=(ci in (1, 8, 15)), border=_data_border())
        start_row += 1

    # Total row
    if total_row:
        def lv(k): return total_row.get(f"{left_pfx}_{k}", 0) or 0
        def rv(k): return total_row.get(f"{right_pfx}_{k}", 0) or 0
        def dv(k): return total_row.get(f"diff_{k}", 0) or 0
        tot_vals = [
            "Total",
            lv("taxable"), lv("igst"), lv("cgst"), lv("sgst"),
            lv("taxable") + lv("igst") + lv("cgst") + lv("sgst"),
            None, "Total",
            rv("taxable"), rv("igst"), rv("cgst"), rv("sgst"),
            rv("taxable") + rv("igst") + rv("cgst") + rv("sgst"),
            None, "Total",
            dv("taxable"), dv("igst"), dv("cgst"), dv("sgst"),
            dv("taxable") + dv("igst") + dv("cgst") + dv("sgst"),
        ]
        for ci, v in enumerate(tot_vals, 1):
            cell = ws.cell(row=start_row, column=ci, value=v)
            is_num = isinstance(v, (int, float)) and v is not None
            _cell_style(cell, bold=True, bg=_GREY_TOTAL, num_fmt=_NUM_FMT if is_num else None,
                       border=_header_border(), align="right" if is_num else "left")
        start_row += 1

    return start_row


# ---------------------------------------------------------------------------
# GSTR-1 Pivot sheet
# ---------------------------------------------------------------------------

def _write_pivot_sheet(ws, pivot_rows):
    ws.cell(row=1, column=1, value="GSTR-1 Pivot — Excel (OCTA) vs PDF Validation")
    _cell_style(ws.cell(row=1, column=1), bold=True, bg=_NAVY, fg="FFFFFF")

    headers_top = ["", "Excel (OCTA)", None, None, None, "", "PDF (GST Portal)", None, None, None, "", "Difference"]
    for ci, v in enumerate(headers_top, 1):
        cell = ws.cell(row=3, column=ci, value=v if v else None)
        if v:
            _cell_style(cell, bold=True, bg=_MED_BLUE, fg="FFFFFF", align="center")

    sub_headers = ["Month", "Taxable", "IGST", "CGST", "SGST", "",
                   "Taxable", "IGST", "CGST", "SGST", "",
                   "Taxable", "IGST", "CGST", "SGST"]
    for ci, h in enumerate(sub_headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        _cell_style(cell, bold=True, bg=_MED_BLUE, fg="FFFFFF", border=_header_border(), align="center")

    data_rows = [r for r in pivot_rows if r.get("month") != "Total"]
    total_row  = next((r for r in pivot_rows if r.get("month") == "Total"), None)

    for ri, row in enumerate(data_rows):
        is_alt = (ri % 2 == 1)
        bg = _LIGHT_BLUE if is_alt else None
        month = row.get("month", "")
        ev = lambda k: row.get(f"excel_{k}", 0) or 0
        pv = lambda k: row.get(f"pdf_{k}", 0) or 0
        dv = lambda k: row.get(f"diff_{k}", 0) or 0
        vals = [month, ev("taxable"), ev("igst"), ev("cgst"), ev("sgst"), None,
                pv("taxable"), pv("igst"), pv("cgst"), pv("sgst"), None,
                dv("taxable"), dv("igst"), dv("cgst"), dv("sgst")]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=5 + ri, column=ci, value=v)
            is_num = isinstance(v, (int, float)) and v is not None
            is_diff = ci >= 12
            if is_diff and is_num and abs(v) > 1:
                _cell_style(cell, bg=bg, fg=_RED_FG if v < 0 else "006100", num_fmt=_NUM_FMT, border=_data_border())
            elif is_num:
                _cell_style(cell, bg=bg, num_fmt=_NUM_FMT, border=_data_border(), align="right")
            else:
                _cell_style(cell, bg=bg, bold=(ci == 1), border=_data_border())

    if total_row:
        tr = 5 + len(data_rows)
        ev = lambda k: total_row.get(f"excel_{k}", 0) or 0
        pv = lambda k: total_row.get(f"pdf_{k}", 0) or 0
        dv = lambda k: total_row.get(f"diff_{k}", 0) or 0
        vals = ["Total", ev("taxable"), ev("igst"), ev("cgst"), ev("sgst"), None,
                pv("taxable"), pv("igst"), pv("cgst"), pv("sgst"), None,
                dv("taxable"), dv("igst"), dv("cgst"), dv("sgst")]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=tr, column=ci, value=v)
            is_num = isinstance(v, (int, float)) and v is not None
            _cell_style(cell, bold=True, bg=_GREY_TOTAL, num_fmt=_NUM_FMT if is_num else None,
                       border=_header_border(), align="right" if is_num else "left")

    _auto_col_width(ws, min_width=10, max_width=18)
    ws.freeze_panes = "B5"


# ---------------------------------------------------------------------------
# B2B Reco sheet
# ---------------------------------------------------------------------------

_REMARK_COLORS = {
    "Match":                    (_GREEN_FG,  _GREEN_BG),
    "Diff":                     (_ORANGE_FG, _ORANGE_BG),
    "Not in GSTR-1":            (_RED_FG,    _RED_BG),
    "Not in Books":             (_RED_FG,    _RED_BG),
    "Amazon Entry As per Tally":  (_BLUE_FG,  _BLUE_BG),
    "Amazon Entry as per GSTR-1": ("5C3317",  "F0E6D3"),
}


def _write_b2b_reco_sheet(ws, b2b_rows: list[dict], tally_cols: list[str]):
    ws.cell(row=1, column=1, value="Books VS GSTR-1 — B2B Reconciliation")
    _cell_style(ws.cell(row=1, column=1), bold=True, bg=_NAVY, fg="FFFFFF")

    # Dynamic columns = all Tally cols + separator + GSTR-1 cols + Diff + Remark
    tally_display = [c for c in tally_cols if not c.startswith("_")]
    g1_cols    = ["GSTR-1 Invoice No", "GSTR-1 GSTIN", "GSTR-1 Taxable", "GSTR-1 IGST", "GSTR-1 CGST", "GSTR-1 SGST"]
    diff_cols  = ["Diff Taxable", "Diff IGST", "Diff CGST", "Diff SGST"]
    remark_col = ["Remark"]
    all_headers = tally_display + [""] + g1_cols + [""] + diff_cols + [" "] + remark_col

    # Group header row (row 2)
    n_tally = len(tally_display)
    group_row = ws.cell(row=2, column=1, value="Sales Register")
    _cell_style(group_row, bold=True, bg=_MED_BLUE, fg="FFFFFF", align="center")
    g1_start = n_tally + 2
    ws.cell(row=2, column=g1_start, value="GSTR-1")
    _cell_style(ws.cell(row=2, column=g1_start), bold=True, bg=_MED_BLUE, fg="FFFFFF", align="center")
    diff_start = g1_start + len(g1_cols) + 1
    ws.cell(row=2, column=diff_start, value="Difference")
    _cell_style(ws.cell(row=2, column=diff_start), bold=True, bg=_MED_BLUE, fg="FFFFFF", align="center")
    remark_start = diff_start + len(diff_cols) + 1
    ws.cell(row=2, column=remark_start, value="Remark")
    _cell_style(ws.cell(row=2, column=remark_start), bold=True, bg=_NAVY, fg="FFFFFF", align="center")

    # Column headers (row 3)
    for ci, h in enumerate(all_headers, 1):
        cell = ws.cell(row=3, column=ci, value=h or None)
        if h and h.strip():
            _cell_style(cell, bold=True, bg=_MED_BLUE, fg="FFFFFF", border=_header_border(), align="center")

    # Map Tally row keys → internal column names
    _g1_key_map = {
        "GSTR-1 Invoice No": "_gstr1_inv_no",
        "GSTR-1 GSTIN":      "_gstr1_gstin",
        "GSTR-1 Taxable":    "_gstr1_taxable",
        "GSTR-1 IGST":       "_gstr1_igst",
        "GSTR-1 CGST":       "_gstr1_cgst",
        "GSTR-1 SGST":       "_gstr1_sgst",
        "Diff Taxable":      "_diff_taxable",
        "Diff IGST":         "_diff_igst",
        "Diff CGST":         "_diff_cgst",
        "Diff SGST":         "_diff_sgst",
    }

    for ri, row in enumerate(b2b_rows):
        xl_row = ri + 4
        is_alt = (ri % 2 == 1)
        bg = _LIGHT_BLUE if is_alt else None
        remark = row.get("_remark", "")

        ci = 1
        # Tally columns (pass-through)
        for col in tally_display:
            v = row.get(col)
            cell = ws.cell(row=xl_row, column=ci, value=_json_to_xl(v))
            is_num = isinstance(v, (int, float)) and v is not None
            _cell_style(cell, bg=bg, num_fmt=_NUM_FMT if is_num else None,
                       border=_data_border(), align="right" if is_num else "left")
            ci += 1

        ci += 1  # separator

        # GSTR-1 matched columns
        for h in g1_cols:
            key = _g1_key_map.get(h, "")
            v = row.get(key)
            cell = ws.cell(row=xl_row, column=ci, value=_json_to_xl(v))
            is_num = isinstance(v, (int, float)) and v is not None
            _cell_style(cell, bg=bg, num_fmt=_NUM_FMT if is_num else None,
                       border=_data_border(), align="right" if is_num else "left")
            ci += 1

        ci += 1  # separator

        # Diff columns
        for h in diff_cols:
            key = _g1_key_map.get(h, "")
            v = row.get(key)
            cell = ws.cell(row=xl_row, column=ci, value=_json_to_xl(v))
            is_num = isinstance(v, (int, float)) and v is not None
            if is_num and abs(v) > 0.5:
                fg = _RED_FG if v > 0 else "006100"
                _cell_style(cell, bg=bg, fg=fg, num_fmt=_NUM_FMT, border=_data_border(), align="right")
            else:
                _cell_style(cell, bg=bg, num_fmt=_NUM_FMT if is_num else None,
                           border=_data_border(), align="right" if is_num else "left")
            ci += 1

        ci += 1  # separator before remark

        # Remark column
        rfg, rbg = _REMARK_COLORS.get(remark, ("000000", None))
        remark_cell = ws.cell(row=xl_row, column=ci, value=remark)
        _cell_style(remark_cell, bold=True, bg=rbg, fg=rfg, border=_header_border(), align="center")

    _auto_col_width(ws, min_width=8, max_width=30)
    ws.freeze_panes = "A4"


def _json_to_xl(v):
    """Convert stored JSON value back to Excel-friendly value."""
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    return str(v)


# ---------------------------------------------------------------------------
# B2C Reco sheet
# ---------------------------------------------------------------------------

def _write_b2c_reco_sheet(ws, b2c_rows: list[dict]):
    ws.cell(row=1, column=1, value="B2C Reconciliation — State + Rate Annual Aggregation")
    _cell_style(ws.cell(row=1, column=1), bold=True, bg=_NAVY, fg="FFFFFF")

    # Group headers (row 2)
    group_map = {1: "GSTR-1", 8: "Books", 15: "Difference"}
    for ci, label in group_map.items():
        cell = ws.cell(row=2, column=ci, value=label)
        _cell_style(cell, bold=True, bg=_MED_BLUE, fg="FFFFFF", align="center")

    # Column headers (row 3)
    headers = [
        "States", "GST Rate", "Taxable Value", "IGST", "CGST", "SGST", "",
        "States", "GST Rate", "Taxable Value", "IGST", "CGST", "SGST", "",
        "Taxable Value", "IGST", "CGST", "SGST",
    ]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=ci, value=h or None)
        if h and h.strip():
            _cell_style(cell, bold=True, bg=_MED_BLUE, fg="FFFFFF", border=_header_border(), align="center")

    # Data rows
    for ri, row in enumerate(b2c_rows):
        xl_row = ri + 4
        is_alt = (ri % 2 == 1)
        bg = _LIGHT_BLUE if is_alt else None
        state = row.get("state", "")
        rate  = row.get("rate", 0)

        gv = lambda k: row.get(f"gstr1_{k}", 0) or 0
        bv = lambda k: row.get(f"books_{k}", 0) or 0
        dv = lambda k: row.get(f"diff_{k}", 0) or 0

        vals = [
            state, rate,
            gv("taxable"), gv("igst"), gv("cgst"), gv("sgst"), None,
            state, rate,
            bv("taxable"), bv("igst"), bv("cgst"), bv("sgst"), None,
            dv("taxable"), dv("igst"), dv("cgst"), dv("sgst"),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=xl_row, column=ci, value=v)
            is_num = isinstance(v, (int, float)) and v is not None
            is_diff = ci >= 15
            if is_diff and is_num and abs(v) > 0.5:
                fg = _RED_FG if v < 0 else "006100"
                _cell_style(cell, bg=bg, fg=fg, num_fmt=_NUM_FMT, border=_data_border(), align="right")
            elif is_num:
                _cell_style(cell, bg=bg, num_fmt=_NUM_FMT, border=_data_border(), align="right")
            else:
                _cell_style(cell, bg=bg, border=_data_border())

    # Totals row
    if b2c_rows:
        tr = len(b2c_rows) + 4
        sum_keys = ["gstr1_taxable", "gstr1_igst", "gstr1_cgst", "gstr1_sgst",
                    "books_taxable", "books_igst", "books_cgst", "books_sgst",
                    "diff_taxable", "diff_igst", "diff_cgst", "diff_sgst"]
        sums = {k: sum(row.get(k, 0) or 0 for row in b2c_rows) for k in sum_keys}
        tot_vals = [
            "Total", "",
            sums["gstr1_taxable"], sums["gstr1_igst"], sums["gstr1_cgst"], sums["gstr1_sgst"], None,
            "Total", "",
            sums["books_taxable"], sums["books_igst"], sums["books_cgst"], sums["books_sgst"], None,
            sums["diff_taxable"], sums["diff_igst"], sums["diff_cgst"], sums["diff_sgst"],
        ]
        for ci, v in enumerate(tot_vals, 1):
            cell = ws.cell(row=tr, column=ci, value=v)
            is_num = isinstance(v, (int, float)) and v is not None
            _cell_style(cell, bold=True, bg=_GREY_TOTAL, num_fmt=_NUM_FMT if is_num else None,
                       border=_header_border(), align="right" if is_num else "left")

    _auto_col_width(ws, min_width=10, max_width=22)
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Raw passthrough sheet writer
# ---------------------------------------------------------------------------

def _write_raw_sheet(wb: Workbook, title: str, records: list[dict]):
    if not records:
        ws = wb.create_sheet(title)
        ws.cell(row=1, column=1, value="(No data)")
        return
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    headers = list(records[0].keys())
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=str(h))
        _cell_style(cell, bold=True, bg=_NAVY, fg="FFFFFF", border=_header_border(), align="center")
    for ri, row in enumerate(records, 2):
        is_alt = (ri % 2 == 0)
        bg = _LIGHT_BLUE if is_alt else None
        for ci, h in enumerate(headers, 1):
            v = row.get(h)
            cell = ws.cell(row=ri, column=ci, value=_json_to_xl(v))
            is_num = isinstance(v, (int, float)) and v is not None
            _cell_style(cell, bg=bg, num_fmt=_NUM_FMT if is_num else None,
                       border=_data_border(), align="right" if is_num else "left")
    _auto_col_width(ws, min_width=8, max_width=30)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Auto column width helper
# ---------------------------------------------------------------------------

def _auto_col_width(ws, min_width=10, max_width=40):
    from openpyxl.utils import get_column_letter
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def main() -> None:
    query = parse_qs(urlparse("?" + "port=8765").query)
    port = int(query.get("port", ["8765"])[0])
    server = ThreadingHTTPServer(("0.0.0.0", port), ReconciliationHandler)
    print(f"CA Reconciliation Tool running at http://127.0.0.1:{port}")
    server.serve_forever()


def build_bank_reco_workbook(results: list[dict], summary: dict[str, int], counts: dict[str, int]) -> Workbook:
    workbook = Workbook()
    
    ws = workbook.active
    ws.title = "OD acc Working"
    headers = ["Txn Date", "Description", "Debit", "Credit", "Balance", "Type ", "Ledger name"]
    ws.append(headers)
    
    for r in results:
        ws.append([
            r.get("txn_date", ""),
            r.get("original_description", ""),
            r.get("debit") if r.get("debit") else None,
            r.get("credit") if r.get("credit") else None,
            r.get("balance", 0),
            r.get("predicted_type", ""),
            r.get("predicted_ledger", "")
        ])
    style_header(ws)
    
    summary_ws = workbook.create_sheet("Summary")
    summary_ws.append(["Bank Statement Classification"])
    summary_ws.append([])
    summary_ws.append(["Metric", "Value"])
    summary_ws.append(["Total Ledgers", counts.get("master_ledgers", 0)])
    summary_ws.append(["Total Bank Rows", counts.get("bank_rows", 0)])
    summary_ws.append([])
    summary_ws.append(["Confidence Level", "Count"])
    for cat, cnt in summary.items():
        summary_ws.append([cat, cnt])
    style_header_row(summary_ws, 3)
    style_header_row(summary_ws, 7)
    
    return workbook

if __name__ == "__main__":
    main()
