from __future__ import annotations

import logging
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from io import BytesIO, StringIO
import re
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook

from .core import json_safe, normalize_doc_no, normalize_doc_type, normalize_gstin, parse_date, round_money
from .parsers import normalize_header

logger = logging.getLogger(__name__)


PORTAL_2A_SHEET_HINTS = ("2a as per gst", "gstr2a", "gstr-2a", "2a")
PORTAL_2B_SHEET_HINTS = ("2b", "gstr2b", "gstr-2b")
BOOKS_SHEET_HINTS = ("working", "books", "purchase", "sheet1")

FIELD_ALIASES = {
    "supplier_gstin": [
        "supplier gstin", "gstin/uin", "gstin uin", "gstin of supplier",
        "vendor gstin", "party gstin", "gstin", "ctin", "supplier gst",
        "gstin of isd", "gstin of ecommerce",
    ],
    "pan": ["pan", "vendor pan", "supplier pan"],
    "supplier_name": [
        "supplier name", "particulars", "party name", "vendor name",
        "trade name", "legal name", "trade/legal name",
    ],
    "doc_type": [
        "doc type", "document type", "voucher type", "invoice type", "type",
        "isd document type",
    ],
    "doc_no": [
        "voucher ref no as per gst",
        "supplier invoice no",
        "supplier invoice no.",
        "doc no",
        "document number",
        "invoice number",
        "invoice no",
        "invoice no.",
        "voucher ref no",
        "voucher ref no.",
        "voucher no",
        "voucher no.",
        "bill no",
        "bill no.",
        "inum",
        "no",
        "isd document number",
        "note number",
        "nt num",
    ],
    "doc_date": [
        "doc date", "document date", "invoice date", "date", "voucher date",
        "bill date", "idt", "dt", "isd document date", "note date",
    ],
    "month": [
        "month", "tax period", "gstr-1 filing period", "gstr1 filing period",
        "return period", "filing period", "isd gstr-6 period", "isd gstr6 period",
    ],
    "doc_value": [
        "invoice value", "doc value", "document value", "gross total",
        "total value", "as per 2a", "as per 2b", "total invoice value", "val",
    ],
    "taxable_value": [
        "item taxable value", "taxable value", "taxable amount",
        "value", "txval", "assessable value",
    ],
    "igst": ["igst", "integrated tax", "integrated tax amount", "iamt"],
    "cgst": ["cgst", "central tax", "central tax amount", "camt"],
    "sgst": ["sgst", "utgst", "state tax", "state/ut tax", "state tax amount", "samt"],
    "cess": ["cess", "cess amount", "csamt"],
    "remarks": ["remarks", "remark", "remarks by manisha"],
    "unique": ["unique no", "unique value", "unique number", "uid"],
}


@dataclass
class ThreeWayRecord:
    source: str
    row_id: str
    supplier_gstin: str = ""
    pan: str = ""
    supplier_name: str = ""
    doc_type: str = "INV"
    doc_no: str = ""
    normalized_doc_no: str = ""
    doc_date: str = ""
    month: str = ""
    doc_value: float = 0.0
    taxable_value: float = 0.0
    igst: float = 0.0
    cgst: float = 0.0
    sgst: float = 0.0
    cess: float = 0.0
    uid: str = ""
    remarks: str = ""
    raw: dict[str, Any] = field(default_factory=dict)

    @property
    def total_tax(self) -> float:
        return round_money(self.igst + self.cgst + self.sgst + self.cess)

    def as_dict(self) -> dict[str, Any]:
        result = json_safe(asdict(self))
        result["total_tax"] = self.total_tax
        return result


@dataclass
class ThreeWayResult:
    category: str
    confidence: int
    gstr2a: ThreeWayRecord | None
    gstr2b: ThreeWayRecord | None
    books: ThreeWayRecord | None
    mismatch_fields: list[str]
    suggested_action: str
    explanation: str

    def as_dict(self) -> dict[str, Any]:
        return {
            "category": self.category,
            "confidence": self.confidence,
            "gstr2a": self.gstr2a.as_dict() if self.gstr2a else None,
            "gstr2b": self.gstr2b.as_dict() if self.gstr2b else None,
            "books": self.books.as_dict() if self.books else None,
            "mismatch_fields": self.mismatch_fields,
            "suggested_action": self.suggested_action,
            "explanation": self.explanation,
        }


def read_three_way_uploads(files: dict[str, dict]) -> tuple[list[ThreeWayRecord], list[ThreeWayRecord], list[ThreeWayRecord]]:
    logger.info("=== read_three_way_uploads ===")
    logger.info("Files received: %s", {k: {'filename': v.get('filename'), 'size': len(v.get('content', b''))} for k, v in files.items()})
    combined = files.get("combined_workbook")
    if combined and combined.get("content"):
        # Validate that the combined workbook actually has multiple relevant sheets
        # (2A, 2B, and/or Books). If it only has one generic sheet like 'Sheet1',
        # treat it as a Books file instead and fall through to separate processing.
        if is_truly_combined_workbook(combined["filename"], combined["content"]):
            logger.info("Using combined workbook: %s (%d bytes)", combined['filename'], len(combined['content']))
            return read_combined_workbook(combined["filename"], combined["content"])
        else:
            logger.info("Combined workbook slot has a single-purpose file, treating as Books: %s", combined['filename'])
            # Move it to be treated as Books if no explicit books file
            if not files.get("books") or not files["books"].get("content"):
                files["books"] = combined

    gstr2a = read_records_from_upload(files.get("gstr2a"), "GSTR-2A")
    gstr2b = read_records_from_upload(files.get("gstr2b_multi"), "GSTR-2B")
    books = read_records_from_upload(files.get("books"), "Books")
    logger.info("Separate files: 2A=%d records, 2B=%d records, Books=%d records", len(gstr2a), len(gstr2b), len(books))
    if not gstr2a and not gstr2b and not books:
        raise ValueError("Upload a combined workbook or separate GSTR-2A, GSTR-2B, and Books files.")
    if not books:
        raise ValueError("Books data is required for GSTR-2A vs 2B vs Books reconciliation.")
    return gstr2a, gstr2b, books


def is_truly_combined_workbook(filename: str, data: bytes) -> bool:
    """Check if the uploaded file actually has multiple relevant sheet types (2A, 2B, Books).
    If it only has generic sheets, it's likely a single-purpose file."""
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if suffix not in {"xlsx", "xlsm"}:
        return False
    try:
        workbook = load_workbook(BytesIO(data), data_only=True, read_only=True)
        sheet_names = [normalize_header(s.title) for s in workbook.worksheets]
        logger.info("  Checking if truly combined: sheets=%s", sheet_names)
        has_2a = any(matches_hint(name, PORTAL_2A_SHEET_HINTS) and "vs" not in name for name in sheet_names)
        has_2b = any(matches_hint(name, PORTAL_2B_SHEET_HINTS) and "vs" not in name for name in sheet_names)
        has_books = any(matches_hint(name, BOOKS_SHEET_HINTS) and "pivot" not in name and "remark" not in name for name in sheet_names)
        # It's truly combined if it has at least 2 different categories of sheets
        categories = sum([has_2a, has_2b, has_books])
        logger.info("  Categories found: 2A=%s, 2B=%s, Books=%s (total=%d)", has_2a, has_2b, has_books, categories)
        return categories >= 2
    except Exception as exc:
        logger.warning("  Failed to inspect combined workbook: %s", exc)
        return False


def read_combined_workbook(filename: str, data: bytes) -> tuple[list[ThreeWayRecord], list[ThreeWayRecord], list[ThreeWayRecord]]:
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if suffix not in {"xlsx", "xlsm"}:
        raise ValueError("Combined workbook must be an Excel .xlsx or .xlsm file.")

    workbook = load_workbook(BytesIO(data), data_only=True, read_only=True)
    logger.info("Combined workbook sheets: %s", [s.title for s in workbook.worksheets])
    gstr2a: list[ThreeWayRecord] = []
    gstr2b: list[ThreeWayRecord] = []
    books_by_sheet: dict[str, list[ThreeWayRecord]] = {}
    for sheet in workbook.worksheets:
        name = normalize_header(sheet.title)
        logger.info("  Sheet '%s' -> normalized: '%s'", sheet.title, name)
        is_2a = matches_hint(name, PORTAL_2A_SHEET_HINTS) and "vs" not in name
        is_2b = matches_hint(name, PORTAL_2B_SHEET_HINTS) and "vs" not in name
        is_books = matches_hint(name, BOOKS_SHEET_HINTS) and "pivot" not in name and "remark" not in name
        logger.info("    matches: 2A=%s, 2B=%s, Books=%s", is_2a, is_2b, is_books)
        if is_2a:
            records = records_from_sheet(sheet, "GSTR-2A")
            logger.info("    -> %d GSTR-2A records from sheet '%s'", len(records), sheet.title)
            gstr2a.extend(records)
        elif is_2b:
            records = records_from_sheet(sheet, "GSTR-2B")
            logger.info("    -> %d GSTR-2B records from sheet '%s'", len(records), sheet.title)
            gstr2b.extend(records)
        elif is_books:
            records = records_from_sheet(sheet, "Books")
            logger.info("    -> %d Books records from sheet '%s'", len(records), sheet.title)
            books_by_sheet[name] = records
        else:
            logger.info("    -> SKIPPED (no match for any category)")
    books = select_books_records(books_by_sheet)
    logger.info("Combined totals: 2A=%d, 2B=%d, Books=%d", len(gstr2a), len(gstr2b), len(books))
    return dedupe_blank_records(gstr2a), dedupe_blank_records(gstr2b), dedupe_blank_records(books)


def read_records_from_upload(file_info: dict | None, source: str) -> list[ThreeWayRecord]:
    if not file_info or not file_info.get("content"):
        logger.info("No file provided for %s", source)
        return []
    filename = file_info["filename"]
    data = file_info["content"]
    logger.info("Reading separate upload for %s: %s (%d bytes)", source, filename, len(data))
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if suffix in {"xlsx", "xlsm"}:
        workbook = load_workbook(BytesIO(data), data_only=True, read_only=True)
        logger.info("  Sheets in %s: %s", filename, [s.title for s in workbook.worksheets])
        records: list[ThreeWayRecord] = []
        for sheet in workbook.worksheets:
            sheet_records = records_from_sheet(sheet, source)
            logger.info("  Sheet '%s' -> %d records for %s", sheet.title, len(sheet_records), source)
            records.extend(sheet_records)
        final = dedupe_blank_records(records)
        logger.info("  Total after dedupe for %s: %d", source, len(final))
        return final
    if suffix == "csv":
        frame = pd.read_csv(StringIO(data.decode("utf-8-sig", errors="replace"))).fillna("")
        return normalize_rows(frame.to_dict(orient="records"), source)
    raise ValueError(f"Unsupported file type for {source}: {filename}")


def matches_hint(sheet_name: str, hints: Iterable[str]) -> bool:
    return any(hint in sheet_name for hint in hints)


def records_from_sheet(sheet, source: str) -> list[ThreeWayRecord]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        logger.info("    Sheet '%s': empty (0 rows)", getattr(sheet, 'title', '?'))
        return []
    logger.info("    Sheet '%s': %d raw rows", getattr(sheet, 'title', '?'), len(rows))
    # Log first few rows for debugging - show ALL columns
    for i, row in enumerate(rows[:5]):
        logger.info("      Row %d: %s", i, [str(v)[:50] if v is not None else None for v in row])
    header_index = detect_header_row(rows)
    if header_index is None:
        logger.warning("    Sheet '%s': NO HEADER ROW DETECTED (score < 3)", getattr(sheet, 'title', '?'))
        return []
    headers = [str(value or "") for value in rows[header_index]]
    logger.info("    Sheet '%s': header at row %d -> ALL columns: %s", getattr(sheet, 'title', '?'), header_index, headers)
    records: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        if not any(value is not None and str(value).strip() for value in row):
            continue
        record = {headers[index]: row[index] if index < len(row) else "" for index in range(len(headers))}
        records.append(record)
    logger.info("    Sheet '%s': %d data rows -> normalizing for %s", getattr(sheet, 'title', '?'), len(records), source)
    result = normalize_rows(records, source)
    logger.info("    Sheet '%s': %d normalized records", getattr(sheet, 'title', '?'), len(result))
    return result


def detect_header_row(rows: list[tuple[Any, ...]]) -> int | None:
    best_index = None
    best_score = 0
    all_aliases = {alias for aliases in FIELD_ALIASES.values() for alias in aliases}
    for index, row in enumerate(rows[:25]):
        normalized = [normalize_header(value) for value in row]
        score = sum(1 for value in normalized if value in all_aliases)
        if "doc no" in normalized or "voucher ref no as per gst" in normalized:
            score += 2
        if "supplier gstin" in normalized or "gstin/uin" in normalized:
            score += 2
        matched_aliases = [v for v in normalized if v in all_aliases]
        if score > 0:
            logger.info("      Header candidate row %d: score=%d, matched=%s", index, score, matched_aliases)
        if score > best_score:
            best_score = score
            best_index = index
    logger.info("      Best header: row %s with score %d (threshold=3)", best_index, best_score)
    return best_index if best_score >= 3 else None


def normalize_rows(rows: list[dict[str, Any]], source: str) -> list[ThreeWayRecord]:
    output: list[ThreeWayRecord] = []
    skipped_no_identity = 0
    skipped_no_uid = 0
    for index, row in enumerate(rows, start=1):
        getter = build_getter(row)
        supplier_gstin = normalize_gstin(getter("supplier_gstin"))
        pan = normalize_pan(getter("pan")) or pan_from_gstin(supplier_gstin)
        supplier_name = str(getter("supplier_name") or "").strip()
        doc_no = clean_doc_display(getter("doc_no"))
        normalized_doc_no = normalize_doc_no(doc_no)
        if not normalized_doc_no and getter("unique"):
            normalized_doc_no = doc_from_uid(getter("unique"), pan)
            doc_no = normalized_doc_no
        # Allow records with supplier_name + doc_no even without GSTIN/PAN
        # (common in Tally Purchase Registers)
        if not pan and not supplier_gstin and not normalized_doc_no and not supplier_name:
            skipped_no_identity += 1
            continue
        doc_value = infer_doc_value(getter)
        uid = make_uid(pan, supplier_gstin, normalized_doc_no)
        # Fallback: use normalized supplier_name as identity for UID if no GSTIN/PAN
        if not uid and supplier_name and normalized_doc_no:
            name_key = normalize_doc_no(supplier_name)  # strip to alphanumeric
            uid = f"{name_key}{normalized_doc_no}"
        record = ThreeWayRecord(
            source=source,
            row_id=f"{source}-{index}",
            supplier_gstin=supplier_gstin,
            pan=pan,
            supplier_name=supplier_name,
            doc_type=normalize_doc_type(getter("doc_type")),
            doc_no=doc_no,
            normalized_doc_no=normalized_doc_no,
            doc_date=parse_date(getter("doc_date")),
            month=format_month(getter("month"), getter("doc_date")),
            doc_value=doc_value,
            taxable_value=round_money(getter("taxable_value")),
            igst=round_money(getter("igst")),
            cgst=round_money(getter("cgst")),
            sgst=round_money(getter("sgst")),
            cess=round_money(getter("cess")),
            uid=uid,
            remarks=str(getter("remarks") or "").strip(),
            raw={str(key): value for key, value in row.items() if str(key).strip()},
        )
        if record.uid:
            output.append(record)
        else:
            skipped_no_uid += 1
    if skipped_no_identity or skipped_no_uid:
        logger.info("    normalize_rows(%s): %d records created, %d skipped (no identity), %d skipped (no uid)",
                    source, len(output), skipped_no_identity, skipped_no_uid)
    return output


def build_getter(row: dict[str, Any]):
    normalized_map = {normalize_header(key): value for key, value in row.items()}

    def get(field_name: str) -> Any:
        for alias in FIELD_ALIASES[field_name]:
            alias_key = normalize_header(alias)
            if alias_key in normalized_map and normalized_map[alias_key] not in (None, ""):
                return normalized_map[alias_key]
        return ""

    return get


def clean_doc_display(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_pan(value: Any) -> str:
    text = re.sub(r"[^0-9A-Z]", "", str(value or "").upper())
    return text if len(text) == 10 else ""


def pan_from_gstin(gstin: str) -> str:
    return gstin[2:12] if len(gstin) >= 12 else ""


def make_uid(pan: str, gstin: str, normalized_doc_no: str) -> str:
    identity = pan or pan_from_gstin(gstin) or gstin
    if not identity or not normalized_doc_no:
        return ""
    return f"{identity}{normalized_doc_no}"


def doc_from_uid(uid: Any, pan: str) -> str:
    text = normalize_doc_no(uid)
    pan_text = normalize_pan(pan)
    if pan_text and text.startswith(pan_text):
        return text[len(pan_text) :]
    return text


def infer_doc_value(getter) -> float:
    value = round_money(getter("doc_value"))
    if value:
        return value
    return round_money(getter("taxable_value")) + round_money(getter("igst")) + round_money(getter("cgst")) + round_money(getter("sgst")) + round_money(getter("cess"))


def format_month(month_value: Any, date_value: Any) -> str:
    if month_value not in (None, ""):
        parsed = parse_date(month_value)
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", parsed):
            return datetime.fromisoformat(parsed).strftime("%b-%Y")
        return str(month_value).strip()
    parsed_date = parse_date(date_value)
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", parsed_date):
        return datetime.fromisoformat(parsed_date).strftime("%b-%Y")
    return ""


def dedupe_blank_records(records: list[ThreeWayRecord]) -> list[ThreeWayRecord]:
    return [record for record in records if record.uid and (record.doc_value or record.taxable_value or record.total_tax or record.doc_no)]


def select_books_records(books_by_sheet: dict[str, list[ThreeWayRecord]]) -> list[ThreeWayRecord]:
    if "working" in books_by_sheet:
        return books_by_sheet["working"]
    for name, records in books_by_sheet.items():
        if "working" in name:
            return records
    if books_by_sheet:
        return max(books_by_sheet.values(), key=len)
    return []


def reconcile_three_way(
    gstr2a_records: Iterable[ThreeWayRecord],
    gstr2b_records: Iterable[ThreeWayRecord],
    books_records: Iterable[ThreeWayRecord],
    tolerance: float = 1.0,
) -> list[ThreeWayResult]:
    list_2a = list(gstr2a_records)
    list_2b = list(gstr2b_records)
    list_books = list(books_records)

    by_2a = bucket(list_2a)
    by_2b = bucket(list_2b)
    by_books = bucket(list_books)

    results: list[ThreeWayResult] = []
    matched_2a_uids: set[str] = set()
    matched_2b_uids: set[str] = set()
    matched_books_uids: set[str] = set()

    # --- Pass 1: UID-based matching (works when all sources have GSTIN/PAN) ---
    all_uids = sorted(set(by_2a) | set(by_2b) | set(by_books))
    for uid in all_uids:
        has_2a = uid in by_2a
        has_2b = uid in by_2b
        has_books = uid in by_books

        # Only process UIDs that appear in at least 2 sources, or in GST sources
        # (we'll handle books-only and gst-only in pass 2 cross-matching)
        sources_count = sum([has_2a, has_2b, has_books])
        if sources_count < 2 and has_books and not has_2a and not has_2b:
            continue  # books-only — defer to cross-matching
        if sources_count < 2 and (has_2a or has_2b) and not has_books:
            # GST-only record — check if it might match a books record in pass 2
            # For now, mark it and let pass 2 handle it
            continue

        duplicate_uid = any(len(records) > 1 for records in (by_2a.get(uid, []), by_2b.get(uid, []), by_books.get(uid, [])))
        rec_2a = aggregate_records(by_2a.get(uid, []))
        rec_2b = aggregate_records(by_2b.get(uid, []))
        rec_books = aggregate_records(by_books.get(uid, []))

        result = classify_three_way(rec_2a, rec_2b, rec_books, tolerance, duplicate_uid)
        results.append(result)

        if has_2a: matched_2a_uids.add(uid)
        if has_2b: matched_2b_uids.add(uid)
        if has_books: matched_books_uids.add(uid)

    # --- Pass 2: Doc-no-based cross-matching for unmatched records ---
    # This handles the common case where Books records have no GSTIN (Tally exports)
    # and GST records have GSTIN-based UIDs — match by normalized_doc_no instead.
    unmatched_books = [r for uid_recs in by_books.values() for r in uid_recs
                       if r.uid not in matched_books_uids]
    unmatched_2a = [r for uid_recs in by_2a.values() for r in uid_recs
                    if r.uid not in matched_2a_uids]
    unmatched_2b = [r for uid_recs in by_2b.values() for r in uid_recs
                    if r.uid not in matched_2b_uids]

    logger.info("Cross-matching pass: %d unmatched books, %d unmatched 2A, %d unmatched 2B",
                len(unmatched_books), len(unmatched_2a), len(unmatched_2b))

    # Build doc_no indexes for unmatched GST records
    gst_2a_by_docno: dict[str, list[ThreeWayRecord]] = defaultdict(list)
    gst_2b_by_docno: dict[str, list[ThreeWayRecord]] = defaultdict(list)
    for r in unmatched_2a:
        if r.normalized_doc_no:
            gst_2a_by_docno[r.normalized_doc_no].append(r)
    for r in unmatched_2b:
        if r.normalized_doc_no:
            gst_2b_by_docno[r.normalized_doc_no].append(r)

    # Also build a doc_no index for unmatched books
    books_by_docno: dict[str, list[ThreeWayRecord]] = defaultdict(list)
    for r in unmatched_books:
        if r.normalized_doc_no:
            books_by_docno[r.normalized_doc_no].append(r)

    # Track which records we've cross-matched
    cross_matched_2a: set[str] = set()
    cross_matched_2b: set[str] = set()
    cross_matched_books: set[str] = set()

    # Match unmatched books records against unmatched GST records by doc_no
    for doc_no in sorted(set(books_by_docno) | set(gst_2a_by_docno) | set(gst_2b_by_docno)):
        books_recs = books_by_docno.get(doc_no, [])
        recs_2a = gst_2a_by_docno.get(doc_no, [])
        recs_2b = gst_2b_by_docno.get(doc_no, [])

        # Skip if already cross-matched
        books_recs = [r for r in books_recs if r.row_id not in cross_matched_books]
        recs_2a = [r for r in recs_2a if r.row_id not in cross_matched_2a]
        recs_2b = [r for r in recs_2b if r.row_id not in cross_matched_2b]

        if not books_recs and not recs_2a and not recs_2b:
            continue

        rec_books = aggregate_records(books_recs) if books_recs else None
        rec_2a = aggregate_records(recs_2a) if recs_2a else None
        rec_2b = aggregate_records(recs_2b) if recs_2b else None

        # Need at least something
        if not rec_books and not rec_2a and not rec_2b:
            continue

        duplicate = any(len(recs) > 1 for recs in (books_recs, recs_2a, recs_2b))
        result = classify_three_way(rec_2a, rec_2b, rec_books, tolerance, duplicate)

        # If this is a cross-match (books + GST found via doc_no), note it
        if rec_books and (rec_2a or rec_2b):
            if result.confidence == 0:
                # Upgrade from "not found" to a cross-match
                result = ThreeWayResult(
                    category="Cross-Matched",
                    confidence=70,
                    gstr2a=rec_2a,
                    gstr2b=rec_2b,
                    books=rec_books,
                    mismatch_fields=["matched_by_doc_no"],
                    suggested_action="Matched by invoice number. Verify vendor GSTIN mapping in books.",
                    explanation="Books record matched with GST data by invoice number (GSTIN was missing in books).",
                )

        results.append(result)

        for r in books_recs:
            cross_matched_books.add(r.row_id)
        for r in recs_2a:
            cross_matched_2a.add(r.row_id)
        for r in recs_2b:
            cross_matched_2b.add(r.row_id)

    # --- Pass 3: Remaining unmatched records ---
    for r in unmatched_books:
        if r.row_id not in cross_matched_books:
            results.append(ThreeWayResult(
                category="Not in GSTR 2A / 2B",
                confidence=0,
                gstr2a=None,
                gstr2b=None,
                books=r,
                mismatch_fields=["missing_in_2a_2b"],
                suggested_action="Hold ITC or follow up with vendor unless this is RCM, timing, or an approved adjustment.",
                explanation="The invoice is in books, but it was not found in either supplier-reported GST source.",
            ))

    for r in unmatched_2a:
        if r.row_id not in cross_matched_2a:
            # Check if this 2A record has a matching 2B record already handled
            results.append(ThreeWayResult(
                category="Not in Books",
                confidence=0,
                gstr2a=r,
                gstr2b=None,
                books=None,
                mismatch_fields=["missing_in_books"],
                suggested_action="Check whether booking is missed, not our bill, cross-GSTIN, or ineligible.",
                explanation="The invoice is visible in GSTR-2A, but no matching books entry was found.",
            ))

    for r in unmatched_2b:
        if r.row_id not in cross_matched_2b:
            results.append(ThreeWayResult(
                category="Not in Books",
                confidence=0,
                gstr2a=None,
                gstr2b=r,
                books=None,
                mismatch_fields=["missing_in_books"],
                suggested_action="Check whether booking is missed, not our bill, cross-GSTIN, or ineligible.",
                explanation="The invoice is visible in GSTR-2B, but no matching books entry was found.",
            ))

    logger.info("Reconciliation complete: %d results total", len(results))
    return results


def classify_three_way(
    rec_2a: ThreeWayRecord | None,
    rec_2b: ThreeWayRecord | None,
    rec_books: ThreeWayRecord | None,
    tolerance: float,
    duplicate_uid: bool,
) -> ThreeWayResult:
    """Classify a single three-way match into the appropriate category."""
    mismatch_fields: list[str] = []
    values = [record.doc_value for record in (rec_2a, rec_2b, rec_books) if record]

    if rec_books and not rec_2a and not rec_2b:
        category = "Not in GSTR 2A / 2B"
        confidence = 0
        action = "Hold ITC or follow up with vendor unless this is RCM, timing, or an approved adjustment."
        explanation = "The invoice is in books, but it was not found in either supplier-reported GST source."
        mismatch_fields.append("missing_in_2a_2b")
    elif (rec_2a or rec_2b) and not rec_books:
        category = "Not in Books"
        confidence = 0
        action = "Check whether booking is missed, not our bill, cross-GSTIN, or ineligible."
        explanation = "The invoice is visible in GST data, but no matching books entry was found."
        mismatch_fields.append("missing_in_books")
    elif rec_2a and not rec_2b:
        category = "2A not in 2B"
        confidence = 45
        action = "Review ITC eligibility and timing before claiming."
        explanation = "The invoice is visible in dynamic 2A but not in static 2B."
        mismatch_fields.append("missing_in_2b")
    elif rec_2b and not rec_2a:
        category = "2B not in 2A"
        confidence = 45
        action = "Review source download period and vendor filing details."
        explanation = "The invoice is visible in 2B but not in the uploaded 2A data."
        mismatch_fields.append("missing_in_2a")
    elif values_match(values, tolerance):
        category = "Matched"
        confidence = 100
        action = "No action required, subject to ITC eligibility review."
        explanation = "UID and invoice value match across available sources within tolerance."
    else:
        category = "Amount Differences"
        confidence = 65
        action = "Review taxable value, GST amount, TDS adjustment, debit/credit notes, and partial booking."
        explanation = "UID is found, but invoice values differ beyond tolerance."
        mismatch_fields.append("invoice_amount")

    if duplicate_uid:
        mismatch_fields.append("duplicate_uid")
        explanation += " Duplicate UID entries were aggregated for this result."

    return ThreeWayResult(
        category=category,
        confidence=confidence,
        gstr2a=rec_2a,
        gstr2b=rec_2b,
        books=rec_books,
        mismatch_fields=mismatch_fields,
        suggested_action=action,
        explanation=explanation,
    )


def bucket(records: Iterable[ThreeWayRecord]) -> dict[str, list[ThreeWayRecord]]:
    grouped: dict[str, list[ThreeWayRecord]] = defaultdict(list)
    for record in records:
        grouped[record.uid].append(record)
    return grouped


def aggregate_records(records: list[ThreeWayRecord]) -> ThreeWayRecord | None:
    if not records:
        return None
    chosen = max(records, key=lambda record: abs(record.doc_value))
    if len(records) == 1:
        return chosen
    return ThreeWayRecord(
        source=chosen.source,
        row_id=f"{chosen.source}-AGG-{chosen.uid}",
        supplier_gstin=chosen.supplier_gstin,
        pan=chosen.pan,
        supplier_name=chosen.supplier_name,
        doc_type=chosen.doc_type,
        doc_no=chosen.doc_no,
        normalized_doc_no=chosen.normalized_doc_no,
        doc_date=chosen.doc_date,
        month=chosen.month,
        doc_value=round_money(sum(record.doc_value for record in records)),
        taxable_value=round_money(sum(record.taxable_value for record in records)),
        igst=round_money(sum(record.igst for record in records)),
        cgst=round_money(sum(record.cgst for record in records)),
        sgst=round_money(sum(record.sgst for record in records)),
        cess=round_money(sum(record.cess for record in records)),
        uid=chosen.uid,
        remarks="Aggregated duplicate UID rows",
        raw={"duplicate_count": len(records), "source": chosen.source},
    )


def values_match(values: list[float], tolerance: float) -> bool:
    if len(values) <= 1:
        return True
    reference = values[0]
    return all(abs(round_money(value) - round_money(reference)) <= tolerance for value in values[1:])


def summarize_three_way(results: Iterable[ThreeWayResult]) -> dict[str, int]:
    summary: dict[str, int] = {}
    for result in results:
        summary[result.category] = summary.get(result.category, 0) + 1
    return summary
