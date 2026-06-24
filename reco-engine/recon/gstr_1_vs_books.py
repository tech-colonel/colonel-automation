"""GSTR-1 vs Books Reconciliation Engine (v2 — Accountant Workflow).

Steps (in order):
  0. GSTR-1 Pivot   (optional — only when gstr1_pdf uploaded)
     Validates OCTA Excel monthly totals against GST-portal PDF.
  1. GSTR-1 vs GSTR-3B monthly comparison  → Section 1 of GST Reco sheet
  2. Books All Sales vs GSTR-1 All monthly → Section 2
  3. Books B2B vs GSTR-1 B2B monthly       → Section 3
  4. Books B2C vs GSTR-1 B2C monthly       → Section 4
  5. B2B Reco  — invoice-level, all Tally columns pass-through + Remark
  6. B2C Reco  — State + Rate annual aggregation

Input files:
  tally_sales  (required)  Tally Sales Register .xlsx
  gstr1_octa   (required)  OCTA download with Final GSTR-1 + GSTR3B + GSTR2B sheets
  gstr1_pdf    (optional)  GST-portal GSTR-1 PDF  → enables Step 0
  credit_note  (optional)  Separate Credit Note Register .xlsx
"""

from __future__ import annotations

import logging
import re
from collections import defaultdict
from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from .core import normalize_doc_no, round_money
from .parsers import normalize_header

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_FY_MONTHS = [
    "April", "May", "June", "July", "August", "September",
    "October", "November", "December", "January", "February", "March",
]

_SHORT_TO_FULL = {
    "apr": "April", "may": "May", "jun": "June", "jul": "July",
    "aug": "August", "sep": "September", "oct": "October",
    "nov": "November", "dec": "December", "jan": "January",
    "feb": "February", "mar": "March",
}

# Keywords that flag an Amazon consolidated Tally ledger row
_AMAZON_TALLY_KEYWORDS = (
    "amazon b2b", "amazon inter", "amazon intra",
    "amazon seller", "amazon b2c", "amazon ecommerce",
)

# Regex to detect Amazon consolidated invoice numbers (AMZ-INTER-04, AMZ-INTRA-01, etc.)
_AMZ_INV_RE = re.compile(r"^AMZ[-_](INTER|INTRA)", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Month helpers
# ---------------------------------------------------------------------------

def _norm_month(val) -> str:
    """Normalise any month representation → full name like 'April'."""
    if val is None:
        return "Unknown"
    # Handle pandas NaT
    try:
        import pandas as _pd
        if _pd.isnull(val):
            return "Unknown"
    except Exception:
        pass
    if isinstance(val, datetime):
        try:
            return val.strftime("%B")
        except Exception:
            return "Unknown"
    text = str(val).strip()
    if not text or text.lower() in ("nat", "nan", "none", ""):
        return "Unknown"
    # Datetime-like string
    try:
        return datetime.fromisoformat(text[:10]).strftime("%B")
    except Exception:
        pass
    # Short abbreviation: "Apr 2025" or just "Apr"
    short = text[:3].lower()
    return _SHORT_TO_FULL.get(short, text.title())


def _month_sort_key(m: str) -> int:
    try:
        return _FY_MONTHS.index(m)
    except ValueError:
        return 99


def _sorted_months(months: set) -> list[str]:
    return sorted(months, key=_month_sort_key)


def _zero_amounts() -> dict:
    return {"taxable": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0}


def _add_amounts(acc: dict, taxable=0, igst=0, cgst=0, sgst=0, sign=1) -> None:
    acc["taxable"] += sign * _f(taxable)
    acc["igst"]    += sign * _f(igst)
    acc["cgst"]    += sign * _f(cgst)
    acc["sgst"]    += sign * _f(sgst)


def _f(v) -> float:
    """Safe float conversion, treating NaN/None as 0."""
    try:
        x = float(v)
        return 0.0 if (x != x) else x  # NaN check
    except (TypeError, ValueError):
        return 0.0


def _norm_inv(v) -> str:
    """Normalise invoice number for matching — strip non-alphanumeric, uppercase."""
    return re.sub(r"[^A-Z0-9]", "", str(v or "").upper())


def _is_amazon_tally_row(particulars: str, inv_no: str) -> bool:
    p = particulars.lower()
    if any(kw in p for kw in _AMAZON_TALLY_KEYWORDS):
        return True
    if _AMZ_INV_RE.match(str(inv_no or "").strip()):
        return True
    return False


# ---------------------------------------------------------------------------
# OCTA Excel reader
# ---------------------------------------------------------------------------

def read_octa_excel(file_info: dict) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Auto-detect and return (gstr1_df, gstr3b_df, gstr2b_df) from the OCTA Excel."""
    data = file_info["content"]
    wb = load_workbook(BytesIO(data), data_only=True, read_only=True)

    gstr1_df = pd.DataFrame()
    gstr3b_df = pd.DataFrame()
    gstr2b_df = pd.DataFrame()

    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        # Scan up to first 15 rows to find the real header row
        # (OCTA sheets often have summary numbers in rows 1-4 before the header)
        header_row = None
        for i, r in enumerate(rows[:15]):
            nh = [normalize_header(v) for v in r if v]
            # GSTR-1 header: "doc no" + "customer gstin"
            if gstr1_df.empty and _has_cols(nh, ["doc no", "customer gstin"]):
                header_row = i
                break
            # GSTR-2B header: "supplier gstin" + ("itc eligible" or "gstr1 filing period")
            if gstr2b_df.empty and _has_cols(nh, ["supplier gstin"]) and any("itc" in h or "filing" in h for h in nh):
                header_row = i
                break
            # GSTR-3B header: has "section" column
            if gstr3b_df.empty and any("section" in h for h in nh) and any("apr" in h or "may" in h or "apr 2" in h for h in nh):
                header_row = i
                break

        if header_row is None:
            continue

        headers = [str(v or "").strip() for v in rows[header_row]]
        norm_headers = [normalize_header(h) for h in headers]

        # Detect Final GSTR-1: has "Doc No" and "Customer GSTIN"
        if gstr1_df.empty and _has_cols(norm_headers, ["doc no", "customer gstin"]):
            data_rows = [dict(zip(headers, r)) for r in rows[header_row + 1:]
                         if any(v is not None and str(v).strip() for v in r)]
            gstr1_df = pd.DataFrame(data_rows)
            logger.info("OCTA GSTR-1 sheet detected: '%s' (%d rows)", sheet.title, len(gstr1_df))
            continue

        # Detect GSTR3B: has "Section" column and month columns
        if gstr3b_df.empty and any("section" in h for h in norm_headers):
            all_text = " ".join(str(v) for row in rows for v in row if v)
            if "3.1.a" in all_text.lower():
                data_rows = [dict(zip(headers, r)) for r in rows[header_row + 1:]
                             if any(v is not None and str(v).strip() for v in r)]
                gstr3b_df = pd.DataFrame(data_rows)
                logger.info("OCTA GSTR-3B sheet detected: '%s' (%d rows)", sheet.title, len(gstr3b_df))
                continue

        # Detect GSTR2B: has "Supplier GSTIN" + ITC columns
        if gstr2b_df.empty and _has_cols(norm_headers, ["supplier gstin"]) and any("itc" in h or "filing" in h for h in norm_headers):
            data_rows = [dict(zip(headers, r)) for r in rows[header_row + 1:]
                         if any(v is not None and str(v).strip() for v in r)]
            gstr2b_df = pd.DataFrame(data_rows)
            logger.info("OCTA GSTR-2B sheet detected: '%s' (%d rows)", sheet.title, len(gstr2b_df))

    return gstr1_df, gstr3b_df, gstr2b_df


def _has_cols(norm_headers: list[str], required: list[str]) -> bool:
    joined = " ".join(norm_headers)
    return all(req in joined for req in required)


# ---------------------------------------------------------------------------
# Tally Sales Register reader (raw — preserves all columns)
# ---------------------------------------------------------------------------

def read_tally_sales_raw(file_info: dict) -> pd.DataFrame:
    """Read Tally Sales Register, auto-detect header row, return full DataFrame."""
    data = file_info["content"]
    filename = file_info.get("filename", "")
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    if suffix in {"xlsx", "xlsm", "xls"}:
        wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
        # Use the sheet that looks most like a sales register
        best_sheet = None
        best_score = 0
        for sheet in wb.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            for i, row in enumerate(rows[:15]):
                nh = [normalize_header(v) for v in row if v]
                score = sum(1 for t in ["date", "voucher no", "total sales", "particulars", "category"] if any(t in h for h in nh))
                if score > best_score:
                    best_score = score
                    best_sheet = (sheet, rows, i)
        if best_sheet is None:
            raise ValueError("No recognisable Sales Register sheet found in Tally file.")
        sheet, rows, hdr_idx = best_sheet
        headers = [str(v or f"Col_{i}").strip() for i, v in enumerate(rows[hdr_idx])]
        data_rows = []
        for row in rows[hdr_idx + 1:]:
            if not any(v is not None and str(v).strip() for v in row):
                continue
            data_rows.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
        df = pd.DataFrame(data_rows)
    else:
        df = pd.read_csv(BytesIO(data), encoding="utf-8-sig")

    logger.info("Tally Sales Register: %d rows, %d cols", len(df), len(df.columns))
    return df


def read_credit_note_raw(file_info: dict) -> pd.DataFrame:
    """Read Credit Note Register as raw DataFrame (same structure as Tally)."""
    return read_tally_sales_raw(file_info)


# ---------------------------------------------------------------------------
# Column detection helpers for DataFrames
# ---------------------------------------------------------------------------

def _find_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    """Return the first DataFrame column that matches any candidate (normalized)."""
    norm_cols = {normalize_header(c): c for c in df.columns}
    for cand in candidates:
        key = normalize_header(cand)
        if key in norm_cols:
            return norm_cols[key]
    return None


def _col_val(row, col: str | None, default=None):
    if col is None or col not in row.index:
        return default
    v = row[col]
    if v is None or (isinstance(v, float) and v != v):
        return default
    return v


# ---------------------------------------------------------------------------
# GSTR-3B monthly totals extractor
# ---------------------------------------------------------------------------

def extract_gstr3b_monthly(gstr3b_df: pd.DataFrame) -> dict[str, dict]:
    """
    Extract monthly taxable/IGST/CGST/SGST from GSTR-3B DataFrame.
    Sums 3.1.A + 3.1.B (zero-rated) for taxable and IGST.
    Returns {month_name: {taxable, igst, cgst, sgst}}.
    """
    if gstr3b_df.empty:
        return {}

    # Identify the "Section" and "Type" columns
    section_col = _find_col(gstr3b_df, ["Section", "Section Name"])
    type_col    = _find_col(gstr3b_df, ["Type", "Row Type"])

    if section_col is None:
        logger.warning("GSTR-3B: no Section column found")
        return {}

    # Month columns: everything after "Type" that looks like a month
    skip_cols = {section_col, type_col} if type_col else {section_col}
    month_cols: dict[str, str] = {}  # month_name → column_name
    for col in gstr3b_df.columns:
        if col in skip_cols:
            continue
        m = _norm_month(col)
        if m in _FY_MONTHS:
            month_cols[m] = col

    if not month_cols:
        logger.warning("GSTR-3B: no month columns detected")
        return {}

    result: dict[str, dict] = {m: _zero_amounts() for m in month_cols}

    for _, row in gstr3b_df.iterrows():
        section = str(_col_val(row, section_col, "")).strip()
        rtype   = str(_col_val(row, type_col, "")).strip().lower() if type_col else ""

        is_3_1a = "3.1.a" in section.lower() or (section.startswith("3.1.A") if section else False)
        is_3_1b = "3.1.b" in section.lower() or (section.startswith("3.1.B") if section else False)

        if not (is_3_1a or is_3_1b):
            continue

        is_supply = "supply value" in rtype or "supply" in rtype
        is_igst   = "integrated" in rtype or "igst" in rtype
        is_cgst   = "central" in rtype or "cgst" in rtype
        is_sgst   = ("state" in rtype and "supply" not in rtype) or "sgst" in rtype

        for month, col in month_cols.items():
            val = _f(_col_val(row, col, 0))
            if val == 0:
                continue
            if is_supply:
                result[month]["taxable"] += val
            elif is_igst:
                result[month]["igst"] += val
            elif is_cgst:
                result[month]["cgst"] += val
            elif is_sgst:
                result[month]["sgst"] += val

    return result


# ---------------------------------------------------------------------------
# GSTR-1 monthly aggregator (from OCTA Final GSTR-1)
# ---------------------------------------------------------------------------

def aggregate_gstr1_monthly(gstr1_df: pd.DataFrame) -> dict[str, dict]:
    """Aggregate Final GSTR-1 sheet by Tax Period month."""
    if gstr1_df.empty:
        return {}

    period_col  = _find_col(gstr1_df, ["Tax Period", "Tax period", "Period", "Return Period"])
    taxable_col = _find_col(gstr1_df, ["Item Taxable Value", "Taxable Value", "Taxable Amount", "Taxable"])
    igst_col    = _find_col(gstr1_df, ["IGST", "Integrated Tax", "IGST Amount"])
    cgst_col    = _find_col(gstr1_df, ["CGST", "Central Tax", "CGST Amount"])
    sgst_col    = _find_col(gstr1_df, ["SGST", "State Tax", "SGST Amount"])

    if period_col is None:
        logger.warning("GSTR-1: Tax Period column not found")
        return {}

    result: dict[str, dict] = defaultdict(_zero_amounts)
    for _, row in gstr1_df.iterrows():
        month = _norm_month(_col_val(row, period_col))
        if month not in _FY_MONTHS:
            continue
        _add_amounts(
            result[month],
            taxable=_col_val(row, taxable_col, 0),
            igst=_col_val(row, igst_col, 0),
            cgst=_col_val(row, cgst_col, 0),
            sgst=_col_val(row, sgst_col, 0),
        )
    return dict(result)


# ---------------------------------------------------------------------------
# Books monthly aggregator (from Tally Sales Register)
# ---------------------------------------------------------------------------

def aggregate_books_monthly(tally_df: pd.DataFrame, category: str | None = None) -> dict[str, dict]:
    """
    Aggregate Tally by month.
    category: None=all rows, 'B2B'=only B2B rows, 'B2C'=only B2C rows.
    """
    if tally_df.empty:
        return {}

    month_col   = _find_col(tally_df, ["Month", "Mth"])
    date_col    = _find_col(tally_df, ["Date", "Invoice Date", "Voucher Date"])
    cat_col     = _find_col(tally_df, ["Category", "Catogary", "Cat", "Type"])
    gstin_col   = _find_col(tally_df, ["GSTIN", "GSTIN/UIN", "Buyer GSTIN", "GST No"])
    taxable_col = _find_col(tally_df, ["Total Sales", "Taxable Value", "Taxable Amount", "Net Amount"])
    igst_col    = _find_col(tally_df, ["Total IGST", "IGST", "Integrated Tax"])
    cgst_col    = _find_col(tally_df, ["Total CGST", "CGST", "Central Tax"])
    sgst_col    = _find_col(tally_df, ["Total SGST", "SGST", "State Tax"])

    result: dict[str, dict] = defaultdict(_zero_amounts)

    for _, row in tally_df.iterrows():
        # Determine month
        month = _norm_month(_col_val(row, month_col) or _col_val(row, date_col))
        if month not in _FY_MONTHS:
            continue

        # Skip header-like rows
        taxable = _f(_col_val(row, taxable_col, 0))
        igst_v  = _f(_col_val(row, igst_col, 0))
        cgst_v  = _f(_col_val(row, cgst_col, 0))
        sgst_v  = _f(_col_val(row, sgst_col, 0))
        if taxable == 0 and igst_v == 0 and cgst_v == 0 and sgst_v == 0:
            continue

        # Category filter
        if category is not None:
            row_cat = _infer_category(row, cat_col, gstin_col)
            if row_cat != category:
                continue

        _add_amounts(result[month], taxable=taxable, igst=igst_v, cgst=cgst_v, sgst=sgst_v)

    return dict(result)


def _infer_category(row, cat_col: str | None, gstin_col: str | None) -> str:
    """Return 'B2B' or 'B2C' for a Tally row."""
    if cat_col:
        cat = str(_col_val(row, cat_col, "")).strip().upper()
        if "B2B" in cat:
            return "B2B"
        if "B2C" in cat:
            return "B2C"
    # Fall back to GSTIN presence
    if gstin_col:
        gstin = str(_col_val(row, gstin_col, "")).strip()
        if gstin and gstin.upper() not in ("#N/A", "N/A", "NA", "NONE", ""):
            if len(re.sub(r"[^A-Z0-9]", "", gstin.upper())) == 15:
                return "B2B"
    return "B2C"


# ---------------------------------------------------------------------------
# Monthly comparison table builder (for GST Reco sections)
# ---------------------------------------------------------------------------

def build_monthly_comparison(
    left: dict[str, dict],
    right: dict[str, dict],
    left_label: str,
    right_label: str,
) -> list[dict]:
    """
    Build list of monthly comparison rows.
    Each row: {month, left_taxable/igst/cgst/sgst, right_taxable/igst/cgst/sgst, diff_*}
    """
    all_months = set(left) | set(right)
    rows = []
    for month in _sorted_months(all_months):
        lv = left.get(month, _zero_amounts())
        rv = right.get(month, _zero_amounts())
        rows.append({
            "month": month,
            f"{left_label}_taxable": round_money(lv["taxable"]),
            f"{left_label}_igst":    round_money(lv["igst"]),
            f"{left_label}_cgst":    round_money(lv["cgst"]),
            f"{left_label}_sgst":    round_money(lv["sgst"]),
            f"{right_label}_taxable": round_money(rv["taxable"]),
            f"{right_label}_igst":    round_money(rv["igst"]),
            f"{right_label}_cgst":    round_money(rv["cgst"]),
            f"{right_label}_sgst":    round_money(rv["sgst"]),
            "diff_taxable": round_money(lv["taxable"] - rv["taxable"]),
            "diff_igst":    round_money(lv["igst"]    - rv["igst"]),
            "diff_cgst":    round_money(lv["cgst"]    - rv["cgst"]),
            "diff_sgst":    round_money(lv["sgst"]    - rv["sgst"]),
        })
    # Totals row
    if rows:
        tot = {
            "month": "Total",
            f"{left_label}_taxable":  sum(r[f"{left_label}_taxable"]  for r in rows),
            f"{left_label}_igst":     sum(r[f"{left_label}_igst"]     for r in rows),
            f"{left_label}_cgst":     sum(r[f"{left_label}_cgst"]     for r in rows),
            f"{left_label}_sgst":     sum(r[f"{left_label}_sgst"]     for r in rows),
            f"{right_label}_taxable": sum(r[f"{right_label}_taxable"] for r in rows),
            f"{right_label}_igst":    sum(r[f"{right_label}_igst"]    for r in rows),
            f"{right_label}_cgst":    sum(r[f"{right_label}_cgst"]    for r in rows),
            f"{right_label}_sgst":    sum(r[f"{right_label}_sgst"]    for r in rows),
            "diff_taxable": sum(r["diff_taxable"] for r in rows),
            "diff_igst":    sum(r["diff_igst"]    for r in rows),
            "diff_cgst":    sum(r["diff_cgst"]    for r in rows),
            "diff_sgst":    sum(r["diff_sgst"]    for r in rows),
        }
        rows.append(tot)
    return rows


# ---------------------------------------------------------------------------
# B2B Reconciliation (new — invoice-level, all Tally columns pass-through)
# ---------------------------------------------------------------------------

def reconcile_b2b_new(
    tally_df: pd.DataFrame,
    gstr1_df: pd.DataFrame,
    tolerance: float = 1.0,
) -> list[dict]:
    """
    Match Tally B2B rows against GSTR-1 by invoice number.
    Returns list of row dicts: all original Tally columns + GSTR-1 matched cols + Diff + Remark.
    Unmatched GSTR-1 rows are appended at the bottom.
    """
    if tally_df.empty:
        return []

    # Detect key Tally columns
    inv_col     = _find_col(tally_df, ["Voucher No.", "Voucher No", "Voucher Ref. No.", "Invoice No", "Doc No", "Bill No"])
    ref_col     = _find_col(tally_df, ["Voucher Ref. No.", "Voucher Ref No", "Ref No"])
    cat_col     = _find_col(tally_df, ["Category", "Catogary"])
    part_col    = _find_col(tally_df, ["Particulars", "Party Name", "Buyer", "Ledger Name"])
    gstin_col   = _find_col(tally_df, ["GSTIN", "GSTIN/UIN", "Buyer GSTIN"])
    taxable_col = _find_col(tally_df, ["Total Sales", "Taxable Value", "Taxable Amount"])
    igst_col    = _find_col(tally_df, ["Total IGST", "IGST"])
    cgst_col    = _find_col(tally_df, ["Total CGST", "CGST"])
    sgst_col    = _find_col(tally_df, ["Total SGST", "SGST"])

    # Build GSTR-1 index by normalised invoice number → row dict
    g1_inv_col     = _find_col(gstr1_df, ["Doc No", "Invoice Number", "Invoice No"])
    g1_gstin_col   = _find_col(gstr1_df, ["Customer GSTIN", "GSTIN of Recipient", "Buyer GSTIN"])
    g1_taxable_col = _find_col(gstr1_df, ["Item Taxable Value", "Taxable Value", "Taxable"])
    g1_igst_col    = _find_col(gstr1_df, ["IGST", "Integrated Tax"])
    g1_cgst_col    = _find_col(gstr1_df, ["CGST", "Central Tax"])
    g1_sgst_col    = _find_col(gstr1_df, ["SGST", "State Tax"])

    # Index GSTR-1 rows
    gstr1_by_inv: dict[str, list[int]] = defaultdict(list)
    gstr1_rows = []
    if not gstr1_df.empty:
        for idx, row in gstr1_df.iterrows():
            # Skip B2C rows (no Customer GSTIN)
            gstin = str(_col_val(row, g1_gstin_col, "")).strip()
            if not gstin or gstin.upper() in ("", "N/A", "NA"):
                continue
            inv = str(_col_val(row, g1_inv_col, "")).strip()
            norm = _norm_inv(inv)
            g1_row = {
                "inv_no":  inv,
                "gstin":   gstin,
                "taxable": _f(_col_val(row, g1_taxable_col, 0)),
                "igst":    _f(_col_val(row, g1_igst_col, 0)),
                "cgst":    _f(_col_val(row, g1_cgst_col, 0)),
                "sgst":    _f(_col_val(row, g1_sgst_col, 0)),
                "_idx":    len(gstr1_rows),
            }
            gstr1_rows.append(g1_row)
            if norm:
                gstr1_by_inv[norm].append(len(gstr1_rows) - 1)

    matched_gstr1: set[int] = set()
    result_rows: list[dict] = []

    for _, tally_row in tally_df.iterrows():
        # Get key fields
        inv_no   = str(_col_val(tally_row, inv_col, "")).strip()
        ref_no   = str(_col_val(tally_row, ref_col, "")).strip() if ref_col else ""
        part     = str(_col_val(tally_row, part_col, "")).strip()
        taxable  = _f(_col_val(tally_row, taxable_col, 0))
        igst_v   = _f(_col_val(tally_row, igst_col, 0))
        cgst_v   = _f(_col_val(tally_row, cgst_col, 0))
        sgst_v   = _f(_col_val(tally_row, sgst_col, 0))

        # Skip empty / header-like rows
        if not inv_no and not part and taxable == 0:
            continue

        # Check category
        row_cat = _infer_category(tally_row, cat_col, gstin_col)
        if row_cat != "B2B":
            continue  # B2C rows handled separately

        # Build output row: all original Tally columns
        out = {col: _json_safe_val(tally_row[col]) for col in tally_df.columns}

        # Amazon detection
        effective_inv = inv_no or ref_no
        if _is_amazon_tally_row(part, effective_inv):
            # Try to find a GSTR-1 match anyway (for reference)
            norm = _norm_inv(effective_inv)
            candidates = [j for j in gstr1_by_inv.get(norm, []) if j not in matched_gstr1]
            if candidates:
                g1 = gstr1_rows[candidates[0]]
                out.update({
                    "_gstr1_inv_no": g1["inv_no"],
                    "_gstr1_gstin":  g1["gstin"],
                    "_gstr1_taxable": g1["taxable"],
                    "_gstr1_igst":   g1["igst"],
                    "_gstr1_cgst":   g1["cgst"],
                    "_gstr1_sgst":   g1["sgst"],
                    "_diff_taxable": round_money(taxable - g1["taxable"]),
                    "_diff_igst":    round_money(igst_v  - g1["igst"]),
                    "_diff_cgst":    round_money(cgst_v  - g1["cgst"]),
                    "_diff_sgst":    round_money(sgst_v  - g1["sgst"]),
                })
                matched_gstr1.add(candidates[0])
            else:
                _set_empty_gstr1(out)
            out["_remark"] = "Amazon Entry As per Tally"
            result_rows.append(out)
            continue

        # Regular B2B matching
        norm = _norm_inv(effective_inv)
        candidates = [j for j in gstr1_by_inv.get(norm, []) if j not in matched_gstr1]

        if candidates:
            # Pick best match by closest taxable value
            best_j = min(candidates, key=lambda j: abs(taxable - gstr1_rows[j]["taxable"]))
            g1 = gstr1_rows[best_j]
            dt = round_money(taxable - g1["taxable"])
            di = round_money(igst_v  - g1["igst"])
            dc = round_money(cgst_v  - g1["cgst"])
            ds = round_money(sgst_v  - g1["sgst"])
            is_diff = any(abs(d) > tolerance for d in (dt, di, dc, ds))
            out.update({
                "_gstr1_inv_no":  g1["inv_no"],
                "_gstr1_gstin":   g1["gstin"],
                "_gstr1_taxable": g1["taxable"],
                "_gstr1_igst":    g1["igst"],
                "_gstr1_cgst":    g1["cgst"],
                "_gstr1_sgst":    g1["sgst"],
                "_diff_taxable":  dt,
                "_diff_igst":     di,
                "_diff_cgst":     dc,
                "_diff_sgst":     ds,
                "_remark":        "Diff" if is_diff else "Match",
            })
            matched_gstr1.add(best_j)
        else:
            _set_empty_gstr1(out)
            out["_remark"] = "Not in GSTR-1"

        result_rows.append(out)

    # Append unmatched GSTR-1 rows at the bottom
    for j, g1 in enumerate(gstr1_rows):
        if j in matched_gstr1:
            continue
        out: dict = {col: None for col in tally_df.columns}
        # Fill in the GSTR-1 columns
        is_amazon = any(kw in g1["inv_no"].lower() for kw in ("amz", "amazon"))
        out.update({
            "_gstr1_inv_no":  g1["inv_no"],
            "_gstr1_gstin":   g1["gstin"],
            "_gstr1_taxable": g1["taxable"],
            "_gstr1_igst":    g1["igst"],
            "_gstr1_cgst":    g1["cgst"],
            "_gstr1_sgst":    g1["sgst"],
            "_diff_taxable": -g1["taxable"],
            "_diff_igst":    -g1["igst"],
            "_diff_cgst":    -g1["cgst"],
            "_diff_sgst":    -g1["sgst"],
            "_remark": "Amazon Entry as per GSTR-1" if is_amazon else "Not in Books",
        })
        result_rows.append(out)

    logger.info("B2B Reco: %d rows (%d GSTR-1 unmatched)", len(result_rows), len(gstr1_rows) - len(matched_gstr1))
    return result_rows


def _set_empty_gstr1(out: dict) -> None:
    out.update({
        "_gstr1_inv_no": "", "_gstr1_gstin": "",
        "_gstr1_taxable": None, "_gstr1_igst": None,
        "_gstr1_cgst": None, "_gstr1_sgst": None,
        "_diff_taxable": None, "_diff_igst": None,
        "_diff_cgst": None, "_diff_sgst": None,
    })


def _json_safe_val(v):
    """Convert a value to JSON-serialisable form."""
    if v is None:
        return None
    # Handle pandas NaT / NaN
    try:
        import pandas as _pd
        if _pd.isnull(v):
            return None
    except Exception:
        pass
    if isinstance(v, bool):
        return v
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return None if v != v else v  # NaN → None
    if isinstance(v, datetime):
        try:
            return v.strftime("%Y-%m-%d")
        except Exception:
            return None
    # pandas Timestamp
    try:
        return v.strftime("%Y-%m-%d")
    except AttributeError:
        pass
    if isinstance(v, str):
        return v
    try:
        f = float(v)
        return None if f != f else f
    except (TypeError, ValueError):
        return str(v)


# ---------------------------------------------------------------------------
# B2C Reconciliation (state + rate annual aggregation)
# ---------------------------------------------------------------------------

def reconcile_b2c_new(
    tally_df: pd.DataFrame,
    gstr1_df: pd.DataFrame,
    tolerance: float = 1.0,
) -> list[dict]:
    """State + rate annual aggregation: GSTR-1 B2C vs Books B2C."""
    # Books B2C aggregation
    cat_col     = _find_col(tally_df, ["Category", "Catogary"])
    gstin_col   = _find_col(tally_df, ["GSTIN", "GSTIN/UIN"])
    state_col   = _find_col(tally_df, ["States", "Place of Supply", "State"])
    rate_col    = _find_col(tally_df, ["Rate", "Tax Rate", "GST Rate"])
    taxable_col = _find_col(tally_df, ["Total Sales", "Taxable Value"])
    igst_col    = _find_col(tally_df, ["Total IGST", "IGST"])
    cgst_col    = _find_col(tally_df, ["Total CGST", "CGST"])
    sgst_col    = _find_col(tally_df, ["Total SGST", "SGST"])

    books_agg: dict[tuple, dict] = defaultdict(_zero_amounts)

    for _, row in tally_df.iterrows():
        cat = _infer_category(row, cat_col, gstin_col)
        if cat != "B2C":
            continue
        taxable = _f(_col_val(row, taxable_col, 0))
        igst_v  = _f(_col_val(row, igst_col, 0))
        cgst_v  = _f(_col_val(row, cgst_col, 0))
        sgst_v  = _f(_col_val(row, sgst_col, 0))
        if taxable == 0 and igst_v == 0 and cgst_v == 0 and sgst_v == 0:
            continue
        state = str(_col_val(row, state_col, "")).strip().title() or "Unknown"
        rate  = _f(_col_val(row, rate_col, 0))
        key   = (state, rate)
        _add_amounts(books_agg[key], taxable=taxable, igst=igst_v, cgst=cgst_v, sgst=sgst_v)

    # GSTR-1 B2C aggregation
    g1_gstin_col   = _find_col(gstr1_df, ["Customer GSTIN", "GSTIN of Recipient"])
    g1_state_col   = _find_col(gstr1_df, ["Place of Supply", "State"])
    g1_rate_col    = _find_col(gstr1_df, ["GST Rate", "Rate", "Tax Rate"])
    g1_taxable_col = _find_col(gstr1_df, ["Item Taxable Value", "Taxable Value"])
    g1_igst_col    = _find_col(gstr1_df, ["IGST", "Integrated Tax"])
    g1_cgst_col    = _find_col(gstr1_df, ["CGST", "Central Tax"])
    g1_sgst_col    = _find_col(gstr1_df, ["SGST", "State Tax"])

    gstr1_agg: dict[tuple, dict] = defaultdict(_zero_amounts)

    if not gstr1_df.empty:
        for _, row in gstr1_df.iterrows():
            gstin = str(_col_val(row, g1_gstin_col, "")).strip()
            if gstin and len(re.sub(r"[^A-Z0-9]", "", gstin.upper())) == 15:
                continue  # B2B row — skip
            state = _extract_pos_state(str(_col_val(row, g1_state_col, "")).strip())
            rate  = _f(_col_val(row, g1_rate_col, 0))
            key   = (state, rate)
            _add_amounts(
                gstr1_agg[key],
                taxable=_f(_col_val(row, g1_taxable_col, 0)),
                igst=_f(_col_val(row, g1_igst_col, 0)),
                cgst=_f(_col_val(row, g1_cgst_col, 0)),
                sgst=_f(_col_val(row, g1_sgst_col, 0)),
            )

    all_keys = sorted(set(books_agg) | set(gstr1_agg), key=lambda k: (k[0], k[1]))
    rows = []
    for state, rate in all_keys:
        bv = books_agg.get((state, rate), _zero_amounts())
        gv = gstr1_agg.get((state, rate), _zero_amounts())
        rows.append({
            "state":           state,
            "rate":            rate,
            "gstr1_taxable":   round_money(gv["taxable"]),
            "gstr1_igst":      round_money(gv["igst"]),
            "gstr1_cgst":      round_money(gv["cgst"]),
            "gstr1_sgst":      round_money(gv["sgst"]),
            "books_taxable":   round_money(bv["taxable"]),
            "books_igst":      round_money(bv["igst"]),
            "books_cgst":      round_money(bv["cgst"]),
            "books_sgst":      round_money(bv["sgst"]),
            "diff_taxable":    round_money(gv["taxable"] - bv["taxable"]),
            "diff_igst":       round_money(gv["igst"]    - bv["igst"]),
            "diff_cgst":       round_money(gv["cgst"]    - bv["cgst"]),
            "diff_sgst":       round_money(gv["sgst"]    - bv["sgst"]),
        })
    logger.info("B2C Reco: %d state+rate rows", len(rows))
    return rows


def _extract_pos_state(pos: str) -> str:
    """Strip state code prefix like '07-Delhi' → 'Delhi'."""
    if "-" in pos:
        parts = pos.split("-", 1)
        if parts[0].strip().isdigit():
            return parts[1].strip().title()
    return pos.strip().title()


# ---------------------------------------------------------------------------
# PDF parser for GSTR-1 portal PDF
# ---------------------------------------------------------------------------

def parse_gstr1_pdf_monthly(pdf_bytes: bytes) -> dict[str, dict] | None:
    """
    Extract monthly totals from GST-portal GSTR-1 PDF.
    Returns {month: {taxable, igst, cgst, sgst}} or None on failure.
    """
    try:
        import pdfplumber
    except ImportError:
        logger.warning("pdfplumber not installed — PDF parsing skipped")
        return None

    try:
        result: dict[str, dict] = {}
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    if not table or len(table) < 2:
                        continue
                    header = [str(c or "").lower().strip() for c in table[0]]
                    # Look for Month | Taxable | IGST | CGST | SGST columns
                    month_idx   = _col_idx(header, ["month", "period", "return period"])
                    taxable_idx = _col_idx(header, ["taxable", "taxable value", "taxable amount"])
                    igst_idx    = _col_idx(header, ["igst", "integrated"])
                    cgst_idx    = _col_idx(header, ["cgst", "central"])
                    sgst_idx    = _col_idx(header, ["sgst", "state"])
                    if month_idx is None or (taxable_idx is None and igst_idx is None):
                        continue
                    for data_row in table[1:]:
                        if not data_row or len(data_row) <= (month_idx or 0):
                            continue
                        month = _norm_month(data_row[month_idx])
                        if month not in _FY_MONTHS:
                            continue
                        result[month] = {
                            "taxable": _parse_pdf_num(data_row, taxable_idx),
                            "igst":    _parse_pdf_num(data_row, igst_idx),
                            "cgst":    _parse_pdf_num(data_row, cgst_idx),
                            "sgst":    _parse_pdf_num(data_row, sgst_idx),
                        }
        return result if result else None
    except Exception as e:
        logger.warning("PDF parsing failed: %s", e)
        return None


def _col_idx(header: list[str], candidates: list[str]) -> int | None:
    for i, h in enumerate(header):
        if any(c in h for c in candidates):
            return i
    return None


def _parse_pdf_num(row: list, idx: int | None) -> float:
    if idx is None or idx >= len(row):
        return 0.0
    return _f(str(row[idx] or "").replace(",", "").strip())


# ---------------------------------------------------------------------------
# DataFrame → JSON-safe records helper
# ---------------------------------------------------------------------------

def df_to_records(df: pd.DataFrame) -> list[dict]:
    """Convert DataFrame to list of JSON-safe dicts."""
    records = []
    for _, row in df.iterrows():
        records.append({col: _json_safe_val(row[col]) for col in df.columns})
    return records


# ---------------------------------------------------------------------------
# Summary builder (for frontend display)
# ---------------------------------------------------------------------------

def build_summary(b2b_rows: list[dict], b2c_rows: list[dict]) -> dict:
    remark_counts: dict[str, int] = defaultdict(int)
    for r in b2b_rows:
        remark_counts[r.get("_remark", "Unknown")] += 1
    return {
        "B2B rows": len(b2b_rows),
        "B2C rows": len(b2c_rows),
        **{f"B2B: {k}": v for k, v in remark_counts.items()},
    }
