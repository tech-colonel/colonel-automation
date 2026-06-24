"""GSTR-3B vs GSTR-2B Line-Item-Wise Reconciliation Engine.

SOP Summary
-----------
1.  Combine month-wise GSTR-2B files into one dataset (sheets: B2B, CDNR, IMPG, ECO, ISD …).
    Add *Month*, *Source* = "2B", *Remarks* = "Match".
2.  The GSTR-3B dataset comes from the 3B return-filing workings (it is NOT a portal download
    because 3B is a summary return).  Source = "3B Working", Remarks = "Match".
3.  Reverse the sign of GSTR-3B values so that a pivot of the combined data shows only the
    variance between the two sources.
4.  Build a Pivot: Rows = Month, Values = SUM of IGST / CGST / SGST, Filters = Source & Remark.
5.  Identify:
      • Entries in 2B but NOT claimed in 3B  → "Not Claimed in 3B"
      • Entries claimed in 3B but NOT in 2B  → "Not in 2B"
6.  For matched entries the pivot totals should net to zero month-wise.

This module produces a combined line-item dataset and a month-wise pivot summary.
"""

from __future__ import annotations

import logging
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from datetime import datetime
from io import BytesIO, StringIO
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook

from .core import (
    json_safe,
    normalize_doc_no,
    normalize_doc_type,
    normalize_gstin,
    parse_date,
    round_money,
)
from .parsers import normalize_header

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Sheet detection hints
# ---------------------------------------------------------------------------
GSTR2B_SHEET_HINTS = ("b2b", "cdnr", "impg", "eco", "isd", "2b")
GSTR3B_SHEET_HINTS = ("3b", "working", "3b working", "gstr3b", "gstr-3b")

# Sheets to SKIP from 3B working files — these are sales/output-side data,
# NOT purchase/ITC.  The reconciliation is 2B (ITC available) vs 3B (ITC claimed).
SALES_SHEET_SKIP = (
    "sale", "sales", "sales register", "sales working",
    "gstr1", "gstr-1", "gstr 1", "output", "output tax",
    "revenue", "b2cs", "b2cl", "hsn", "at", "atadj",
    "cdna", "b2ba", "exp", "nil",
    "cn register", "credit note register",       # CN issued = sales side
    "dn register", "debit note register",         # DN issued = sales side
    "compliment", "complimentry", "custom complimentry",
)

# Sheets to ALSO SKIP because they are embedded copies of 2B data inside the
# 3B working file — user uploads the actual 2B separately.
EMBEDDED_2B_SKIP = ("2b", "2bdn", "2b dn", "2b-dn", "gstr2b", "gstr-2b")

# PURCHASE-side sheet whitelist — sheets that are definitely purchase/ITC data.
PURCHASE_SHEET_HINTS = (
    "purchase", "purchase register", "purchase working",
    "input", "input tax", "itc", "rcm",
    "3b working", "3b purchase", "gstr3b", "gstr-3b",
    "working",
)

MONTH_NAMES = {
    "january": "January", "february": "February", "march": "March",
    "april": "April", "may": "May", "june": "June",
    "july": "July", "august": "August", "september": "September",
    "october": "October", "november": "November", "december": "December",
    "jan": "January", "feb": "February", "mar": "March",
    "apr": "April", "jun": "June", "jul": "July",
    "aug": "August", "sep": "September", "oct": "October",
    "nov": "November", "dec": "December",
}

# ---------------------------------------------------------------------------
# Voucher type keywords that indicate SALES / OUTPUT entries (to be skipped)
# ---------------------------------------------------------------------------
SALES_VOUCHER_KEYWORDS = (
    "sales", "sale", "output", "revenue", "gstr-1", "gstr1",
    "tax invoice", "debit note issued", "credit note issued",
)

# Voucher type keywords that indicate PURCHASE / ITC entries (to be kept)
PURCHASE_VOUCHER_KEYWORDS = (
    "purchase", "input", "expense", "rcm", "import", "isd",
    "debit note received", "credit note received",
    "journal", "payment",  # may contain ITC reversals
)

# ---------------------------------------------------------------------------
# Field aliases — covers portal 2B downloads AND typical 3B working sheets
# ---------------------------------------------------------------------------
FIELD_ALIASES = {
    "supplier_gstin": [
        "gstin of supplier", "supplier gstin", "gstin/uin", "gstin uin",
        "gstin", "ctin", "vendor gstin", "party gstin", "supplier gst",
        "gstin of isd", "gstin of ecommerce",
    ],
    "supplier_name": [
        "trade/legal name", "trade name", "legal name", "supplier name",
        "party name", "vendor name", "particulars",
    ],
    "doc_no": [
        "supplier invoice no", "supplier invoice no.",  # CRITICAL: must be before voucher no
        "voucher ref. no.", "voucher ref no", "voucher ref. no",  # Debit Note section
        "invoice number", "invoice no", "invoice no.", "document number",
        "document no", "doc no", "inum", "no", "bill no", "bill no.",
        "note number", "nt num", "isd document number",
        "voucher no", "voucher no.",  # fallback — only used if no supplier invoice no
    ],
    "doc_date": [
        "invoice date", "document date", "doc date", "date",
        "voucher date", "bill date", "idt", "dt",
        "isd document date", "note date",
    ],
    "doc_type": [
        "document type", "doc type", "invoice type", "type", "typ",
        "note type", "isd document type",
    ],
    "voucher_type": [
        "voucher type", "vch type", "transaction type", "entry type",
        "voucher category",
    ],
    "month": [
        "month", "tax period", "return period", "filing period",
        "gstr-1 filing period", "gstr1 filing period",
        "gstr-1/iff/gstr-5 period",
    ],
    "value": [
        "invoice value", "invoice value(₹)", "invoice value (₹)",
        "note value", "note value (₹)", "note value(₹)",
        "document value", "document value(₹)",
        "value", "total value", "gross total",
        "doc value", "total invoice value", "val",
    ],
    "taxable_value": [
        "taxable value", "taxable value (₹)", "taxable value(₹)",
        "taxable amount", "txval", "assessable value",
        "taxable",
    ],
    "igst": [
        "igst", "integrated tax", "integrated tax(₹)",
        "integrated tax amount", "iamt",
    ],
    "cgst": [
        "cgst", "central tax", "central tax(₹)",
        "central tax amount", "camt",
    ],
    "sgst": [
        "sgst", "utgst", "state tax", "state/ut tax",
        "state/ut tax(₹)", "state tax(₹)",
        "state tax amount", "samt",
    ],
    "cess": ["cess", "cess amount", "cess(₹)", "csamt"],
    "source": ["source"],
    "remarks": ["remarks", "remark"],
    "remarks2": [
        "remarks 2", "remarks 2 detail explanation",
        "remarks 2 detail", "remarks2",
    ],
}


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------
@dataclass
class LineItem:
    """A single line-item from either GSTR-2B or GSTR-3B working."""
    source: str          # "2B - B2B", "2B - CDNR", "3B Working", etc.
    row_id: str
    month: str = ""
    date: str = ""
    invoice_no: str = ""
    normalized_invoice_no: str = ""
    party_name: str = ""
    supplier_gstin: str = ""
    value: float = 0.0
    cgst: float = 0.0
    sgst: float = 0.0
    igst: float = 0.0
    cess: float = 0.0
    taxable: float = 0.0
    remarks: str = "Match"
    remarks2: str = ""
    effect_in_books: str = ""
    effect_in_9c: str = ""
    raw: dict[str, Any] = field(default_factory=dict)

    @property
    def total_tax(self) -> float:
        return round_money(self.igst + self.cgst + self.sgst + self.cess)

    @property
    def uid(self) -> str:
        """Identity key for matching: party_name (normalized) + invoice_no."""
        name_key = normalize_doc_no(self.party_name) if self.party_name else ""
        gstin_key = self.supplier_gstin or ""
        identity = gstin_key or name_key
        return f"{identity}|{self.normalized_invoice_no}" if identity and self.normalized_invoice_no else ""

    def as_dict(self) -> dict[str, Any]:
        result = json_safe(asdict(self))
        result["total_tax"] = self.total_tax
        result["uid"] = self.uid
        return result


@dataclass
class RecoResult:
    """One reconciled row in the output."""
    category: str        # "Matched", "Not in 2B", "Not Claimed in 3B", "Amount Mismatch"
    confidence: int
    rec_2b: LineItem | None
    rec_3b: LineItem | None
    mismatch_fields: list[str]
    suggested_action: str
    explanation: str

    def as_dict(self) -> dict[str, Any]:
        return {
            "category": self.category,
            "confidence": self.confidence,
            "rec_2b": self.rec_2b.as_dict() if self.rec_2b else None,
            "rec_3b": self.rec_3b.as_dict() if self.rec_3b else None,
            "mismatch_fields": self.mismatch_fields,
            "suggested_action": self.suggested_action,
            "explanation": self.explanation,
        }


# ---------------------------------------------------------------------------
# File readers
# ---------------------------------------------------------------------------
def read_3b_2b_uploads(files: dict[str, dict]) -> tuple[list[LineItem], list[LineItem]]:
    """Read uploaded files and return (gstr2b_records, gstr3b_records)."""
    logger.info("=== read_3b_2b_uploads ===")
    logger.info("Files received: %s", {k: {'filename': v.get('filename'), 'size': len(v.get('content', b''))} for k, v in files.items()})

    gstr2b_file = files.get("gstr2b_3b")
    gstr3b_file = files.get("gstr3b")

    if not gstr2b_file or not gstr2b_file.get("content"):
        raise ValueError("Upload GSTR-2B file (combined monthly Excel with B2B/CDNR/IMPG sheets).")
    if not gstr3b_file or not gstr3b_file.get("content"):
        raise ValueError("Upload GSTR-3B working file (Excel from 3B return filing workings).")

    gstr2b_records = _read_file(gstr2b_file, "2B")
    gstr3b_records = _read_file(gstr3b_file, "3B Working")

    logger.info("Parsed: %d GSTR-2B records, %d GSTR-3B records", len(gstr2b_records), len(gstr3b_records))
    if not gstr2b_records:
        raise ValueError("No records found in the GSTR-2B file. Check that the file has B2B/CDNR/IMPG sheets with invoice data.")
    if not gstr3b_records:
        raise ValueError("No records found in the GSTR-3B working file. Ensure it has invoice-level working data.")
    return gstr2b_records, gstr3b_records


def _read_file(file_info: dict, source_prefix: str) -> list[LineItem]:
    filename = file_info["filename"]
    data = file_info["content"]
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    if suffix in {"xlsx", "xlsm"}:
        return _read_excel(data, source_prefix)
    if suffix == "csv":
        frame = pd.read_csv(StringIO(data.decode("utf-8-sig", errors="replace"))).fillna("")
        return _normalize_rows(frame.to_dict(orient="records"), source_prefix, "CSV")
    raise ValueError(f"Unsupported file type: {filename}")


def _read_excel(data: bytes, source_prefix: str) -> list[LineItem]:
    workbook = load_workbook(BytesIO(data), data_only=True, read_only=True)
    all_records: list[LineItem] = []

    for sheet in workbook.worksheets:
        name = normalize_header(sheet.title)
        logger.info("  Sheet '%s' -> normalized: '%s' (source_prefix=%s)", sheet.title, name, source_prefix)

        rows_raw = list(sheet.iter_rows(values_only=True))
        if not rows_raw:
            continue

        # SPECIAL LOGIC: Look for 'Custom Complimentry' in working or summary sheets
        if source_prefix == "3B Working" and ("working" in name or "summary" in name):
            for row in rows_raw:
                col0 = str(row[0] or "").strip()
                if "custom complimentry" in col0.lower() or "custom complimentary" in col0.lower():
                    try:
                        taxable = float(row[1] or 0)
                        igst = float(row[2] or 0)
                        cgst = float(row[3] or 0)
                        sgst = float(row[4] or 0)
                        val = taxable + igst + cgst + sgst
                        
                        month_str = ""
                        parts = col0.split()
                        if len(parts) >= 3:
                            month_str = parts[-1] # e.g. "Dec-24"

                        item = LineItem(
                            source="3B Working",
                            row_id=f"COMPLIMENTRY_{taxable}",
                            party_name="Business Promotion & Marketing",
                            invoice_no="",
                            month=month_str,
                            taxable=taxable,
                            igst=igst,
                            cgst=cgst,
                            sgst=sgst,
                            value=val,
                            remarks2="ITC Reversed",
                            supplier_gstin=""
                        )
                        all_records.append(item)
                        logger.info("    -> Found and injected Custom Complimentry line: %s (Value: %s)", col0, val)
                    except (ValueError, TypeError, IndexError) as e:
                        logger.warning("    -> Found Custom Complimentry line but failed to parse values: %s", e)

        # Determine the source label for this sheet
        source_label = _detect_source_label(name, source_prefix)
        if not source_label:
            logger.info("    -> SKIPPED (not relevant for %s)", source_prefix)
            continue

        if source_prefix == "2B" or name in ("2b", "2bdn", "2b-dn", "2b dn"):
            # GSTR-2B portal files and embedded 2B copies have multi-row headers
            records = _parse_2b_portal_sheet(rows_raw, source_label, sheet.title)
            all_records.extend(records)
        else:
            # Other 3B working files: may have multiple sections with different headers
            records = _parse_3b_multi_section_sheet(rows_raw, source_label, sheet.title)
            all_records.extend(records)

    return all_records


def _parse_2b_portal_sheet(rows_raw: list[tuple], source_label: str, sheet_title: str) -> list[LineItem]:
    """Parse a GSTR-2B portal sheet with multi-row merged headers.

    Portal sheets have headers split across 2 ADJACENT rows:
      Row 5: GSTIN of supplier | Trade/Legal name | Invoice Details | None | None | None | Place of supply | ... | Taxable Value | Tax Amount | ...
      Row 6: None              | None             | Invoice number  | Invoice type | Invoice Date | Invoice Value(₹) | None | ... | None | Integrated Tax(₹) | Central Tax(₹) | ...

    We find both rows, merge them (taking the non-empty value for each column),
    and use the merged result as the final header.
    """
    all_aliases = set()
    for aliases in FIELD_ALIASES.values():
        for a in aliases:
            all_aliases.add(normalize_header(a))

    # Score every row in the first 25 rows for how many alias columns it has
    row_scores: list[tuple[int, int]] = []  # (index, score)
    for idx, row in enumerate(rows_raw[:25]):
        normalized = [normalize_header(v) for v in row]
        score = sum(1 for v in normalized if v in all_aliases)
        if score >= 2:
            row_scores.append((idx, score))

    if not row_scores:
        logger.info("    -> SKIPPED (no header detected in 2B sheet '%s')", sheet_title)
        return []

    # Find pairs of adjacent header rows (the portal always uses 2 rows)
    best_pair = None
    best_combined_score = 0
    for i in range(len(row_scores)):
        for j in range(i + 1, len(row_scores)):
            idx_a, score_a = row_scores[i]
            idx_b, score_b = row_scores[j]
            if idx_b - idx_a == 1:  # Adjacent rows
                combined = score_a + score_b
                if combined > best_combined_score:
                    best_combined_score = combined
                    best_pair = (idx_a, idx_b)

    if best_pair:
        row_a_idx, row_b_idx = best_pair
        row_a = [str(v or "") for v in rows_raw[row_a_idx]]
        row_b = [str(v or "") for v in rows_raw[row_b_idx]]
        max_cols = max(len(row_a), len(row_b))
        merged = []
        for i in range(max_cols):
            a = row_a[i].strip() if i < len(row_a) else ""
            b = row_b[i].strip() if i < len(row_b) else ""
            # Take whichever is non-empty; if both non-empty, prefer the one
            # that is a recognized alias (more specific)
            if a and b:
                a_is_alias = normalize_header(a) in all_aliases
                b_is_alias = normalize_header(b) in all_aliases
                merged.append(b if b_is_alias and not a_is_alias else a)
            elif b:
                merged.append(b)
            elif a:
                merged.append(a)
            else:
                merged.append(f"col_{i}")
        headers = merged
        data_start = row_b_idx + 1
        logger.info("    Merged 2-row header (rows %d+%d): %s", row_a_idx, row_b_idx, headers[:15])
    else:
        # No adjacent pair found — use the single best header row
        best_idx = max(row_scores, key=lambda x: x[1])[0]
        headers = [str(v or "") for v in rows_raw[best_idx]]
        data_start = best_idx + 1
        logger.info("    Single header at row %d: %s", best_idx, headers[:15])

    logger.info("    Final header: %s", headers[:15])

    records: list[dict[str, Any]] = []
    for row in rows_raw[data_start:]:
        if not any(v is not None and str(v).strip() for v in row):
            continue
        record = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))}
        records.append(record)

    normalized = _normalize_rows(records, source_label, sheet_title)
    logger.info("    -> %d records from 2B sheet '%s' as '%s'", len(normalized), sheet_title, source_label)
    return normalized


def _parse_3b_multi_section_sheet(rows_raw: list[tuple], source_label: str, sheet_title: str) -> list[LineItem]:
    """Parse a 3B working sheet that may have multiple sections with different headers.

    Purchase Register files often have:
      Section 1 (rows 1-N): Purchases/Journals with 'Supplier Invoice No.' column
      Section 2 (rows N+3-M): Debit Notes with 'Voucher Ref. No.' column
    Each section has its own header row. We detect and parse each section separately.
    """
    all_records: list[LineItem] = []
    all_aliases = set()
    for aliases in FIELD_ALIASES.values():
        for a in aliases:
            all_aliases.add(normalize_header(a))

    # Find ALL header rows in the sheet
    header_indices = []
    for idx, row in enumerate(rows_raw[:60]):  # scan first 60 rows
        normalized = [normalize_header(v) for v in row]
        score = sum(1 for v in normalized if v in all_aliases)
        if score >= 3:  # need at least 3 matching columns for a section header
            # Avoid treating data rows as headers by checking if previous was also a header
            if header_indices and idx - header_indices[-1] <= 1:
                continue
            header_indices.append(idx)

    if not header_indices:
        # Fall back to best single header
        header_idx = _detect_header_row(rows_raw)
        if header_idx is not None:
            header_indices = [header_idx]

    logger.info("    Found %d section header(s) at rows: %s", len(header_indices), header_indices)

    for sec_idx, h_idx in enumerate(header_indices):
        headers = [str(v or "") for v in rows_raw[h_idx]]
        # Determine where this section ends (next header or end of data)
        if sec_idx + 1 < len(header_indices):
            end_idx = header_indices[sec_idx + 1]
        else:
            end_idx = len(rows_raw)

        records: list[dict[str, Any]] = []
        for row in rows_raw[h_idx + 1:end_idx]:
            if not any(v is not None and str(v).strip() for v in row):
                continue
            record = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))}
            records.append(record)

        normalized = _normalize_rows(records, source_label, sheet_title)
        logger.info("    -> %d records from section %d (header row %d) of '%s'",
                    len(normalized), sec_idx + 1, h_idx, sheet_title)
        all_records.extend(normalized)

    return all_records


def _detect_source_label(sheet_name: str, source_prefix: str) -> str | None:
    """Determine the source label for a sheet given the file's source prefix."""
    if source_prefix == "2B":
        # Accept standard GSTR-2B section sheets: B2B, B2BA, B2B-CDNR, B2B-CDNRA
        # CRITICAL: check 'cdnr' BEFORE 'b2b' because sheet names like
        # 'B2B-CDNR' contain both substrings; CDNR must win.
        if "cdnr" in sheet_name:
            return "2B - CDNR"
        for hint in ("b2b", "b2ba", "impg", "eco", "isd", "impgsez"):
            if hint in sheet_name:
                return f"2B - {hint.upper()}"
        # Also accept sheets explicitly named 2b or gstr-2b
        if "2b" in sheet_name and "3b" not in sheet_name:
            return "2B - B2B"
        return None
    else:
        # For 3B working — FOCUS ONLY ON "2B" and "2B-DN" SHEETS
        # The user's specific process copies 2B portal data directly into "2B" and "2B-DN" 
        # tabs in their working file to mark what is actually claimed. 
        # They DO NOT want to use the raw Purchase Register.
        if sheet_name in ("2b", "2bdn", "2b-dn", "2b dn", "b2b", "b2b-cdnr", "b2ba", "b2b-cdnra"):
            return "3B Working"
            
        logger.info("    -> SKIPPED sheet: '%s' (Core logic: only use 2B and 2B-DN sheets for 3B Working)", sheet_name)
        return None


def _detect_header_row(rows: list[tuple[Any, ...]]) -> int | None:
    all_aliases = set()
    for aliases in FIELD_ALIASES.values():
        for a in aliases:
            all_aliases.add(normalize_header(a))

    best_idx = None
    best_score = 0

    for idx, row in enumerate(rows[:25]):
        normalized = [normalize_header(v) for v in row]
        score = sum(1 for v in normalized if v in all_aliases)
        if score > best_score:
            best_score = score
            best_idx = idx

    return best_idx if best_score >= 2 else None


def _normalize_rows(rows: list[dict[str, Any]], source: str, sheet_name: str) -> list[LineItem]:
    output: list[LineItem] = []
    skipped_sales = 0
    for idx, row in enumerate(rows, start=1):
        getter = _build_getter(row)

        # --- Row-level sales filtering for 3B Working ---
        # GSTR-3B working files often mix purchase and sales entries.
        # We must ONLY keep purchase/ITC entries for this reconciliation.
        if "3B" in source:
            if _is_sales_row(getter, row):
                skipped_sales += 1
                continue

        invoice_no = _clean_value(getter("doc_no"))
        party_name = str(getter("supplier_name") or "").strip()
        supplier_gstin = normalize_gstin(getter("supplier_gstin"))

        # Skip total/summary rows (e.g., "Total", "Grand Total")
        if party_name.lower() in ("total", "grand total", ""):
            if not invoice_no and not supplier_gstin:
                continue

        # Skip re-header rows (column headers repeated mid-data)
        if party_name.lower() in ("particulars", "party name", "supplier name", "trade/legal name"):
            continue

        # Need at least invoice_no or party_name to be useful
        if not invoice_no and not party_name and not supplier_gstin:
            continue

        month_raw = getter("month")
        date_raw = getter("doc_date")
        month = _format_month(month_raw, date_raw)

        value = round_money(getter("value"))
        taxable = round_money(getter("taxable_value"))
        cgst = round_money(getter("cgst"))
        sgst = round_money(getter("sgst"))
        igst = round_money(getter("igst"))
        cess = round_money(getter("cess"))

        # Infer value from components if missing
        if not value and (taxable or cgst or sgst or igst):
            value = round_money(taxable + cgst + sgst + igst + cess)

        # Infer taxable from value - taxes if missing
        if not taxable and value:
            taxable = round_money(value - cgst - sgst - igst - cess)

        remarks = str(getter("remarks") or "Match").strip()
        remarks2 = str(getter("remarks2") or "").strip()

        record = LineItem(
            source=source,
            row_id=f"{source}-{sheet_name}-{idx}",
            month=month,
            date=parse_date(date_raw),
            invoice_no=invoice_no,
            normalized_invoice_no=normalize_doc_no(invoice_no),
            party_name=party_name,
            supplier_gstin=supplier_gstin,
            value=value,
            cgst=cgst,
            sgst=sgst,
            igst=igst,
            cess=cess,
            taxable=taxable,
            remarks=remarks,
            remarks2=remarks2,
            raw={str(k): v for k, v in row.items() if str(k).strip()},
        )
        output.append(record)

    if skipped_sales:
        logger.info("    [%s] Skipped %d sales/output rows (purchase-only filter)", sheet_name, skipped_sales)
    return output


def _is_sales_row(getter, row: dict[str, Any]) -> bool:
    """Check if a row is a sales/output entry that should be excluded.

    Returns True if the row looks like a sales entry (output GST),
    False if it looks like purchase/ITC or is ambiguous (keep it).
    """
    # 1. Check voucher_type column (common in Tally-exported 3B workings)
    voucher_type = str(getter("voucher_type") or "").strip().lower()
    if voucher_type:
        for kw in SALES_VOUCHER_KEYWORDS:
            if kw in voucher_type:
                return True
        # If voucher_type explicitly says purchase, keep it
        for kw in PURCHASE_VOUCHER_KEYWORDS:
            if kw in voucher_type:
                return False

    # 2. Check source column — if the file pre-labels rows
    source_val = str(getter("source") or "").strip().lower()
    if source_val:
        for kw in SALES_VOUCHER_KEYWORDS:
            if kw in source_val:
                return True

    # 3. Check all column headers for sales-indicator columns
    #    e.g. "Output CGST", "Output IGST", "Sales" columns with values
    for key, val in row.items():
        key_lower = normalize_header(key)
        if any(kw in key_lower for kw in ("output cgst", "output sgst", "output igst",
                                           "sales cgst", "sales sgst", "sales igst",
                                           "output tax")):
            # If this sales-tax column has a non-zero value, it's a sales row
            money = round_money(val)
            if money != 0:
                return True

    return False


def _build_getter(row: dict[str, Any]):
    normalized_map = {normalize_header(k): v for k, v in row.items()}

    def get(field_name: str) -> Any:
        for alias in FIELD_ALIASES.get(field_name, []):
            key = normalize_header(alias)
            if key in normalized_map and normalized_map[key] not in (None, ""):
                return normalized_map[key]
        return ""
    return get


def _clean_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _format_month(month_value: Any, date_value: Any) -> str:
    """Convert to a display month like 'April', 'May', etc."""
    if month_value not in (None, ""):
        text = str(month_value).strip().lower()
        # Check if it's a known month name
        if text in MONTH_NAMES:
            return MONTH_NAMES[text]
        # Try parsing as a date
        parsed = parse_date(month_value)
        if parsed and "-" in parsed:
            try:
                return datetime.fromisoformat(parsed).strftime("%B")
            except Exception:
                pass
        return str(month_value).strip()

    parsed = parse_date(date_value)
    if parsed and "-" in parsed:
        try:
            return datetime.fromisoformat(parsed).strftime("%B")
        except Exception:
            pass
    return ""


# ---------------------------------------------------------------------------
# Reconciliation logic
# ---------------------------------------------------------------------------
def reconcile_3b_vs_2b(
    gstr2b_records: list[LineItem],
    gstr3b_records: list[LineItem],
    tolerance: float = 1.0,
) -> list[RecoResult]:
    """
    Match GSTR-2B line items against GSTR-3B working line items.

    Matching strategy:
    1. Primary: Match by UID (GSTIN/party + invoice_no)
    2. Secondary: Match by invoice_no + approximate amount
    3. Remaining 2B → "Not Claimed in 3B"
    4. Remaining 3B → "Not in 2B"
    """
    results: list[RecoResult] = []
    matched_2b: set[int] = set()
    matched_3b: set[int] = set()

    # Build indexes
    idx_3b_by_uid: dict[str, list[int]] = defaultdict(list)
    idx_3b_by_docno: dict[str, list[int]] = defaultdict(list)
    for i, rec in enumerate(gstr3b_records):
        if rec.uid:
            idx_3b_by_uid[rec.uid].append(i)
        if rec.normalized_invoice_no:
            idx_3b_by_docno[rec.normalized_invoice_no].append(i)

    # --- Pass 1: Exact UID match ---
    for i_2b, rec_2b in enumerate(gstr2b_records):
        if not rec_2b.uid:
            continue
        candidates = [j for j in idx_3b_by_uid.get(rec_2b.uid, []) if j not in matched_3b]
        if not candidates:
            continue

        # Pick best candidate by amount closeness
        best_j = min(candidates, key=lambda j: abs(rec_2b.value - gstr3b_records[j].value))
        rec_3b = gstr3b_records[best_j]

        mismatches = _compare_amounts(rec_2b, rec_3b, tolerance)
        if not mismatches:
            results.append(RecoResult(
                category="Matched",
                confidence=100,
                rec_2b=rec_2b,
                rec_3b=rec_3b,
                mismatch_fields=[],
                suggested_action="No action. ITC matched between 2B and 3B filing.",
                explanation="Invoice found in both GSTR-2B and GSTR-3B working with matching amounts.",
            ))
        else:
            results.append(RecoResult(
                category="Amount Mismatch",
                confidence=75,
                rec_2b=rec_2b,
                rec_3b=rec_3b,
                mismatch_fields=mismatches,
                suggested_action="Review tax amounts. Check for partial claim, TDS adjustment, or rounding.",
                explanation=f"Invoice matched but {', '.join(mismatches)} differ beyond tolerance.",
            ))

        matched_2b.add(i_2b)
        matched_3b.add(best_j)

    # --- Pass 2: Doc-no cross-match for remaining ---
    for i_2b, rec_2b in enumerate(gstr2b_records):
        if i_2b in matched_2b:
            continue
        if not rec_2b.normalized_invoice_no:
            continue

        candidates = [j for j in idx_3b_by_docno.get(rec_2b.normalized_invoice_no, [])
                       if j not in matched_3b]
        if not candidates:
            continue

        best_j = min(candidates, key=lambda j: abs(rec_2b.value - gstr3b_records[j].value))
        rec_3b = gstr3b_records[best_j]
        mismatches = _compare_amounts(rec_2b, rec_3b, tolerance)

        if not mismatches:
            results.append(RecoResult(
                category="Matched",
                confidence=85,
                rec_2b=rec_2b,
                rec_3b=rec_3b,
                mismatch_fields=["matched_by_doc_no"],
                suggested_action="Matched by invoice number. Verify GSTIN mapping.",
                explanation="Invoice matched by document number (party identity differed).",
            ))
        else:
            results.append(RecoResult(
                category="Amount Mismatch",
                confidence=60,
                rec_2b=rec_2b,
                rec_3b=rec_3b,
                mismatch_fields=mismatches,
                suggested_action="Review amounts. Matched by invoice number only.",
                explanation=f"Invoice cross-matched by doc no, but {', '.join(mismatches)} differ.",
            ))

        matched_2b.add(i_2b)
        matched_3b.add(best_j)

    # --- Pass 3: Unmatched 2B records → "Not Claimed in 3B" ---
    for i_2b, rec_2b in enumerate(gstr2b_records):
        if i_2b in matched_2b:
            continue
        results.append(RecoResult(
            category="Not Claimed in 3B",
            confidence=0,
            rec_2b=rec_2b,
            rec_3b=None,
            mismatch_fields=["not_in_3b"],
            suggested_action="Check if ITC was intentionally not claimed (ineligible, blocked u/s 17(5), or timing).",
            explanation="Invoice appears in GSTR-2B but was not found in the GSTR-3B working/filing.",
        ))

    # --- Pass 4: Unmatched 3B records → "Not in 2B" ---
    for i_3b, rec_3b in enumerate(gstr3b_records):
        if i_3b in matched_3b:
            continue
        results.append(RecoResult(
            category="Not in 2B",
            confidence=0,
            rec_2b=None,
            rec_3b=rec_3b,
            mismatch_fields=["not_in_2b"],
            suggested_action="Verify if this is an ITC reversal, RCM entry, or claimed from books without 2B support.",
            explanation="Entry claimed in GSTR-3B working but not appearing in GSTR-2B data.",
        ))

    logger.info("3B vs 2B reconciliation: %d results (%d matched, %d not-in-3B, %d not-in-2B)",
                len(results),
                sum(1 for r in results if r.category in ("Matched", "Amount Mismatch")),
                sum(1 for r in results if r.category == "Not Claimed in 3B"),
                sum(1 for r in results if r.category == "Not in 2B"))
    return results


def _compare_amounts(rec_2b: LineItem, rec_3b: LineItem, tolerance: float) -> list[str]:
    """Compare amounts between 2B and 3B records.

    Accountant-grade logic:
    - Compare TAX amounts (IGST, CGST, SGST) using absolute values
      because CDNR entries have opposite signs between 2B and PR.
    - If tax amounts match, classify as Matched even if value/taxable differ
      (2B Note Value includes tax, PR Taxable Value excludes tax — this is normal).
    - Only flag Amount Mismatch when actual tax components diverge.
    """
    mismatches = []

    # Compare tax amounts using absolute values (handles sign differences in CDNR)
    if abs(abs(rec_2b.cgst) - abs(rec_3b.cgst)) > tolerance:
        mismatches.append("cgst")
    if abs(abs(rec_2b.sgst) - abs(rec_3b.sgst)) > tolerance:
        mismatches.append("sgst")
    if abs(abs(rec_2b.igst) - abs(rec_3b.igst)) > tolerance:
        mismatches.append("igst")

    # Only check value/taxable if tax amounts also mismatch
    # (value/taxable often differ between 2B and PR due to different reporting bases)
    if mismatches:
        if abs(abs(rec_2b.value) - abs(rec_3b.value)) > tolerance:
            mismatches.append("value")
        if abs(abs(rec_2b.taxable) - abs(rec_3b.taxable)) > tolerance:
            mismatches.append("taxable")

    return mismatches


# ---------------------------------------------------------------------------
# Pivot / summary builders
# ---------------------------------------------------------------------------
def summarize_3b_vs_2b(results: list[RecoResult]) -> dict[str, int]:
    summary: dict[str, int] = {}
    for r in results:
        summary[r.category] = summary.get(r.category, 0) + 1
    return summary


def build_month_pivot(results: list[RecoResult]) -> list[dict[str, Any]]:
    """
    Build a month-wise pivot table as per the SOP.
    For each month, sum IGST/CGST/SGST:
      - 2B values positive
      - 3B values negative (sign-reversed)
    Filters: Source, Remark.
    Net should be zero for matched entries.
    """
    pivot: dict[str, dict[str, float]] = defaultdict(lambda: {
        "igst_2b": 0.0, "cgst_2b": 0.0, "sgst_2b": 0.0,
        "igst_3b": 0.0, "cgst_3b": 0.0, "sgst_3b": 0.0,
        "value_2b": 0.0, "value_3b": 0.0,
        "taxable_2b": 0.0, "taxable_3b": 0.0,
        "count_2b": 0, "count_3b": 0,
    })

    for r in results:
        if r.rec_2b:
            month = r.rec_2b.month or "Unknown"
            pivot[month]["igst_2b"] += r.rec_2b.igst
            pivot[month]["cgst_2b"] += r.rec_2b.cgst
            pivot[month]["sgst_2b"] += r.rec_2b.sgst
            pivot[month]["value_2b"] += r.rec_2b.value
            pivot[month]["taxable_2b"] += r.rec_2b.taxable
            pivot[month]["count_2b"] += 1
        if r.rec_3b:
            month = r.rec_3b.month or "Unknown"
            # Sign-reversed for 3B (step 8 of SOP)
            pivot[month]["igst_3b"] += r.rec_3b.igst
            pivot[month]["cgst_3b"] += r.rec_3b.cgst
            pivot[month]["sgst_3b"] += r.rec_3b.sgst
            pivot[month]["value_3b"] += r.rec_3b.value
            pivot[month]["taxable_3b"] += r.rec_3b.taxable
            pivot[month]["count_3b"] += 1

    # Sort by FY month order
    month_order = [
        "April", "May", "June", "July", "August", "September",
        "October", "November", "December", "January", "February", "March",
    ]
    order_map = {m: i for i, m in enumerate(month_order)}

    rows = []
    for month, data in sorted(pivot.items(), key=lambda x: order_map.get(x[0], 99)):
        rows.append({
            "month": month,
            "count_2b": data["count_2b"],
            "count_3b": data["count_3b"],
            "igst_2b": round_money(data["igst_2b"]),
            "cgst_2b": round_money(data["cgst_2b"]),
            "sgst_2b": round_money(data["sgst_2b"]),
            "igst_3b": round_money(data["igst_3b"]),
            "cgst_3b": round_money(data["cgst_3b"]),
            "sgst_3b": round_money(data["sgst_3b"]),
            # Net difference (2B - 3B)
            "igst_diff": round_money(data["igst_2b"] - data["igst_3b"]),
            "cgst_diff": round_money(data["cgst_2b"] - data["cgst_3b"]),
            "sgst_diff": round_money(data["sgst_2b"] - data["sgst_3b"]),
            "value_2b": round_money(data["value_2b"]),
            "value_3b": round_money(data["value_3b"]),
            "value_diff": round_money(data["value_2b"] - data["value_3b"]),
            "taxable_2b": round_money(data["taxable_2b"]),
            "taxable_3b": round_money(data["taxable_3b"]),
            "taxable_diff": round_money(data["taxable_2b"] - data["taxable_3b"]),
        })
    return rows
