"""
GSTR-2B vs Books reconciliation.

Handles:
  - GSTR-2B portal Excel: two-row header at rows 4+5, data from row 6
  - Purchase Register / Debit Note Register: 6-row Tally preamble, header row 6, data from row 7+

Matching strategy:
  - Primary key: normalised invoice/note number (no GSTIN because Tally doesn't export GSTIN)
  - Amount tolerances applied to taxable_value, igst, cgst, sgst
  - Remarks: "Matched", "In 2B but not in the books" (only 2B), "Need to 2B" (only Books),
              "Amount Mismatch" (both sides but amounts differ)
"""
from __future__ import annotations

import re
from io import BytesIO
from typing import Any

import pandas as pd

from .core import (
    NormalizedInvoice,
    MatchResult,
    normalize_doc_no,
    parse_date,
    round_money,
    amount_equal,
    summarize,
)


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

ALLOWED_SHEETS = {"B2B", "B2BA", "B2B-CDNR", "B2B-CDNRA"}
# Ordered so originals (B2B, B2B-CDNR) are parsed before amendments (B2BA, B2B-CDNRA).
# This ensures the original entry is first in the records list and gets matched with Books
# first in Pass 1 — leaving the amendment entry (B2BA/B2B-CDNRA) unmatched as expected.
ALLOWED_SHEETS_ORDER = ["B2B", "B2B-CDNR", "B2BA", "B2B-CDNRA"]

GST_STATE_CODES: dict[str, str] = {
    "01": "Jammu & Kashmir",        "02": "Himachal Pradesh",
    "03": "Punjab",                 "04": "Chandigarh",
    "05": "Uttarakhand",            "06": "Haryana",
    "07": "Delhi",                  "08": "Rajasthan",
    "09": "Uttar Pradesh",          "10": "Bihar",
    "11": "Sikkim",                 "12": "Arunachal Pradesh",
    "13": "Nagaland",               "14": "Manipur",
    "15": "Mizoram",                "16": "Tripura",
    "17": "Meghalaya",              "18": "Assam",
    "19": "West Bengal",            "20": "Jharkhand",
    "21": "Odisha",                 "22": "Chhattisgarh",
    "23": "Madhya Pradesh",         "24": "Gujarat",
    "25": "Daman & Diu",            "26": "Dadra & Nagar Haveli",
    "27": "Maharashtra",            "28": "Andhra Pradesh",
    "29": "Karnataka",              "30": "Goa",
    "31": "Lakshadweep",            "32": "Kerala",
    "33": "Tamil Nadu",             "34": "Puducherry",
    "35": "Andaman & Nicobar",      "36": "Telangana",
    "37": "Andhra Pradesh (New)",   "38": "Ladakh",
    "97": "Other Territory",        "99": "Centre Jurisdiction",
}

# Matches the 15-character Indian GSTIN format: 2 digits + 5 letters + 4 digits + letter + alphanum + Z + alphanum
_GSTIN_RE = re.compile(r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][0-9A-Z]Z[0-9A-Z]$')


def _extract_gstin(rec) -> str:
    """Return GSTIN from a NormalizedInvoice: try supplier_gstin first, then scan raw dict."""
    direct = str(_get_val(rec, "supplier_gstin", "") or "").strip().upper()
    if direct and _GSTIN_RE.match(direct):
        return direct
    raw = _get_val(rec, "raw", {})
    if isinstance(raw, dict):
        for v in raw.values():
            v_str = str(v or "").strip().upper()
            if _GSTIN_RE.match(v_str):
                return v_str
    return ""


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def convert_xls_to_xlsx(xls_bytes: bytes) -> bytes:
    import pandas as pd
    from io import BytesIO
    xls_file = pd.ExcelFile(BytesIO(xls_bytes), engine="xlrd")
    xlsx_buffer = BytesIO()
    with pd.ExcelWriter(xlsx_buffer, engine="openpyxl") as writer:
        for sheet_name in xls_file.sheet_names:
            df = xls_file.parse(sheet_name, header=None)
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
    xlsx_buffer.seek(0)
    return xlsx_buffer.read()


def _ensure_xlsx(data: bytes) -> bytes:
    if data and data.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        try:
            return convert_xls_to_xlsx(data)
        except Exception as e:
            print(f"Error converting legacy .xls to .xlsx: {e}")
    return data



def _norm(text: Any) -> str:
    s = str(text or "").strip().lower()
    s = re.sub(r"[\n\r\t]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _build_getter(row: dict[str, Any], col_map: dict[str, list[str]]):
    norm_row = {_norm(k): v for k, v in row.items()}

    def get(field: str) -> Any:
        for alias in col_map.get(field, []):
            key = _norm(alias)
            if key in norm_row and str(norm_row[key]).strip() not in ("", "nan"):
                return norm_row[key]
        return ""

    return get


# ---------------------------------------------------------------------------
# GSTR-2B Parser
# ---------------------------------------------------------------------------

GSTR2B_COL_MAP: dict[str, list[str]] = {
    "supplier_gstin":  ["gstin of supplier", "gstin"],
    "supplier_name":   ["trade/legal name", "trade name", "legal name"],
    "doc_no":          ["invoice number", "note number", "invoice no", "note no"],
    "doc_type":        ["invoice type", "note type", "type"],
    "doc_date":        ["invoice date", "note date", "date"],
    "taxable_value":   ["taxable value (₹)", "taxable value"],
    "igst":            ["integrated tax (₹)", "integrated tax(₹)", "integrated tax"],
    "cgst":            ["central tax (₹)", "central tax(₹)", "central tax"],
    "sgst":            ["state/ut tax (₹)", "state/ut tax(₹)", "state/ut tax"],
    "cess":            ["cess (₹)", "cess(₹)", "cess"],
    "invoice_value":   ["invoice value (₹)", "invoice value(₹)", "invoice value"],
}


def _read_gstr2b_sheet(data: bytes, sheet_name: str) -> list[dict[str, Any]]:
    """
    GSTR-2B sheets have two merged header rows at index 4 and 5 (0-based).
    Data starts from row index 6.
    """
    raw = pd.read_excel(BytesIO(data), sheet_name=sheet_name, header=None, dtype=object)
    if len(raw) < 7:
        return []

    row4 = [str(v).strip() if str(v) != "nan" else "" for v in raw.iloc[4]]
    row5 = [str(v).strip() if str(v) != "nan" else "" for v in raw.iloc[5]]

    # Forward-fill group headers
    last = ""
    row4_filled = []
    for v in row4:
        if v:
            last = v
        row4_filled.append(last)

    # Prefer sub-header; make columns unique by appending index if duplicate
    combined = []
    seen: dict[str, int] = {}
    for top, sub in zip(row4_filled, row5):
        name = sub if (sub and sub != top) else top
        count = seen.get(name, 0)
        seen[name] = count + 1
        combined.append(f"{name}_{count}" if count else name)

    data_rows = raw.iloc[6:].copy()
    data_rows.columns = combined
    data_rows = data_rows.dropna(how="all")

    # GSTR-2B portal exports merge supplier GSTIN and name cells across all invoices
    # for the same supplier — pandas reads merged cells as NaN after the first row.
    # Forward-fill those columns so every invoice row carries its supplier info.
    gstin_aliases = {"gstin of supplier", "gstin"}
    name_aliases = {"trade/legal name", "trade name", "legal name"}
    for col in data_rows.columns:
        if _norm(col) in gstin_aliases or _norm(col) in name_aliases:
            data_rows[col] = data_rows[col].replace("", pd.NA).ffill()

    rows = data_rows.fillna("").to_dict(orient="records")
    for row in rows:
        row["_sheet"] = sheet_name
    return rows


# ---------------------------------------------------------------------------
# OCTA GSTR-2B flat-format support
# (single sheet, one header row at the top, one row per invoice — NOT the portal
#  multi-sheet B2B/B2BA/CDNR layout). Detected and parsed separately; the portal
#  path below is left completely unchanged.)
# ---------------------------------------------------------------------------

GSTR2B_OCTA_COL_MAP: dict[str, list[str]] = {
    "supplier_gstin":  ["supplier gstin"],
    "supplier_name":   ["supplier name"],
    "tally_name":      ["as per tally name"],
    "doc_no":          ["doc no", "document number", "invoice number"],
    "doc_type":        ["doc type", "document type"],
    "doc_date":        ["doc date", "document date"],
    "taxable_value":   ["item taxable value", "taxable value"],
    "igst":            ["igst", "integrated tax"],
    "cgst":            ["cgst", "central tax"],
    "sgst":            ["sgst", "state/ut tax"],
    "cess":            ["cess"],
    "invoice_value":   ["doc value", "document value"],
    "reverse_charge":  ["reverse charge"],
    "place_of_supply": ["place of supply"],
    "is_amendment":    ["is amendment"],
    "original_doc_no": ["original doc no", "original document number"],
}

# Signature headers that uniquely identify the OCTA export vs the GST portal export.
_OCTA_SIGNATURE = ("item taxable value", "supplier gstin")


def _find_octa_header_idx(raw) -> int:
    """Return the 0-based index of the OCTA header row, or -1 if not an OCTA sheet."""
    for idx in range(min(15, len(raw))):
        cells = {_norm(v) for v in raw.iloc[idx]}
        if all(sig in cells for sig in _OCTA_SIGNATURE):
            return idx
    return -1


def _is_octa_format(data: bytes) -> bool:
    """True when the first sheet carries the OCTA signature headers."""
    try:
        raw = pd.read_excel(BytesIO(data), sheet_name=0, header=None, dtype=object)
    except Exception:
        return False
    return _find_octa_header_idx(raw) >= 0


def _read_octa_sheet(data: bytes) -> list[dict[str, Any]]:
    """Read the OCTA flat sheet into row dicts (header detected dynamically)."""
    raw = pd.read_excel(BytesIO(data), sheet_name=0, header=None, dtype=object)
    header_idx = _find_octa_header_idx(raw)
    if header_idx < 0:
        return []
    raw_headers = [str(v).strip() if str(v) != "nan" else f"_col{i}"
                   for i, v in enumerate(raw.iloc[header_idx])]
    seen: dict[str, int] = {}
    headers = []
    for h in raw_headers:
        count = seen.get(h, 0)
        seen[h] = count + 1
        headers.append(f"{h}_{count}" if count else h)
    data_rows = raw.iloc[header_idx + 1:].copy()
    data_rows.columns = headers
    data_rows = data_rows.dropna(how="all")
    return data_rows.fillna("").to_dict(orient="records")


def _parse_gstr2b_octa(data: bytes) -> list[NormalizedInvoice]:
    """Parse the OCTA flat GSTR-2B export into NormalizedInvoice records.

    Mirrors parse_gstr2b's NormalizedInvoice construction. Credit Notes are stored
    PRE-SIGNED (negative) in OCTA, so values are normalised with -abs() for CRN —
    never re-negated like the portal path (which receives positive CN values)."""
    records: list[NormalizedInvoice] = []
    index = 0
    for row in _read_octa_sheet(data):
        get = _build_getter(row, GSTR2B_OCTA_COL_MAP)
        gstin = str(get("supplier_gstin") or "").strip().upper()
        if not _GSTIN_RE.match(gstin):
            gstin = ""
        doc_no = str(get("doc_no") or "").strip()
        # Skip blank / summary / footer-total rows (no GSTIN and no Doc No).
        if not gstin and not doc_no:
            continue

        doc_type_raw = str(get("doc_type") or "").upper()
        if "CREDIT" in doc_type_raw or "CN" in doc_type_raw:
            doc_type = "CRN"
        elif "DEBIT" in doc_type_raw or "DN" in doc_type_raw:
            doc_type = "DBN"
        else:
            doc_type = "INV"

        taxable = round_money(get("taxable_value"))
        igst = round_money(get("igst"))
        cgst = round_money(get("cgst"))
        sgst = round_money(get("sgst"))
        cess = round_money(get("cess"))
        invoice_value = round_money(get("invoice_value"))

        # OCTA pre-signs Credit Notes negative — force consistent CN sign without
        # double-negating (robust whether source value is signed or not).
        if doc_type == "CRN":
            taxable, igst, cgst, sgst, cess, invoice_value = (
                -abs(taxable), -abs(igst), -abs(cgst),
                -abs(sgst), -abs(cess), -abs(invoice_value),
            )

        # Enrich-only: OCTA already maps each supplier to its Tally ledger name
        # ("As per Tally Name", e.g. "SIMK LABELS PRIVATE LIMITED (Maharashtra)").
        # Use it as supplier_name so Pass-2 name similarity and the Vendor Summary
        # common name align with the Tally Books side. Matching itself is unchanged
        # (still keyed on invoice number). Portal name kept in raw for reference.
        portal_name = str(get("supplier_name") or "").strip()
        tally_name  = str(get("tally_name") or "").strip()
        supplier_name = tally_name or portal_name

        # Amendments: OCTA flags them inline (Is Amendment + Original Doc No) rather
        # than via separate B2BA/CDNRA sheets. Map onto the sheet labels Pass-5's
        # amendment detection expects so those remarks fire.
        is_amendment = str(get("is_amendment") or "").strip().lower() in ("yes", "true", "1")
        if doc_type == "CRN":
            sheet_label = "B2B-CDNRA" if is_amendment else "B2B-CDNR"
        else:
            sheet_label = "B2BA" if is_amendment else "B2B"

        index += 1
        rec = NormalizedInvoice(
            source="GSTR-2B",
            row_id=f"GSTR2B-OCTA-{index}",
            supplier_gstin=gstin,
            supplier_name=supplier_name,
            doc_type=doc_type,
            doc_no=doc_no,
            normalized_doc_no=normalize_doc_no(doc_no),
            doc_date=parse_date(get("doc_date")),
            taxable_value=taxable,
            igst=igst,
            cgst=cgst,
            sgst=sgst,
            cess=cess,
            invoice_value=invoice_value,
            sheet_name=sheet_label,
            raw={str(k): v for k, v in row.items() if not str(k).startswith("_")},
        )
        rec.raw["_tally_name"] = tally_name
        rec.raw["_octa_portal_name"] = portal_name
        records.append(rec)
    return records


def parse_gstr2b(data: bytes) -> list[NormalizedInvoice]:
    data = _ensure_xlsx(data)
    # OCTA flat-format export uses a single sheet with inline columns — detect and
    # route to its dedicated parser. The GST portal multi-sheet logic below is
    # left completely unchanged.
    if _is_octa_format(data):
        return _parse_gstr2b_octa(data)
    records: list[NormalizedInvoice] = []
    index = 0

    for sheet in ALLOWED_SHEETS_ORDER:
        try:
            rows = _read_gstr2b_sheet(data, sheet)
        except Exception:
            continue

        for row in rows:
            get = _build_getter(row, GSTR2B_COL_MAP)
            gstin = str(get("supplier_gstin") or "").strip().upper()
            # Fallback: scan all row values for a GSTIN-format string
            if not gstin or not _GSTIN_RE.match(gstin):
                gstin = ""
                for _v in row.values():
                    _vs = str(_v or "").strip().upper()
                    if _GSTIN_RE.match(_vs):
                        gstin = _vs
                        break
            doc_no = str(get("doc_no") or "").strip()
            if not gstin and not doc_no:
                continue

            doc_type_raw = str(get("doc_type") or "").upper()
            if "CREDIT" in doc_type_raw or "CN" in doc_type_raw:
                doc_type = "CRN"
            elif "DEBIT" in doc_type_raw or "DN" in doc_type_raw:
                doc_type = "DBN"
            else:
                doc_type = "INV"

            taxable = round_money(get("taxable_value"))
            igst = round_money(get("igst"))
            cgst = round_money(get("cgst"))
            sgst = round_money(get("sgst"))
            cess = round_money(get("cess"))
            invoice_value = round_money(get("invoice_value"))

            if doc_type == "CRN":
                taxable = -taxable
                igst = -igst
                cgst = -cgst
                sgst = -sgst
                cess = -cess
                invoice_value = -invoice_value

            index += 1
            records.append(NormalizedInvoice(
                source="GSTR-2B",
                row_id=f"GSTR2B-{sheet}-{index}",
                supplier_gstin=gstin,
                supplier_name=str(get("supplier_name") or "").strip(),
                doc_type=doc_type,
                doc_no=doc_no,
                normalized_doc_no=normalize_doc_no(doc_no),
                doc_date=parse_date(get("doc_date")),
                taxable_value=taxable,
                igst=igst,
                cgst=cgst,
                sgst=sgst,
                cess=cess,
                invoice_value=invoice_value,
                sheet_name=sheet,
                raw={str(k): v for k, v in row.items() if not str(k).startswith("_")},
            ))

    return records


# ---------------------------------------------------------------------------
# Books Parser (Purchase Register + Debit Note Register)
# Tally export: 6-row preamble, header on row 6, data from row 7
# ---------------------------------------------------------------------------

def _sum_cols(row: dict[str, Any], aliases: list[str]) -> float:
    norm_row = {_norm(k): v for k, v in row.items()}
    total = 0.0
    for alias in aliases:
        key = _norm(alias)
        if key in norm_row:
            total += round_money(norm_row[key])
    return round(total, 2)


def _is_tax_col(c: str) -> bool:
    c_lower = c.lower()
    if any(t in c_lower for t in ["cgst", "sgst", "igst", "utgst"]):
        return True
    if "cess" in c_lower and "process" not in c_lower:
        return True
    return False


def _is_tds_col(c: str) -> bool:
    return "tds" in c.lower()


def _is_round_off_col(c: str) -> bool:
    """True for a Tally 'Round Off' rounding-adjustment column.
    Collapses spaces so 'Round Off' / 'RoundOff' match; 'Ground Rent' etc. do not."""
    return "roundoff" in str(c).lower().replace(" ", "")


def _find_value_col(row: dict[str, Any]) -> str | None:
    """Return the source header that holds the explicit taxable value.

    Prefers an explicit "Taxable Value"/"Taxable Amount" column, otherwise an
    exact "Value" column. Never matches "Gross Total" (or any "gross" header) —
    Gross Total includes tax/round-off and must never be used as the taxable."""
    exact_value = None
    for k in row.keys():
        n = _norm(k)
        if "gross" in n:
            continue
        if n in ("taxable value", "taxable amount") or ("taxable" in n and "value" in n):
            return k
        if n == "value" and exact_value is None:
            exact_value = k
    return exact_value


def _is_present_number(cell: Any) -> bool:
    """True only when the cell holds an actual number (not blank/placeholder)."""
    if cell is None:
        return False
    s = str(cell).replace(",", "").replace("₹", "").strip()
    if s.lower() in ("", "-", "nan", "nat", "none", "null", "n/a"):
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


IGST_ALIASES = [
    "input igst dl", "input igst 18% gj", "input igst 18% kr", "input igst 18% hr",
    "input igst 18% mh", "input igst 18% rj", "input igst 18% tn", "input igst 18% tg",
    "input igst 18% up", "input igst 18% wb", "input igst 18% isd", "igst",
]
CGST_ALIASES = [
    "input cgst dl", "input cgst 9% gj", "input cgst 9% hr", "input cgst 9% kr",
    "input cgst 9% mp", "input cgst 9% mh", "input cgst 9% pj", "input cgst 9% rj",
    "input cgst 9% tn", "input cgst 9% tg", "input cgst 9% up", "input cgst 9% wb", "cgst",
]
SGST_ALIASES = [
    "input sgst dl", "input sgst 9% gj", "input sgst 9% hr", "input sgst 9% kr",
    "input sgst 9% mp", "input sgst 9% mh", "input sgst 9% pj", "input sgst 9% rj",
    "input sgst 9% tn", "input sgst 9% tg", "input sgst 9% up", "input sgst 9% wb",
    "sgst", "utgst",
]


def _read_tally_sheet(data: bytes) -> list[dict[str, Any]]:
    xl = pd.ExcelFile(BytesIO(data))
    sheet_name = xl.sheet_names[0]
    raw = xl.parse(sheet_name, header=None, dtype=object)
    if len(raw) < 2:
        return []

    # Find the header row dynamically by scanning the first 20 rows
    header_idx = -1
    for idx in range(min(20, len(raw))):
        row_str = [str(x).strip().lower() for x in raw.iloc[idx]]
        if any("date" in x for x in row_str) and any("particulars" in x for x in row_str):
            header_idx = idx
            break
            
    if header_idx == -1:
        # Fallback to row index 6 (row 7)
        header_idx = min(6, len(raw) - 1)

    # Make column names unique
    raw_headers = [str(v).strip() if str(v) != "nan" else f"_col{i}" for i, v in enumerate(raw.iloc[header_idx])]
    seen: dict[str, int] = {}
    headers = []
    for h in raw_headers:
        count = seen.get(h, 0)
        seen[h] = count + 1
        headers.append(f"{h}_{count}" if count else h)

    data_rows = raw.iloc[header_idx + 1:].copy()
    data_rows.columns = headers
    data_rows = data_rows.dropna(how="all")
    return data_rows.fillna("").to_dict(orient="records")


def parse_books(purchase_data: bytes, debit_data: bytes) -> list[NormalizedInvoice]:
    purchase_data = _ensure_xlsx(purchase_data)
    debit_data = _ensure_xlsx(debit_data)
    records: list[NormalizedInvoice] = []
    index = 0

    datasets = [
        (purchase_data, "Purchase Register", "INV"),
        (debit_data, "Debit Note Register", "DBN"),
    ]

    for data, source, default_type in datasets:
        try:
            rows = _read_tally_sheet(data)
        except Exception:
            continue

        for row in rows:
            # Skip if no date (summary rows)
            date_raw = row.get("Date", "")
            if not str(date_raw).strip() or str(date_raw).strip() in ("nan", ""):
                continue

            # Supplier Invoice No. is the matching key
            doc_no = ""
            for col in ["Supplier Invoice No.", "Voucher Ref. No.", "Voucher No."]:
                val = str(row.get(col, "") or "").strip()
                if val and val not in ("nan", ""):
                    doc_no = val
                    break

            # Date
            doc_date = ""
            for col in ["Supplier Invoice Date", "Voucher Ref. Date", "Date"]:
                val = row.get(col, "")
                parsed = parse_date(val)
                if parsed:
                    doc_date = parsed
                    break

            # Supplier name
            supplier_name = ""
            for col in ["Particulars", "Supplier", "Buyer"]:
                val = str(row.get(col, "") or "").strip()
                if val and val not in ("nan", ""):
                    supplier_name = val
                    break

            if not doc_no and not supplier_name:
                continue

            igst = 0.0
            cgst = 0.0
            sgst = 0.0
            cess = 0.0

            for k, v in row.items():
                k_lower = str(k).lower().strip()
                if "igst" in k_lower:
                    igst += round_money(v)
                elif "cgst" in k_lower:
                    cgst += round_money(v)
                elif "sgst" in k_lower or "utgst" in k_lower:
                    sgst += round_money(v)
                elif "cess" in k_lower and "process" not in k_lower:
                    cess += round_money(v)

            # ── Books taxable value ──────────────────────────────────────────
            # ALWAYS the sum of the purchase/expense ledger columns positioned
            # AFTER "Gross Total". The "Value" column is intentionally NOT used:
            # it is blank for no-GRN invoices and omits freight/penalty/etc.
            # Excluded from the sum:
            #   • GST input tax (Input IGST/CGST/SGST/UTGST/Cess) — _is_tax_col.
            #     "GST Interstate Purchase X%" is a PURCHASE ledger (no igst/cgst/
            #     sgst token) and is therefore correctly INCLUDED — it is the key
            #     amount on these files.
            #   • TDS columns                                      — _is_tds_col
            #   • Round Off (rounding adjustment, not an expense)  — _is_round_off_col
            gross = round_money(row.get("Gross Total", 0))
            invoice_value = gross

            taxable_breakdown: list[tuple[str, float]] = []
            keys = list(row.keys())
            gt_pos = next(
                (i for i, k in enumerate(keys) if _norm(k) == "gross total"),
                -1,
            )
            running = 0.0
            for k in keys[gt_pos + 1:]:
                ks = str(k)
                if _is_tax_col(ks) or _is_tds_col(ks) or _is_round_off_col(ks):
                    continue
                amt = round_money(row.get(k))
                if amt:
                    running += amt
                    taxable_breakdown.append((ks, round(amt, 2)))
            taxable = round(running, 2)
            taxable_derived = True

            if default_type == "DBN":
                taxable = -taxable
                igst = -igst
                cgst = -cgst
                sgst = -sgst
                cess = -cess
                invoice_value = -invoice_value
                taxable_breakdown = [(h, -a) for h, a in taxable_breakdown]

            index += 1
            rec = NormalizedInvoice(
                source=source,
                row_id=f"{source}-{index}",
                supplier_gstin="",           # Tally doesn't export GSTIN
                supplier_name=supplier_name,
                doc_type=default_type,
                doc_no=doc_no,
                normalized_doc_no=normalize_doc_no(doc_no),
                doc_date=doc_date,
                taxable_value=taxable,
                igst=igst,
                cgst=cgst,
                sgst=sgst,
                cess=cess,
                invoice_value=invoice_value,
                raw={str(k): v for k, v in row.items()},
            )
            # Display-only audit trail: when taxable was derived (no Value column),
            # remember which expense heads were summed so the output can show a
            # live formula + cell comment. Never affects the numeric value above.
            rec.raw["_taxable_derived"] = taxable_derived
            if taxable_derived:
                rec.raw["_taxable_breakdown"] = taxable_breakdown
            records.append(rec)

    return records


# ---------------------------------------------------------------------------
# Custom reconciliation (match by invoice number; GSTIN not available in books)
# ---------------------------------------------------------------------------

def _name_sim(a: str, b: str) -> float:
    """Simple word-overlap similarity between two supplier names."""
    wa = set(re.sub(r"[^a-z0-9]", " ", a.lower()).split())
    wb = set(re.sub(r"[^a-z0-9]", " ", b.lower()).split())
    if not wa or not wb:
        return 0.0
    return len(wa & wb) / max(len(wa), len(wb))


def _core_digits(doc_no: str) -> str:
    """Extract the core numeric/alpha digits stripping common year suffixes (2025/2026)."""
    s = re.sub(r"(2025|2026)$", "", doc_no.upper())
    return re.sub(r"[^0-9A-Z]", "", s)


_CN_PREFIX_RE = re.compile(r'^(?:CDNR|CDN|LCN|CRN|CN|CR|CD)', re.IGNORECASE)

def _looks_like_cn(doc_no: str) -> bool:
    """True if doc_no starts with a common credit-note prefix (CDN, CN, LCN, CR, CRN…)."""
    return bool(_CN_PREFIX_RE.match(re.sub(r"[^A-Za-z0-9]", "", doc_no)))


def reconcile_by_invoice_no(
    gstr2b: list[NormalizedInvoice],
    books: list[NormalizedInvoice],
    tolerance: float = 1.0,
) -> list[MatchResult]:
    """
    Match GSTR-2B records to Books records using normalised invoice number as the
    primary key. GSTIN is not checked because Tally registers don't contain it.
    Falls back to supplier name + core-digit matching for minor year-suffix differences.
    """
    # Build books indexes
    books_by_no: dict[str, list[int]] = {}
    books_by_core: dict[str, list[int]] = {}
    for idx, rec in enumerate(books):
        books_by_no.setdefault(rec.normalized_doc_no, []).append(idx)
        books_by_core.setdefault(_core_digits(rec.normalized_doc_no), []).append(idx)

    results: list[MatchResult] = []
    used_books: set[int] = set()
    used_gstr2b: set[int] = set()

    def _make_result(g_rec, b_rec, category_override=None):
        mismatches: list[str] = []
        for field in ("taxable_value", "igst", "cgst", "sgst"):
            if not amount_equal(getattr(g_rec, field), getattr(b_rec, field), tolerance):
                mismatches.append(field)
        if category_override:
            category = category_override
            confidence = 70
            action = "Review — invoice number year suffix differs between GSTR-2B and Books."
            explanation = f"Matched by supplier name and core invoice digits; year suffix differs."
        elif not mismatches:
            category = "Matched"
            confidence = 100
            action = "Claim ITC if otherwise eligible."
            explanation = "Invoice number matches and all amounts are within tolerance."
        else:
            category = "Amount Mismatch"
            confidence = 75
            action = "Review amount differences before claiming ITC."
            explanation = f"Invoice number matches but {', '.join(mismatches)} differs."
        return MatchResult(
            category=category,
            confidence=confidence,
            gstr2b=g_rec,
            purchase=b_rec,
            mismatch_fields=mismatches,
            suggested_action=action,
            explanation=explanation,
        )

    # Pass 1 — exact normalised invoice number match
    for g_idx, g_rec in enumerate(gstr2b):
        candidates = [i for i in books_by_no.get(g_rec.normalized_doc_no, []) if i not in used_books]
        if not candidates:
            continue
        best_idx = candidates[0]
        used_gstr2b.add(g_idx)
        used_books.add(best_idx)
        results.append(_make_result(g_rec, books[best_idx]))

    # Pass 2 — fuzzy fallback: core digits + supplier name similarity >= 0.5
    for g_idx, g_rec in enumerate(gstr2b):
        if g_idx in used_gstr2b:
            continue
        g_core = _core_digits(g_rec.normalized_doc_no)
        candidates = [i for i in books_by_core.get(g_core, []) if i not in used_books]
        if not candidates:
            continue
        # Pick best name similarity
        best_idx = max(candidates, key=lambda i: _name_sim(g_rec.supplier_name, books[i].supplier_name))
        if _name_sim(g_rec.supplier_name, books[best_idx].supplier_name) >= 0.5:
            used_gstr2b.add(g_idx)
            used_books.add(best_idx)
            results.append(_make_result(g_rec, books[best_idx], category_override="Matched"))

    # Pass 2.5 — CN (GSTR-2B) ↔ DN (Books) cross-type match
    # A supplier's Credit Note number (e.g. LCN-38/24-25) and the buyer's Debit Note
    # number (e.g. 44/24-25) are completely different strings — Passes 1 & 2 cannot link them.
    # Only attempt when the GSTR-2B doc_no has a recognisable CN prefix to avoid false positives.
    for g_idx, g_rec in enumerate(gstr2b):
        if g_idx in used_gstr2b:
            continue
        if g_rec.doc_type != "CRN" or not _looks_like_cn(g_rec.doc_no):
            continue
        best_b, best_sim = None, 0.0
        for b_idx, b_rec in enumerate(books):
            if b_idx in used_books or b_rec.doc_type != "DBN":
                continue
            sim = _name_sim(g_rec.supplier_name, b_rec.supplier_name)
            if sim >= 0.5 and abs(g_rec.taxable_value - b_rec.taxable_value) <= tolerance:
                if sim > best_sim:
                    best_b, best_sim = b_idx, sim
        if best_b is not None:
            used_gstr2b.add(g_idx)
            used_books.add(best_b)
            results.append(_make_result(g_rec, books[best_b], category_override="Partially Matched"))

    # Pass 3 — GSTR-2B records with no match in books
    for g_idx, g_rec in enumerate(gstr2b):
        if g_idx not in used_gstr2b:
            results.append(MatchResult(
                category="In GSTR-2B not in Books",
                confidence=0,
                gstr2b=g_rec,
                purchase=None,
                mismatch_fields=["missing_in_books"],
                suggested_action="In 2B but not in the books",
                explanation="Invoice present in GSTR-2B but not found in Purchase/Debit Note Register.",
            ))

    # Pass 4 — Books records with no match in GSTR-2B
    for b_idx, b_rec in enumerate(books):
        if b_idx not in used_books:
            results.append(MatchResult(
                category="In Books not in GSTR-2B",
                confidence=0,
                gstr2b=None,
                purchase=b_rec,
                mismatch_fields=["missing_in_gstr2b"],
                suggested_action="Need to 2B",
                explanation="Invoice present in Books but not found in GSTR-2B.",
            ))

    # Pass 5 — Post-processing Calibration & Remark Overhaul

    # ── Amendment detection sets ─────────────────────────────────────────────
    # Two cases both need flagging:
    #   Case A: B2BA unmatched + B2B has same invoice (B2B matched with Books)
    #           → flag the B2BA row: "Amendment entry in B2BA"
    #   Case B: B2B unmatched + B2BA has same invoice (B2BA matched with Books)
    #           → flag the B2B row: "Original invoice amended in B2BA"
    # Same logic applies for B2B-CDNR / B2B-CDNRA.
    b2b_orig_nos  = {r.normalized_doc_no for r in gstr2b
                     if r.sheet_name == "B2B"       and r.normalized_doc_no}
    cdnr_orig_nos = {r.normalized_doc_no for r in gstr2b
                     if r.sheet_name == "B2B-CDNR"  and r.normalized_doc_no}
    b2ba_nos      = {r.normalized_doc_no for r in gstr2b
                     if r.sheet_name == "B2BA"       and r.normalized_doc_no}
    cdnra_nos     = {r.normalized_doc_no for r in gstr2b
                     if r.sheet_name == "B2B-CDNRA"  and r.normalized_doc_no}

    # ── Duplicate-type maps keyed by normalized_doc_no ───────────────────────
    # "dup"              = same doc_no, same sign, same-ish party → true duplicate
    # "same_no_diff_party" = same doc_no, same sign, different parties → different vendors, same number
    # "cross_ref"        = same doc_no with OPPOSITE signs → INV+CRN or PR+DN pair
    # ""                 = unique

    def _calc_dup_type_enhanced(records: list) -> str:
        """records: list of (taxable_value, supplier_name)"""
        if len(records) <= 1:
            return ""
        vals  = [r[0] for r in records]
        names = [r[1] for r in records]
        all_pos = all(v >= 0 for v in vals)
        all_neg = all(v <= 0 for v in vals)
        if all_pos or all_neg:
            any_diff = any(
                _name_sim(names[i], names[j]) < 0.5
                for i in range(len(names))
                for j in range(i + 1, len(names))
            )
            return "same_no_diff_party" if any_diff else "dup"
        return "cross_ref"

    g_key_records: dict = {}
    for g_rec in gstr2b:
        if g_rec.normalized_doc_no:
            g_key_records.setdefault(g_rec.normalized_doc_no, []).append(
                (g_rec.taxable_value, g_rec.supplier_name or "")
            )
    g_key_dup_type: dict = {k: _calc_dup_type_enhanced(v) for k, v in g_key_records.items()}

    b_key_records: dict = {}
    for b_rec in books:
        if b_rec.normalized_doc_no:
            b_key_records.setdefault(b_rec.normalized_doc_no, []).append(
                (b_rec.taxable_value, b_rec.supplier_name or "")
            )
    b_key_dup_type: dict = {k: _calc_dup_type_enhanced(v) for k, v in b_key_records.items()}

    for r in results:
        g = r.gstr2b
        b = r.purchase

        g_dup_type = g_key_dup_type.get(g.normalized_doc_no, "") if g else ""
        b_dup_type = b_key_dup_type.get(b.normalized_doc_no, "") if b else ""

        r.suggested_action_2 = ""

        # Scenario A: BOTH GSTR-2B and Books exist
        if g and b:
            g_tax = round(g.igst + g.cgst + g.sgst + g.cess, 2)
            b_tax = round(b.igst + b.cgst + b.sgst + b.cess, 2)

            sec_remarks = []

            if _get_val(g, "doc_type", "") == "CRN" and _get_val(b, "doc_type", "") == "DBN":
                r.suggested_action = "Partially Matched"
                r.category = "Partially Matched"
                sec_remarks.insert(0,
                    f"CN-DN Match: {_get_val(g, 'doc_no', '')} ↔ {_get_val(b, 'doc_no', '')}"
                )
            else:
                r.suggested_action = "Matched"

            if g_tax - b_tax > 1.0:
                r.category = "Amount Mismatch"
                sec_remarks.append("Tax Amount Mismatch, Excess in 2B")
            elif b_tax - g_tax > 1.0:
                r.category = "Amount Mismatch"
                sec_remarks.append("Tax Amount Mismatch, Excess in Books")

            if g.taxable_value - b.taxable_value > 1.0:
                r.category = "Amount Mismatch"
                sec_remarks.append("Taxable Value Mismatch, Excess in 2B")
            elif b.taxable_value - g.taxable_value > 1.0:
                r.category = "Amount Mismatch"
                sec_remarks.append("Taxable Value Mismatch, Excess in Books")

            from recon.core import parse_date
            if parse_date(g.doc_date) != parse_date(b.doc_date):
                sec_remarks.append("Invoice Date Mismatch")

            # Books duplicate/cross-ref
            if b_dup_type == "dup":
                sec_remarks.append("Invoice appears twice in Books — verify if duplicate booking")
            elif b_dup_type == "same_no_diff_party":
                sec_remarks.append("Same voucher number in Books for different parties — verify")
            elif b_dup_type == "cross_ref":
                sec_remarks.append("Invoice also referenced in DN Register")

            # 2B duplicate/cross-ref (independent — both can appear)
            if g_dup_type == "dup":
                sec_remarks.append("Invoice appears twice in GSTR-2B — possible duplicate upload by vendor")
            elif g_dup_type == "same_no_diff_party":
                sec_remarks.append("Same invoice number in GSTR-2B for different parties — verify")
            elif g_dup_type == "cross_ref":
                sec_remarks.append("Same invoice number also exists as Credit Note in GSTR-2B")

            r.suggested_action_2 = ", ".join(sec_remarks) if sec_remarks else ""

        # Scenario B: In GSTR-2B but NOT in Books
        elif g and not b:
            r.suggested_action = "Showing in 2B but Not in Books"

            sec_remarks = []

            # RCM check
            is_rcm = any(
                "reverse charge" in k.lower() and str(v).strip().lower() == "yes"
                for k, v in g.raw.items()
            )
            if is_rcm:
                sec_remarks.append("RCM")

            # Amendment detection — covers both directions:
            # Case A: amendment sheet (B2BA/CDNRA) is unmatched, original (B2B/CDNR) matched Books
            # Case B: original sheet (B2B/CDNR) is unmatched, amendment (B2BA/CDNRA) matched Books
            # When amendment is flagged, skip the dup remark — dup is a false positive here.
            is_amendment = False
            if g.sheet_name == "B2BA" and g.normalized_doc_no in b2b_orig_nos:
                sec_remarks.append(
                    f"Amendment entry in B2BA — original invoice {g.doc_no} already present in B2B"
                )
                is_amendment = True
            elif g.sheet_name == "B2B-CDNRA" and g.normalized_doc_no in cdnr_orig_nos:
                sec_remarks.append(
                    f"Amendment entry in B2B-CDNRA — original note {g.doc_no} already present in B2B-CDNR"
                )
                is_amendment = True
            elif g.sheet_name == "B2B" and g.normalized_doc_no in b2ba_nos:
                sec_remarks.append(
                    f"Original invoice {g.doc_no} amended in B2BA — amendment entry also present in 2B"
                )
                is_amendment = True
            elif g.sheet_name == "B2B-CDNR" and g.normalized_doc_no in cdnra_nos:
                sec_remarks.append(
                    f"Original note {g.doc_no} amended in B2B-CDNRA — amendment entry also present in 2B"
                )
                is_amendment = True

            if not is_amendment:
                if g_dup_type == "dup":
                    sec_remarks.append("Invoice appears twice in GSTR-2B — possible duplicate upload by vendor")
                elif g_dup_type == "same_no_diff_party":
                    sec_remarks.append("Same invoice number in GSTR-2B for different parties — verify")
                elif g_dup_type == "cross_ref":
                    sec_remarks.append("Same invoice number also exists as Credit Note in GSTR-2B")

            r.suggested_action_2 = ", ".join(sec_remarks) if sec_remarks else ""

        # Scenario C: In Books but NOT in GSTR-2B
        elif b and not g:
            r.suggested_action = "Showing in Books but Not in 2B"

            sec_remarks_c = []
            if b_dup_type == "dup":
                sec_remarks_c.append("Invoice appears twice in Books — verify if duplicate booking")
            elif b_dup_type == "same_no_diff_party":
                sec_remarks_c.append("Same voucher number in Books for different parties — verify")
            elif b_dup_type == "cross_ref":
                sec_remarks_c.append("Invoice also referenced in DN Register")
            r.suggested_action_2 = ", ".join(sec_remarks_c) if sec_remarks_c else ""

    return results


# ---------------------------------------------------------------------------
# Excel export in SOP format
# ---------------------------------------------------------------------------

def _get_val(obj: Any, key: str, default: Any = "") -> Any:
    if obj is None:
        return default
    if isinstance(obj, dict):
        return obj.get(key, default)
    return getattr(obj, key, default)


def _populate_workbook(ws, results: list[Any]) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Styles
    header_font = Font(bold=True, size=10)
    section_font = Font(bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_2b = PatternFill("solid", fgColor="DCE6F1")    # light blue
    fill_books = PatternFill("solid", fgColor="E2EFDA")  # light green
    fill_diff = PatternFill("solid", fgColor="FCE4D6")   # light orange
    fill_matched = PatternFill("solid", fgColor="E2EFDA")
    fill_only_2b = PatternFill("solid", fgColor="FFF2CC")
    fill_only_books = PatternFill("solid", fgColor="FCE4D6")
    fill_mismatch = PatternFill("solid", fgColor="F4B942")
    fill_partial  = PatternFill("solid", fgColor="D6EAF8")  # light blue — CN↔DN partial match

    # Scan for dup types: "dup" (same sign), "cross_ref" (opposite signs), "" (unique)
    def _wb_dup_type(vals: list) -> str:
        if len(vals) <= 1:
            return ""
        return "dup" if (all(v >= 0 for v in vals) or all(v <= 0 for v in vals)) else "cross_ref"

    gstr2b_doc_vals_wb: dict = {}
    books_doc_vals_wb: dict = {}
    for r in results:
        g = _get_val(r, "gstr2b", None)
        b = _get_val(r, "purchase", None)
        g_doc = str(_get_val(g, "doc_no", "")).strip().upper()
        b_doc = str(_get_val(b, "doc_no", "")).strip().upper()
        if g_doc and g_doc not in ("NAN", ""):
            gstr2b_doc_vals_wb.setdefault(g_doc, []).append(float(_get_val(g, "taxable_value", 0) or 0))
        if b_doc and b_doc not in ("NAN", ""):
            books_doc_vals_wb.setdefault(b_doc, []).append(float(_get_val(b, "taxable_value", 0) or 0))

    gstr2b_doc_dup_type = {d: _wb_dup_type(v) for d, v in gstr2b_doc_vals_wb.items()}
    books_doc_dup_type  = {d: _wb_dup_type(v) for d, v in books_doc_vals_wb.items()}

    fill_duplicate  = PatternFill("solid", fgColor="FFC7CE")  # soft red   — true duplicate
    font_duplicate  = Font(color="9C0006", bold=True)         # dark red bold
    fill_cross_ref  = PatternFill("solid", fgColor="FFE0B2")  # soft orange — cross-reference
    font_cross_ref  = Font(color="E65100", bold=True)         # dark orange bold

    col_headers = [
        # GSTR-2B (cols A-J)
        "Party Name (GSTR-2B)", "GSTIN (2B)", "Invoice Number", "Invoice Date",
        "Taxable Value (₹)", "Central Tax (₹)", "State/UT Tax (₹)", "Integrated Tax (₹)",
        "Source Tab (2B)", "State (2B)",
        "",  # spacer col K
        # Books (cols L-U)
        "Particulars", "GSTIN (Books)", "Voucher No.", "Date",
        "Taxable Value (₹)", "Central Tax (₹)", "State/UT Tax (₹)", "Integrated Tax (₹)",
        "Source (Books)", "State (Books)",
        "",  # spacer col V
        # Difference (cols W-Z)
        "Taxable Value (₹)", "Central Tax (₹)", "State/UT Tax (₹)", "Integrated Tax (₹)",
        "Remark 1",   # col AA (27)
        "Remark 2",   # col AB (28)
    ]

    # Row 1: blank
    ws.append([""])

    # Row 2: section headers
    ws.append([
        "As Per GSTR 2B Data", "", "", "", "", "", "", "", "", "",
        "",
        "As Per Books Data", "", "", "", "", "", "", "", "", "",
        "",
        "Difference", "", "", "",
        "", "",
    ])
    ws.merge_cells("A2:J2")
    ws.merge_cells("L2:U2")
    ws.merge_cells("W2:Z2")
    for cell_addr in ["A2", "L2", "W2"]:
        c = ws[cell_addr]
        c.font = section_font
        c.alignment = center
    ws["A2"].fill = fill_2b
    ws["L2"].fill = fill_books
    ws["W2"].fill = fill_diff

    # Row 3: column headers
    ws.append(col_headers)
    for col_idx, _ in enumerate(col_headers, start=1):
        c = ws.cell(row=3, column=col_idx)
        c.font = header_font
        c.alignment = center
        c.border = border
        if col_idx in range(1, 11):       # A-J: 2B section
            c.fill = fill_2b
        elif col_idx in range(12, 22):    # L-U: Books section
            c.fill = fill_books
        elif col_idx in range(23, 29):    # W-Z + AA-AB: Diff + Remarks
            c.fill = fill_diff

    # Data rows
    for result in results:
        g = _get_val(result, "gstr2b", None)
        b = _get_val(result, "purchase", None)
        category = _get_val(result, "category", "")

        # Determine remarks
        remark_1 = _get_val(result, "suggested_action", "")
        remark_2 = _get_val(result, "suggested_action_2", "")

        g_taxable = _get_val(g, "taxable_value", 0)
        g_cgst    = _get_val(g, "cgst", 0)
        g_sgst    = _get_val(g, "sgst", 0)
        g_igst    = _get_val(g, "igst", 0)
        b_taxable = _get_val(b, "taxable_value", 0)
        b_cgst    = _get_val(b, "cgst", 0)
        b_sgst    = _get_val(b, "sgst", 0)
        b_igst    = _get_val(b, "igst", 0)

        diff_taxable = round(g_taxable - b_taxable, 2)
        diff_cgst    = round(g_cgst    - b_cgst,    2)
        diff_sgst    = round(g_sgst    - b_sgst,    2)
        diff_igst    = round(g_igst    - b_igst,    2)

        # New: GSTIN, Source Tab, State for 2B
        gstin_2b   = str(_get_val(g, "supplier_gstin", "") or "")
        source_tab = str(_get_val(g, "sheet_name",     "") or "")
        raw_g      = _get_val(g, "raw", {}) or {}
        entity_g   = str(raw_g.get("_entity_gstin", "") or "")
        state_2b   = GST_STATE_CODES.get(entity_g[:2], "") if entity_g else ""

        # New: GSTIN, Source, State for Books
        # Tally has no GSTIN column — use enriched value, fall back to 2B GSTIN for matched rows
        gstin_bks    = str(_get_val(b, "supplier_gstin", "") or "") or str(_get_val(g, "supplier_gstin", "") or "")
        src_bks_full = str(_get_val(b, "source",          "") or "")
        source_bks   = "PR" if "Purchase" in src_bks_full else ("DN" if "Debit" in src_bks_full else "")
        raw_b        = _get_val(b, "raw", {}) or {}
        entity_b     = str(raw_b.get("_entity_gstin", "") or "")
        state_bks    = GST_STATE_CODES.get(entity_b[:2], "") if entity_b else ""

        row = [
            _get_val(g, "supplier_name", ""),
            gstin_2b,
            _get_val(g, "doc_no", ""),
            _get_val(g, "doc_date", ""),
            g_taxable, g_cgst, g_sgst, g_igst,
            source_tab, state_2b,
            "",  # spacer K
            _get_val(b, "supplier_name", ""),
            gstin_bks,
            _get_val(b, "doc_no", ""),
            _get_val(b, "doc_date", ""),
            b_taxable, b_cgst, b_sgst, b_igst,
            source_bks, state_bks,
            "",  # spacer V
            diff_taxable, diff_cgst, diff_sgst, diff_igst,
            remark_1,
            remark_2,
        ]
        ws.append(row)

        # Row fill based on category
        row_idx = ws.max_row
        if category == "Matched":
            fill = fill_matched
        elif category == "Partially Matched":
            fill = fill_partial
        elif category == "Amount Mismatch":
            fill = fill_mismatch
        elif category == "In GSTR-2B not in Books" or category == "In GSTR-2B not in PR":
            fill = fill_only_2b
        else:
            fill = fill_only_books

        for col_idx in range(1, len(col_headers) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            if c.value != "" and col_idx != 11 and col_idx != 22:  # spacers K(11) and V(22)
                c.fill = fill
            c.border = border

        # Books Taxable (col P = 16): when the value was DERIVED (no source Value
        # column), replace the static number with a live Excel formula summing the
        # expense heads + a hover comment naming each head, so it's auditable.
        raw_b_disp = _get_val(b, "raw", {}) or {}
        if raw_b_disp.get("_taxable_derived") and raw_b_disp.get("_taxable_breakdown"):
            from openpyxl.comments import Comment
            breakdown = raw_b_disp["_taxable_breakdown"]
            note = " + ".join(f"{head} {amt:,.2f}" for head, amt in breakdown)
            # Keep the NUMERIC taxable already written by ws.append — an Excel formula
            # here renders blank/0 in viewers that don't auto-recalc (e.g. Numbers),
            # which made the whole column look empty. Show the make-up as a comment.
            ws.cell(row=row_idx, column=16).comment = Comment(f"= {note}", "Colonel")

        # Highlight GSTR-2B duplicates/cross-refs — Invoice Number is now col C (3)
        g_doc = str(_get_val(g, "doc_no", "")).strip().upper()
        g_dt = gstr2b_doc_dup_type.get(g_doc, "")
        if g_doc and g_doc not in ("NAN", "") and g_dt:
            cell_g = ws.cell(row=row_idx, column=3)
            if g_dt == "dup":
                cell_g.fill = fill_duplicate
                cell_g.font = font_duplicate
            else:
                cell_g.fill = fill_cross_ref
                cell_g.font = font_cross_ref

        # Highlight Books duplicates/cross-refs — Voucher No. is now col N (14)
        b_doc = str(_get_val(b, "doc_no", "")).strip().upper()
        b_dt = books_doc_dup_type.get(b_doc, "")
        if b_doc and b_doc not in ("NAN", "") and b_dt:
            cell_b = ws.cell(row=row_idx, column=14)
            if b_dt == "dup":
                cell_b.fill = fill_duplicate
                cell_b.font = font_duplicate
            else:
                cell_b.fill = fill_cross_ref
                cell_b.font = font_cross_ref

    # Column widths
    # A-J (2B): name, gstin, inv_no, date, taxable, cgst, sgst, igst, source_tab, state
    # K: spacer
    # L-U (Books): name, gstin, voucher, date, taxable, cgst, sgst, igst, source, state
    # V: spacer
    # W-Z (Diff) + AA-AB (Remarks)
    widths = [
        30, 20, 20, 14, 16, 14, 14, 16, 12, 14,   # A-J
        2,                                          # K spacer
        30, 20, 20, 14, 16, 14, 14, 16,  6, 14,   # L-U
        2,                                          # V spacer
        16, 14, 14, 16, 22, 22,                     # W-Z + AA-AB
    ]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"


def _fix_company_name(raw: str) -> str:
    """Title-case a vendor name and restore common Indian company abbreviations."""
    if not raw:
        return raw
    result = raw.title()
    for wrong, right in [
        ("Llp", "LLP"), ("Llc", "LLC"), (" Lp ", " LP "),
        ("Pvt.Ltd.", "Pvt Ltd"), ("Pvt.Ltd", "Pvt Ltd"), ("P.Ltd.", "P Ltd"),
        ("Ltd.", "Ltd"), (" Of ", " of "), (" And ", " and "),
        (" The ", " the "), (" A ", " a "), (" An ", " an "),
        (" For ", " for "), ("Isd", "ISD"), ("Gst", "GST"),
    ]:
        result = result.replace(wrong, right)
    return result


def _build_vendor_summary_sheet(ws, results: list[Any]) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin        = Side(style="thin")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right",  vertical="center")

    title_fill   = PatternFill("solid", fgColor="1F3864")
    title_font   = Font(bold=True, size=12, color="FFFFFF")
    section_font = Font(bold=True, size=11)
    header_font  = Font(bold=True, size=10)

    fill_2b          = PatternFill("solid", fgColor="DCE6F1")
    fill_books_hdr   = PatternFill("solid", fgColor="E2EFDA")
    fill_diff_hdr    = PatternFill("solid", fgColor="FCE4D6")
    fill_matched     = PatternFill("solid", fgColor="E2EFDA")
    fill_mismatch    = PatternFill("solid", fgColor="F4B942")
    fill_only_2b     = PatternFill("solid", fgColor="FFF2CC")
    fill_only_books  = PatternFill("solid", fgColor="FCE4D6")

    TOL     = 1.0
    NUM_COLS = 20
    # Column index map (1-based):
    # A=1 Common Name, B=2 GSTIN
    # C=3 Name 2B, D=4 Taxable 2B, E=5 CGST 2B, F=6 SGST 2B, G=7 IGST 2B
    # H=8 spacer
    # I=9 Name Books, J=10 Taxable Books, K=11 CGST Books, L=12 SGST Books, M=13 IGST Books
    # N=14 spacer
    # O=15 Diff Taxable, P=16 Diff CGST, Q=17 Diff SGST, R=18 Diff IGST
    # S=19 Remark 1, T=20 Remark 2
    SPACERS      = {8, 14}
    NUMERIC_COLS = set(range(4, 8)) | set(range(10, 14)) | set(range(15, 19))

    # ── Step 0: build name→GSTIN lookup from all GSTR-2B records ────────────
    # Uses _extract_gstin() which scans the raw row dict for the GSTIN pattern,
    # so it works even when supplier_gstin wasn't set by the column-name lookup.
    name_to_gstin_map: dict[str, str] = {}
    for _r in results:
        _g = _get_val(_r, "gstr2b", None)
        if _g:
            _gstin = _extract_gstin(_g)
            _name  = str(_get_val(_g, "supplier_name", "") or "").strip().lower()
            if _gstin and _name and _name not in name_to_gstin_map:
                name_to_gstin_map[_name] = _gstin

    # ── Step 1: aggregate by GSTIN ───────────────────────────────────────────
    vendor_groups: dict[str, dict] = {}

    for r in results:
        g = _get_val(r, "gstr2b",    None)
        b = _get_val(r, "purchase",  None)

        # Use _extract_gstin: checks supplier_gstin attr first, then scans raw dict
        gstin = _extract_gstin(g) if g else ""
        if not gstin and b:
            gstin = _extract_gstin(b)

        # If GSTIN still missing, look it up by supplier name from 2B records
        if not gstin:
            _g_name = str(_get_val(g, "supplier_name", "") or "").strip().lower() if g else ""
            _b_name = str(_get_val(b, "supplier_name", "") or "").strip().lower() if b else ""
            gstin = name_to_gstin_map.get(_g_name, "") or name_to_gstin_map.get(_b_name, "")

        fallback_key = ""
        if not gstin:
            if g:
                fallback_key = str(_get_val(g, "supplier_name", "") or "").strip().upper()
            if not fallback_key and b:
                fallback_key = str(_get_val(b, "supplier_name", "") or "").strip().upper()
            if not fallback_key:
                continue

        group_key = gstin if gstin else f"__name__{fallback_key}"

        if group_key not in vendor_groups:
            vendor_groups[group_key] = {
                "gstin":         gstin,
                "name_2b":       "",    "name_books":    "",
                "taxable_2b":    0.0,   "taxable_books": 0.0,
                "cgst_2b":       0.0,   "cgst_books":    0.0,
                "sgst_2b":       0.0,   "sgst_books":    0.0,
                "igst_2b":       0.0,   "igst_books":    0.0,
                "has_2b":        False, "has_books":     False,
                "sec_tags":      set(),
            }
        else:
            # If this group was created earlier with an empty GSTIN (name-based key)
            # but we now have a GSTIN, upgrade it.
            if gstin and not vendor_groups[group_key]["gstin"]:
                vendor_groups[group_key]["gstin"] = gstin

        vg = vendor_groups[group_key]

        try:
            if g:
                n = str(_get_val(g, "supplier_name", "") or "").strip()
                if n and not vg["name_2b"]:
                    vg["name_2b"] = n
                vg["taxable_2b"] += float(_get_val(g, "taxable_value", 0) or 0)
                vg["cgst_2b"]    += float(_get_val(g, "cgst",          0) or 0)
                vg["sgst_2b"]    += float(_get_val(g, "sgst",          0) or 0)
                vg["igst_2b"]    += float(_get_val(g, "igst",          0) or 0)
                vg["has_2b"]      = True

            if b:
                n = str(_get_val(b, "supplier_name", "") or "").strip()
                if n and not vg["name_books"]:
                    vg["name_books"] = n
                vg["taxable_books"] += float(_get_val(b, "taxable_value", 0) or 0)
                vg["cgst_books"]    += float(_get_val(b, "cgst",          0) or 0)
                vg["sgst_books"]    += float(_get_val(b, "sgst",          0) or 0)
                vg["igst_books"]    += float(_get_val(b, "igst",          0) or 0)
                vg["has_books"]      = True

            r2_str = str(_get_val(r, "suggested_action_2", "") or "").strip()
            for tag in r2_str.split(","):
                tag = tag.strip()
                if tag:
                    vg["sec_tags"].add(tag)
        except Exception as _agg_err:
            print(f"Vendor Summary aggregation error for {group_key}: {_agg_err}")

    # ── Step 2: build sorted row lists ──────────────────────────────────────
    def _fmt(val, present):
        return round(val, 2) if present else ""

    def _common_name(vg):
        # Prefer Books name (Tally exports are already properly cased)
        base = vg["name_books"] if vg["name_books"] else _fix_company_name(vg["name_2b"])
        gstin = vg["gstin"]
        if gstin and len(gstin) >= 2:
            state = GST_STATE_CODES.get(gstin[:2], "")
            if state:
                base = f"{base} ({state})"
        return base

    matched_rows    = []
    only_2b_rows    = []
    only_books_rows = []
    no_gstin_rows   = []

    for group_key, vg in vendor_groups.items():
        has_2b    = vg["has_2b"]
        has_books = vg["has_books"]
        gstin     = vg["gstin"]
        both      = has_2b and has_books

        # Remark 2 — compute mismatches first so Remark 1 can reflect them
        sec = []
        if both:
            tax_2b    = round(vg["igst_2b"]   + vg["cgst_2b"]    + vg["sgst_2b"],    2)
            tax_books = round(vg["igst_books"] + vg["cgst_books"] + vg["sgst_books"], 2)
            if   tax_2b    - tax_books  > TOL:
                sec.append("Tax Amount Mismatch, Excess in 2B")
            elif tax_books - tax_2b     > TOL:
                sec.append("Tax Amount Mismatch, Excess in Books")
            tv_diff = round(vg["taxable_2b"] - vg["taxable_books"], 2)
            if   tv_diff  > TOL:
                sec.append("Taxable Value Mismatch, Excess in 2B")
            elif -tv_diff > TOL:
                sec.append("Taxable Value Mismatch, Excess in Books")
        else:
            for tag in sorted(vg["sec_tags"]):
                if "rcm" in tag.lower() or "duplicate" in tag.lower():
                    sec.append(tag)

        # Remark 1 — reflects actual match quality at vendor level
        if both and sec:
            remark_1 = "Amount Mismatch"
        elif both:
            remark_1 = "Matched"
        elif has_2b:
            remark_1 = "Showing in 2B but Not in Books"
        else:
            remark_1 = "Showing in Books but Not in 2B"

        rd = {
            "common_name":   _common_name(vg),
            "gstin":         gstin,
            "name_2b":       vg["name_2b"],
            "taxable_2b":    _fmt(vg["taxable_2b"],    has_2b),
            "cgst_2b":       _fmt(vg["cgst_2b"],       has_2b),
            "sgst_2b":       _fmt(vg["sgst_2b"],       has_2b),
            "igst_2b":       _fmt(vg["igst_2b"],       has_2b),
            "name_books":    vg["name_books"],
            "taxable_books": _fmt(vg["taxable_books"], has_books),
            "cgst_books":    _fmt(vg["cgst_books"],    has_books),
            "sgst_books":    _fmt(vg["sgst_books"],    has_books),
            "igst_books":    _fmt(vg["igst_books"],    has_books),
            "diff_taxable":  _fmt(round(vg["taxable_2b"] - vg["taxable_books"], 2), both),
            "diff_cgst":     _fmt(round(vg["cgst_2b"]    - vg["cgst_books"],    2), both),
            "diff_sgst":     _fmt(round(vg["sgst_2b"]    - vg["sgst_books"],    2), both),
            "diff_igst":     _fmt(round(vg["igst_2b"]    - vg["igst_books"],    2), both),
            "remark_1":      remark_1,
            "remark_2":      ", ".join(sec),
        }

        if remark_1 in ("Matched", "Amount Mismatch"):
            matched_rows.append(rd)
        elif remark_1 == "Showing in 2B but Not in Books":
            only_2b_rows.append(rd)
        elif remark_1 == "Showing in Books but Not in 2B":
            only_books_rows.append(rd)
        else:
            no_gstin_rows.append(rd)

    for grp in (matched_rows, only_2b_rows, only_books_rows, no_gstin_rows):
        grp.sort(key=lambda x: x["common_name"].lower())

    all_rows = matched_rows + only_2b_rows + only_books_rows + no_gstin_rows

    # ── Step 3: write worksheet ───────────────────────────────────────────────

    # Row 1: title
    ws.append(["DETAILED PARTY NAME MAPPING"] + [""] * (NUM_COLS - 1))
    ws.merge_cells(f"A1:{get_column_letter(NUM_COLS)}1")
    tc = ws["A1"]
    tc.fill      = title_fill
    tc.font      = title_font
    tc.alignment = center
    ws.row_dimensions[1].height = 28

    # Row 2: section headers
    ws.append([
        "", "",
        "As Per GSTR 2B Data", "", "", "", "",
        "",
        "As Per Books Data", "", "", "", "",
        "",
        "Difference", "", "", "",
        "", "",
    ])
    ws.merge_cells("C2:G2")
    ws.merge_cells("I2:M2")
    ws.merge_cells("O2:R2")
    for addr, fill in [("C2", fill_2b), ("I2", fill_books_hdr), ("O2", fill_diff_hdr)]:
        c = ws[addr]
        c.font      = section_font
        c.alignment = center
        c.fill      = fill
    ws.row_dimensions[2].height = 18

    # Row 3: column headers
    ws.append([
        "Common Name", "GSTIN",
        "Name as per 2B",    "Taxable Value (₹)", "CGST (₹)", "SGST (₹)", "IGST (₹)",
        "",
        "Name as per Books", "Taxable Value (₹)", "CGST (₹)", "SGST (₹)", "IGST (₹)",
        "",
        "Taxable Value (₹)", "CGST (₹)", "SGST (₹)", "IGST (₹)",
        "Remark 1", "Remark 2",
    ])
    for col_idx in range(1, NUM_COLS + 1):
        c = ws.cell(row=3, column=col_idx)
        c.font      = header_font
        c.alignment = center
        c.border    = border
        if   col_idx in range(3, 8):
            c.fill = fill_2b
        elif col_idx in range(9, 14):
            c.fill = fill_books_hdr
        elif col_idx in range(15, 19):
            c.fill = fill_diff_hdr
    ws.row_dimensions[3].height = 18

    # Data rows
    for rd in all_rows:
        ws.append([
            rd["common_name"],   rd["gstin"],
            rd["name_2b"],       rd["taxable_2b"],   rd["cgst_2b"],    rd["sgst_2b"],   rd["igst_2b"],
            "",
            rd["name_books"],    rd["taxable_books"], rd["cgst_books"], rd["sgst_books"],rd["igst_books"],
            "",
            rd["diff_taxable"],  rd["diff_cgst"],     rd["diff_sgst"],  rd["diff_igst"],
            rd["remark_1"],      rd["remark_2"],
        ])
        row_idx = ws.max_row

        r1 = rd["remark_1"]
        if r1 == "Matched":
            fill = fill_matched
        elif r1 == "Amount Mismatch":
            fill = fill_mismatch
        elif r1 == "Showing in 2B but Not in Books":
            fill = fill_only_2b
        else:
            fill = fill_only_books

        for col_idx in range(1, NUM_COLS + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            if c.value != "" and col_idx not in SPACERS:
                c.fill = fill
            c.border = border
            if col_idx in NUMERIC_COLS:
                c.alignment = right_align

    # Column widths: A B | C D E F G | (H) | I J K L M | (N) | O P Q R | S T
    for i, w in enumerate(
        [35, 20,  35, 14, 12, 12, 12,  2,  35, 14, 12, 12, 12,  2,  14, 12, 12, 12,  32, 42], 1
    ):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"


def build_excel_output(
    results: list[MatchResult],
    output_path: str | None = None,
) -> bytes:
    import openpyxl
    from io import BytesIO as BIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reco 2B vs Books"
    _populate_workbook(ws, results)

    bio = BIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def _copy_workbook_sheets(src_bytes: bytes, target_wb: Any, title_prefix: str) -> None:
    src_bytes = _ensure_xlsx(src_bytes)
    import openpyxl
    from io import BytesIO
    import copy
    # OCTA 2B exports are a single flat sheet (e.g. "Sheet1"), not the portal
    # B2B/B2BA/CDNR tabs — copy it as-is instead of filtering it out.
    is_octa_2b = title_prefix == "2B" and _is_octa_format(src_bytes)
    try:
        src_wb = openpyxl.load_workbook(BytesIO(src_bytes), data_only=True)
        for name in src_wb.sheetnames:
            # ONLY copy the allowed 2B sheets (case-insensitive checking) if prefix is 2B
            if title_prefix == "2B" and not is_octa_2b:
                name_clean = str(name).strip().upper()
                allowed_clean = {s.upper() for s in ALLOWED_SHEETS}
                if name_clean not in allowed_clean:
                    continue
            
            src_sheet = src_wb[name]
            # Ensure unique and beautiful sheet title
            # Max sheet title length in Excel is 31 characters
            prefix = f"{title_prefix} - "
            max_name_len = 31 - len(prefix)
            short_name = name[:max_name_len]
            tgt_title = f"{prefix}{short_name}"
            
            # Avoid duplicate sheet titles
            base_title = tgt_title
            counter = 1
            while tgt_title in target_wb.sheetnames:
                suffix = f"_{counter}"
                tgt_title = base_title[:31 - len(suffix)] + suffix
                counter += 1
                
            tgt_sheet = target_wb.create_sheet(title=tgt_title)
            
            # Copy merged cell ranges
            for merged_range in src_sheet.merged_cells.ranges:
                tgt_sheet.merge_cells(str(merged_range))
                
            # Copy dimensions (row heights)
            for r_idx in src_sheet.row_dimensions:
                tgt_sheet.row_dimensions[r_idx].height = src_sheet.row_dimensions[r_idx].height
                
            # Copy dimensions (column widths)
            for col_letter in src_sheet.column_dimensions:
                tgt_sheet.column_dimensions[col_letter].width = src_sheet.column_dimensions[col_letter].width
                
            # Copy cell values and styles
            for r in range(1, src_sheet.max_row + 1):
                for c in range(1, src_sheet.max_column + 1):
                    cell = src_sheet.cell(row=r, column=c)
                    tgt_cell = tgt_sheet.cell(row=r, column=c, value=cell.value)
                    if cell.has_style:
                        tgt_cell.font = copy.copy(cell.font)
                        tgt_cell.fill = copy.copy(cell.fill)
                        tgt_cell.alignment = copy.copy(cell.alignment)
                        tgt_cell.border = copy.copy(cell.border)
                        tgt_cell.number_format = cell.number_format
    except Exception as e:
        print(f"Error copying sheets for prefix {title_prefix}: {e}")


def _build_month_summary_sheet(wb: Any, results: list[Any]) -> None:
    """Add a 'Month Summary' sheet when results span ≥ 2 distinct invoice months.
    Groups by YYYY-MM (preferring 2B date, falling back to Books date).
    Columns: Month | 2B Taxable/IGST/CGST/SGST | Books Taxable/IGST/CGST/SGST |
             Diff Taxable/IGST | Matched | 2B Only | Books Only"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from collections import defaultdict

    month_data: dict = defaultdict(lambda: {
        "taxable_2b": 0.0, "igst_2b": 0.0, "cgst_2b": 0.0, "sgst_2b": 0.0,
        "taxable_bk": 0.0, "igst_bk": 0.0, "cgst_bk": 0.0, "sgst_bk": 0.0,
        "matched": 0, "only_2b": 0, "only_bk": 0,
    })

    for r in results:
        g = _get_val(r, "gstr2b",   None)
        b = _get_val(r, "purchase", None)
        g_date = str(_get_val(g, "doc_date", "") or "") if g else ""
        b_date = str(_get_val(b, "doc_date", "") or "") if b else ""
        month  = (g_date or b_date)[:7]   # "YYYY-MM" or ""
        if not month or len(month) < 7:
            month = "Unknown"
        else:
            try:
                yr, mn = int(month[:4]), int(month[5:7])
                if not (2000 <= yr <= 2099) or not (1 <= mn <= 12):
                    month = "Unknown"
            except (ValueError, IndexError):
                month = "Unknown"

        md = month_data[month]
        if g:
            md["taxable_2b"] += float(_get_val(g, "taxable_value", 0) or 0)
            md["igst_2b"]    += float(_get_val(g, "igst",          0) or 0)
            md["cgst_2b"]    += float(_get_val(g, "cgst",          0) or 0)
            md["sgst_2b"]    += float(_get_val(g, "sgst",          0) or 0)
        if b:
            md["taxable_bk"] += float(_get_val(b, "taxable_value", 0) or 0)
            md["igst_bk"]    += float(_get_val(b, "igst",          0) or 0)
            md["cgst_bk"]    += float(_get_val(b, "cgst",          0) or 0)
            md["sgst_bk"]    += float(_get_val(b, "sgst",          0) or 0)

        status = str(_get_val(r, "suggested_action", "") or "")
        if "Not in Books" in status:
            md["only_2b"] += 1
        elif "Not in 2B" in status:
            md["only_bk"] += 1
        else:
            md["matched"] += 1

    # Only add sheet when spanning ≥ 2 months
    real_months = [m for m in month_data if m != "Unknown"]
    if len(real_months) < 2:
        return

    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")

    fill_hdr   = PatternFill("solid", fgColor="1F3864")
    fill_total = PatternFill("solid", fgColor="D9E1F2")
    fill_alt   = PatternFill("solid", fgColor="F2F2F2")
    font_hdr   = Font(bold=True, color="FFFFFF", size=10)
    font_total = Font(bold=True, size=10)
    num_fmt    = "#,##0.00"
    int_fmt    = "#,##0"

    ws = wb.create_sheet(title="Month Summary")

    headers = [
        "Month",
        "2B Taxable", "2B IGST", "2B CGST", "2B SGST",
        "Books Taxable", "Books IGST", "Books CGST", "Books SGST",
        "Diff Taxable", "Diff IGST",
        "Matched", "2B Only", "Books Only",
    ]
    for col, hdr in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=hdr)
        c.font      = font_hdr
        c.fill      = fill_hdr
        c.alignment = center
        c.border    = border
    ws.column_dimensions["A"].width = 14
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 16

    sorted_months = sorted(month_data.keys())
    totals = {k: 0.0 for k in ["taxable_2b","igst_2b","cgst_2b","sgst_2b",
                                 "taxable_bk","igst_bk","cgst_bk","sgst_bk",
                                 "matched","only_2b","only_bk"]}

    for row_idx, month in enumerate(sorted_months, 2):
        md   = month_data[month]
        fill = fill_alt if row_idx % 2 == 0 else PatternFill()
        diff_t = round(md["taxable_2b"] - md["taxable_bk"], 2)
        diff_i = round(md["igst_2b"]    - md["igst_bk"],    2)

        row_vals = [
            month,
            md["taxable_2b"], md["igst_2b"], md["cgst_2b"], md["sgst_2b"],
            md["taxable_bk"], md["igst_bk"], md["cgst_bk"], md["sgst_bk"],
            diff_t, diff_i,
            md["matched"], md["only_2b"], md["only_bk"],
        ]
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border = border
            c.fill   = fill
            if col == 1:
                c.alignment = center
            elif col <= 11:
                c.alignment    = right
                c.number_format = num_fmt
            else:
                c.alignment    = right
                c.number_format = int_fmt

        for k in totals:
            totals[k] += md.get(k, 0)

    # Totals row
    tr = len(sorted_months) + 2
    diff_t_tot = round(totals["taxable_2b"] - totals["taxable_bk"], 2)
    diff_i_tot = round(totals["igst_2b"]    - totals["igst_bk"],    2)
    total_vals = [
        "TOTAL",
        totals["taxable_2b"], totals["igst_2b"], totals["cgst_2b"], totals["sgst_2b"],
        totals["taxable_bk"], totals["igst_bk"], totals["cgst_bk"], totals["sgst_bk"],
        diff_t_tot, diff_i_tot,
        int(totals["matched"]), int(totals["only_2b"]), int(totals["only_bk"]),
    ]
    for col, val in enumerate(total_vals, 1):
        c = ws.cell(row=tr, column=col, value=val)
        c.font   = font_total
        c.fill   = fill_total
        c.border = border
        c.alignment = center if col == 1 else right
        if 2 <= col <= 11:
            c.number_format = num_fmt
        elif col > 11:
            c.number_format = int_fmt

    ws.freeze_panes = "B2"


def _copy_rcm_sheet(src_bytes: bytes, target_wb: Any) -> None:
    src_bytes = _ensure_xlsx(src_bytes)
    import openpyxl
    from io import BytesIO
    import copy
    try:
        src_wb = openpyxl.load_workbook(BytesIO(src_bytes), data_only=True)
        # Find the B2B sheet
        b2b_sheet = None
        for name in src_wb.sheetnames:
            if name.upper().strip() == "B2B":
                b2b_sheet = src_wb[name]
                break
        
        if b2b_sheet is None:
            # Fallback to check if "B2B" is in the name
            for name in src_wb.sheetnames:
                if "B2B" in name.upper():
                    b2b_sheet = src_wb[name]
                    break
                    
        if b2b_sheet is None:
            print("No B2B sheet found in GSTR-2B for RCM copying.")
            return

        tgt_sheet = target_wb.create_sheet(title="RCM")
        
        # Copy column dimensions (widths)
        for col_letter in b2b_sheet.column_dimensions:
            tgt_sheet.column_dimensions[col_letter].width = b2b_sheet.column_dimensions[col_letter].width
            
        # Copy merged cells that are within the header (rows 1-6)
        for merged_range in b2b_sheet.merged_cells.ranges:
            if merged_range.max_row <= 6:
                tgt_sheet.merge_cells(str(merged_range))
                
        # Find the column containing 'Supply Attract Reverse Charge' in rows 4-7
        rcm_col_idx = 8  # Default fallback to Column H (8th column)
        found_col = False
        for r_idx in range(4, 8):
            if found_col:
                break
            for col_idx in range(1, b2b_sheet.max_column + 1):
                cell_val = str(b2b_sheet.cell(row=r_idx, column=col_idx).value or "").strip().lower()
                if "supply attract reverse charge" in cell_val or "reverse charge" in cell_val:
                    rcm_col_idx = col_idx
                    found_col = True
                    break
        
        # Copy rows: first 6 are headers, remaining are filtered by 'Yes' in the rcm column
        tgt_row_idx = 1
        for r in range(1, b2b_sheet.max_row + 1):
            is_header = (r <= 6)
            is_rcm = False
            if not is_header:
                rcm_val = str(b2b_sheet.cell(row=r, column=rcm_col_idx).value or "").strip().lower()
                if rcm_val == "yes":
                    is_rcm = True
            
            if is_header or is_rcm:
                # Copy row height
                if r in b2b_sheet.row_dimensions:
                    tgt_sheet.row_dimensions[tgt_row_idx].height = b2b_sheet.row_dimensions[r].height
                    
                for c in range(1, b2b_sheet.max_column + 1):
                    cell = b2b_sheet.cell(row=r, column=c)
                    tgt_cell = tgt_sheet.cell(row=tgt_row_idx, column=c, value=cell.value)
                    if cell.has_style:
                        tgt_cell.font = copy.copy(cell.font)
                        tgt_cell.fill = copy.copy(cell.fill)
                        tgt_cell.alignment = copy.copy(cell.alignment)
                        tgt_cell.border = copy.copy(cell.border)
                        tgt_cell.number_format = cell.number_format
                tgt_row_idx += 1
    except Exception as e:
        print(f"Error copying RCM sheet: {e}")


def _append_rcm_rows(gstr2b_bytes: bytes, target_wb: Any) -> None:
    """Append RCM data rows (reverse charge = yes) from an additional state's GSTR-2B
    into the existing 'RCM' sheet in target_wb. Headers are NOT re-copied (state 1 wrote them).
    Called by the multistate builder for states 2-N."""
    gstr2b_bytes = _ensure_xlsx(gstr2b_bytes)
    import openpyxl
    from io import BytesIO
    import copy
    try:
        src_wb = openpyxl.load_workbook(BytesIO(gstr2b_bytes), data_only=True)
        b2b_sheet = None
        for name in src_wb.sheetnames:
            if name.upper().strip() == "B2B":
                b2b_sheet = src_wb[name]
                break
        if b2b_sheet is None:
            for name in src_wb.sheetnames:
                if "B2B" in name.upper():
                    b2b_sheet = src_wb[name]
                    break
        if b2b_sheet is None:
            return

        # Locate RCM column
        rcm_col_idx = 8
        for r_idx in range(4, 8):
            found = False
            for col_idx in range(1, b2b_sheet.max_column + 1):
                cell_val = str(b2b_sheet.cell(row=r_idx, column=col_idx).value or "").strip().lower()
                if "supply attract reverse charge" in cell_val or "reverse charge" in cell_val:
                    rcm_col_idx = col_idx
                    found = True
                    break
            if found:
                break

        if "RCM" not in target_wb.sheetnames:
            target_wb.create_sheet(title="RCM")
        tgt_sheet = target_wb["RCM"]

        # Append only RCM data rows (skip header rows 1-6)
        for r in range(7, b2b_sheet.max_row + 1):
            rcm_val = str(b2b_sheet.cell(row=r, column=rcm_col_idx).value or "").strip().lower()
            if rcm_val == "yes":
                new_row = tgt_sheet.max_row + 1
                for c in range(1, b2b_sheet.max_column + 1):
                    cell = b2b_sheet.cell(row=r, column=c)
                    tgt_cell = tgt_sheet.cell(row=new_row, column=c, value=cell.value)
                    if cell.has_style:
                        tgt_cell.font      = copy.copy(cell.font)
                        tgt_cell.fill      = copy.copy(cell.fill)
                        tgt_cell.alignment = copy.copy(cell.alignment)
                        tgt_cell.border    = copy.copy(cell.border)
                        tgt_cell.number_format = cell.number_format
        src_wb.close()
    except Exception as e:
        print(f"Error appending RCM rows: {e}")


def build_gstr2b_books_workbook(results: list[dict], payload: dict | None = None) -> Any:
    import openpyxl
    wb = openpyxl.Workbook()

    # 1. Copy GSTR-2B, Purchase Register, and Debit Note Register sheets first if payload is provided
    if payload:
        import base64
        # Decode and copy GSTR-2B sheets
        if "_gstr2b_b64" in payload:
            try:
                gstr2b_bytes = base64.b64decode(payload["_gstr2b_b64"])
                _copy_workbook_sheets(gstr2b_bytes, wb, "2B")
            except Exception as e:
                print(f"Failed to copy GSTR-2B sheets: {e}")
                
        # Decode and copy Purchase Register sheets
        if "_purchase_b64" in payload:
            try:
                purchase_bytes = base64.b64decode(payload["_purchase_b64"])
                _copy_workbook_sheets(purchase_bytes, wb, "PR")
            except Exception as e:
                print(f"Failed to copy Purchase Register sheets: {e}")
                
        # Decode and copy Debit Note Register sheets
        if "_debit_b64" in payload:
            try:
                debit_bytes = base64.b64decode(payload["_debit_b64"])
                _copy_workbook_sheets(debit_bytes, wb, "DN")
            except Exception as e:
                print(f"Failed to copy Debit Note sheets: {e}")
                
    # 2. Add RCM entries sheet
    if payload and "_gstr2b_b64" in payload:
        import base64
        try:
            gstr2b_bytes = base64.b64decode(payload["_gstr2b_b64"])
            _copy_rcm_sheet(gstr2b_bytes, wb)
        except Exception as e:
            print(f"Failed to dynamically populate RCM sheet: {e}")
            wb.create_sheet(title="RCM")
    else:
        wb.create_sheet(title="RCM")
    
    # 3. Add Vendor Summary sheet
    ws_vendor = wb.create_sheet(title="Vendor Summary")
    _build_vendor_summary_sheet(ws_vendor, results)

    # 4. Add Month Summary sheet (only when results span ≥ 2 distinct months)
    _build_month_summary_sheet(wb, results)

    # 5. Add main Output sheet (Reco 2B vs Books)
    ws_output = wb.create_sheet(title="Reco 2B vs Books")
    _populate_workbook(ws_output, results)

    # Remove the default 'Sheet' that openpyxl creates automatically if we added other sheets
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb["Sheet"])

    return wb


# ---------------------------------------------------------------------------
# Top-level entry point
# ---------------------------------------------------------------------------

def reconcile_gstr2b_vs_books(
    gstr2b_data: bytes,
    purchase_data: bytes,
    debit_data: bytes,
    tolerance: float = 1.0,
) -> tuple[list[NormalizedInvoice], list[NormalizedInvoice], list[MatchResult]]:
    gstr2b_data = _ensure_xlsx(gstr2b_data)
    purchase_data = _ensure_xlsx(purchase_data)
    debit_data = _ensure_xlsx(debit_data)
    gstr2b_records = parse_gstr2b(gstr2b_data)
    books_records = parse_books(purchase_data, debit_data)
    
    # Map Books supplier_name to supplier_gstin from GSTR-2B to avoid false negatives from spelling
    name_to_gstin = {}
    for g in gstr2b_records:
        name_clean = str(g.supplier_name or "").strip().lower()
        if name_clean and g.supplier_gstin:
            name_to_gstin[name_clean] = g.supplier_gstin
            
    for b in books_records:
        if not b.supplier_gstin:
            name_clean = str(b.supplier_name or "").strip().lower()
            if name_clean in name_to_gstin:
                b.supplier_gstin = name_to_gstin[name_clean]
                
    results = reconcile_by_invoice_no(gstr2b_records, books_records, tolerance=tolerance)
    return gstr2b_records, books_records, results
