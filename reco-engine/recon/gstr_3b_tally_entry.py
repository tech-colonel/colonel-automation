"""
GSTR-3B Tally Entry Generator
Parses a GSTR-3B file (PDF or Excel) and generates Tally journal entries per SOP.

SOP (GSTR-3B_SOP.docx):
  Part 1 — ITC Transfer to Credit Ledger
    Debit  : Credit Ledger [Tax] [StateAbbr]  ← Row 4(C) Net ITC Available
    Credit : Input [Tax] [StateAbbr]          ← Row 4(4) Inward ISD + Row 4(5) All Other ITC

  Part 2 — Setting Off Output Liability
    Debit  : Output [Tax] [StateAbbr]         ← Row 6.1(A) total output tax (ITC + Cash)
    Credit : Credit Ledger [Tax] [StateAbbr]  ← Row 6.1(A) Tax paid through ITC
    Credit : Cash Ledger [StateAbbr]          ← Row 6.1(A) Tax paid in Cash (if > 0)

  Part 3 — RCM (only if Row 6.1(B) has values)
    Credit : RCM [Tax] [StateAbbr]            ← Row 6.1(B) Reverse Charge amounts
"""

from __future__ import annotations

import calendar
import re
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from recon.gstr_2b_books import GST_STATE_CODES, _GSTIN_RE, _ensure_xlsx


# ---------------------------------------------------------------------------
# State abbreviation map  (vehicle-registration / RTO codes used in Tally)
# ---------------------------------------------------------------------------

GST_STATE_SHORT: dict[str, str] = {
    "01": "JK",  "02": "HP",  "03": "PB",  "04": "CH",
    "05": "UK",  "06": "HR",  "07": "DL",  "08": "RJ",
    "09": "UP",  "10": "BR",  "11": "SK",  "12": "AR",
    "13": "NL",  "14": "MN",  "15": "MZ",  "16": "TR",
    "17": "ML",  "18": "AS",  "19": "WB",  "20": "JH",
    "21": "OD",  "22": "CT",  "23": "MP",  "24": "GJ",
    "25": "DD",  "26": "DN",  "27": "MH",  "28": "AP",
    "29": "KA",  "30": "GA",  "31": "LD",  "32": "KL",
    "33": "TN",  "34": "PY",  "35": "AN",  "36": "TS",
    "37": "AP",  "38": "LA",  "97": "OT",  "99": "CJ",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]
_PERIOD_RE = re.compile(
    r'(?:' + '|'.join(_MONTHS) + r')\s+\d{4}', re.IGNORECASE
)
_PERIOD_MM_YYYY = re.compile(r'\b(0[1-9]|1[0-2])[/-](\d{4})\b')
_GSTIN_LOOSE   = re.compile(r'[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][0-9A-Z]Z[0-9A-Z]')
_DATE_DDMMYYYY = re.compile(r'\b(\d{2}/\d{2}/\d{4})\b')


def _parse_num(val: Any) -> float:
    """Parse Indian-formatted number string to float. Returns 0.0 on failure."""
    if val is None:
        return 0.0
    s = str(val).strip()
    # pdfplumber splits numbers across visual lines: '4521481.\n00' → '4521481.00'
    s = s.replace('\n', '').replace('\r', '')
    # PDF watermark letters prepended to values: 'F4521481.00' → '4521481.00'
    s = re.sub(r'^[A-Za-z]{1,4}', '', s).strip()
    s = s.replace(',', '').replace('₹', '')
    if not s or s in ('-', '--', 'nan', 'NaN', 'None', 'N/A', ''):
        return 0.0
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _is_numeric(val: Any) -> bool:
    """Return True only if val can be parsed as a float (not just any non-empty string)."""
    s = str(val or '').strip()
    s = s.replace('\n', '').replace('\r', '')
    s = re.sub(r'^[A-Za-z]{1,4}', '', s).strip()
    s = s.replace(',', '').replace('₹', '')
    if not s or s in ('-', '--', 'nan', 'NaN', 'None', 'N/A', ''):
        return False
    try:
        float(s)
        return True
    except (ValueError, TypeError):
        return False


def _empty_tax() -> dict:
    return {'igst': 0.0, 'cgst': 0.0, 'sgst': 0.0}


def _nums_from_text(text: str) -> list[float]:
    """Extract all numbers from a text string."""
    tokens = re.findall(r'[\d,]+\.?\d*', text)
    out = []
    for t in tokens:
        try:
            out.append(float(t.replace(',', '')))
        except ValueError:
            pass
    return out


def _state_info(gstin: str) -> tuple[str, str]:
    """Return (full_name, abbreviation) from a GSTIN's first 2 digits."""
    code = gstin[:2] if len(gstin) >= 2 else ''
    full  = GST_STATE_CODES.get(code, code)
    short = GST_STATE_SHORT.get(code, code)
    return full, short


# ---------------------------------------------------------------------------
# COA helpers (for 3B ledger-name matching)
# ---------------------------------------------------------------------------

def _load_coa(data: bytes, filename: str) -> dict:
    """Load a 3B COA file (single-column ledger list) → {lower_name: original_name}."""
    suffix = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''
    if data[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1' or suffix == 'xls':
        data = _ensure_xlsx(data)
    try:
        sheets = pd.read_excel(BytesIO(data), sheet_name=None, dtype=str, header=None)
        sheet_names = list(sheets.keys())
        target = None
        for name in sheet_names:
            if 'ledger' in name.lower():
                target = sheets[name]
                break
        if target is None:
            target = sheets[sheet_names[0]]
        result: dict = {}
        for _, row in target.iterrows():
            for cell in row:
                val = str(cell).strip() if cell is not None else ''
                low = val.lower()
                if val and low not in ('nan', '', 'none', 'ledger name', 'name', 'particular', 'particulars'):
                    result[low] = val
        return result
    except Exception:
        return {}


def _load_coa_from_list(ledger_list: list) -> dict:
    """Build COA dict from a plain list of ledger name strings."""
    result: dict = {}
    for name in ledger_list:
        val = str(name).strip() if name is not None else ''
        if val:
            result[val.lower()] = val
    return result


def _match_coa_ledger(generated: str, coa: dict, state_short: str = '') -> str:
    """Match a generated ledger name to the COA. 3-tier: exact → state-stripped → original."""
    if not coa:
        return generated
    key = generated.lower()
    if key in coa:
        return coa[key]
    if state_short:
        stripped = re.sub(r'\s+' + re.escape(state_short.lower()) + r'\s*$', '', key).strip()
        if stripped != key and stripped in coa:
            return coa[stripped]
    return generated


def _load_voucher_types(data: bytes, filename: str) -> dict:
    """Load a Voucher Type master (single-column list) → {lower: original}. Reuses _load_coa logic."""
    return _load_coa(data, filename)


def _load_voucher_types_from_list(vt_list: list) -> dict:
    """Build VT dict from a plain list of voucher type name strings."""
    return _load_coa_from_list(vt_list)


def _match_voucher_type(base: str, vt_master: dict, state_short: str) -> str:
    """Match base voucher type to state-specific variant. 'Journal' + 'UP' → 'Journal UP' if in master."""
    if not vt_master:
        return base
    candidate = f"{base} {state_short}"
    if candidate.lower() in vt_master:
        return vt_master[candidate.lower()]
    if base.lower() in vt_master:
        return vt_master[base.lower()]
    return base


def _derive_voucher_date(period: str) -> str:
    """Convert 'April 2026' → '30-Apr-2026' (last day of month)."""
    try:
        parts = period.strip().split()
        if len(parts) == 2:
            month_name, year_str = parts
            month_num = next(
                (i + 1 for i, m in enumerate(_MONTHS) if m.lower() == month_name.lower()),
                None
            )
            if month_num and year_str.isdigit():
                year = int(year_str)
                last_day = calendar.monthrange(year, month_num)[1]
                month_abbr = month_name[:3].capitalize()
                return f"{last_day:02d}-{month_abbr}-{year}"
    except Exception:
        pass
    return period


# ---------------------------------------------------------------------------
# PDF Parser
# ---------------------------------------------------------------------------

def _collect_nums_from_rows(all_rows: list[list[str]], start: int, count: int) -> list[float]:
    """Collect strictly-numeric values from all_rows[start : start+count]."""
    nums: list[float] = []
    end = min(start + count, len(all_rows))
    for ri in range(start, end):
        for cell in all_rows[ri]:
            if _is_numeric(cell):
                nums.append(_parse_num(cell))
    return nums


def _extract_61_section(nums: list[float]) -> tuple[dict, dict, dict]:
    """
    GSTR-3B 6.1 table comes in TWO PDF layouts:
      Layout A (columns-first): [TaxIGST,TaxCGST,TaxSGST,TaxCess, ITCIGST,ITCCGST,ITCSGST,ITCCess, CashIGST,CashCGST,CashSGST,...]
      Layout B (rows-first):    [IGST(Tax,ITC,Cash), CGST(Tax,ITC,Cash), SGST(Tax,ITC,Cash), Cess(Tax,ITC,Cash)]

    Validates using Tax = ITC + Cash (±1 tolerance). Returns (tax_amt, itc, cash).
    """
    TOL = 1.0

    def valid(t, i, c):
        return all(abs(t[k] - (i[k] + c[k])) <= TOL for k in ('igst', 'cgst', 'sgst'))

    if len(nums) >= 11:
        # Layout A
        ta = {'igst': nums[0], 'cgst': nums[1],  'sgst': nums[2]}
        ia = {'igst': nums[4], 'cgst': nums[5],  'sgst': nums[6]}
        ca = {'igst': nums[8], 'cgst': nums[9],  'sgst': nums[10]}
        if valid(ta, ia, ca):
            return ta, ia, ca

        # Layout B
        tb = {'igst': nums[0], 'cgst': nums[3], 'sgst': nums[6]}
        ib = {'igst': nums[1], 'cgst': nums[4], 'sgst': nums[7]}
        cb = {'igst': nums[2], 'cgst': nums[5], 'sgst': nums[8]}
        if valid(tb, ib, cb):
            return tb, ib, cb

    if len(nums) >= 6:
        t = {'igst': nums[0], 'cgst': nums[1], 'sgst': nums[2]}
        i = {'igst': nums[0], 'cgst': nums[1], 'sgst': nums[2]}
        c = {'igst': nums[3], 'cgst': nums[4], 'sgst': nums[5]}
        if valid(t, i, c):
            return t, i, c
        # Try: Tax in first 3, Cash in next 3 with ITC=0
        i2 = _empty_tax()
        if valid(t, i2, c):
            return t, i2, c

    if len(nums) >= 3:
        t = {'igst': nums[0], 'cgst': nums[1], 'sgst': nums[2]}
        return t, t.copy(), _empty_tax()

    return _empty_tax(), _empty_tax(), _empty_tax()


def _parse_gstr3b_pdf(data: bytes) -> dict:
    try:
        import pdfplumber
    except ImportError:
        raise ValueError(
            "PDF parsing requires pdfplumber. "
            "Install with: pip3 install pdfplumber"
        )

    result = {
        'gstin': '', 'state': '', 'state_short': '', 'period': '', 'filing_date': '',
        'row_4c':       _empty_tax(),
        'row_4_45':     _empty_tax(),
        'row_61a_tax':         _empty_tax(),  # Tax Amount (output liability) per tax type
        'row_61a_itc':         _empty_tax(),
        'row_61a_cash':        _empty_tax(),
        'row_61a_credit_type': _empty_tax(),  # ITC by credit ledger type (IGST/CGST/SGST column sums)
        'row_61b':      _empty_tax(),  # RCM Tax Amount (= Cash paid for RCM)
    }

    all_rows: list[list[str]] = []
    full_text = ''

    with pdfplumber.open(BytesIO(data)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ''
            full_text += page_text + '\n'
            for table in page.extract_tables() or []:
                for row in table:
                    clean = [str(c or '').strip() for c in row]
                    if any(c for c in clean):
                        all_rows.append(clean)

    # ---- GSTIN ----
    gstin_match = _GSTIN_LOOSE.search(full_text)
    if gstin_match:
        g = gstin_match.group()
        result['gstin'] = g
        result['state'], result['state_short'] = _state_info(g)

    # ---- Period ----
    pm = _PERIOD_RE.search(full_text)
    if pm:
        result['period'] = pm.group().title()
    else:
        mm = _PERIOD_MM_YYYY.search(full_text)
        if mm:
            result['period'] = f"{_MONTHS[int(mm.group(1)) - 1]} {mm.group(2)}"

    # ---- Filing Date (ARN date from PDF — DD/MM/YYYY) ----
    dates_found = _DATE_DDMMYYYY.findall(full_text)
    if dates_found:
        try:
            d, m, y = dates_found[0].split('/')
            month_abbr = _MONTHS[int(m) - 1][:3].capitalize()
            result['filing_date'] = f"{d}-{month_abbr}-{y}"
        except Exception:
            pass

    # ---- Parse table rows ----
    row_4_4 = _empty_tax()
    row_4_5 = _empty_tax()

    n_rows = len(all_rows)

    for i, row in enumerate(all_rows):
        label = ' '.join(row).lower()

        # Row 4(C) – Net ITC Available
        # Values may be on this row or the next; use a 2-row window
        if re.search(r'net\s+itc\s+available|4\s*\(c\)', label):
            nums = _collect_nums_from_rows(all_rows, i, 2)
            if len(nums) >= 3 and nums[0] + nums[1] + nums[2] > 0:
                result['row_4c'] = {'igst': nums[0], 'cgst': nums[1], 'sgst': nums[2]}

        # Row 4(A)(4) – Inward supplies from ISD
        # Use ^ so '16(4)' in other rows doesn't match; keep ISD aliases unanchored
        if re.search(r'^\(4\)|^\(iv\)|inward.*isd|from\s+isd', label):
            nums = _collect_nums_from_rows(all_rows, i, 2)
            if len(nums) >= 3:
                row_4_4 = {'igst': row_4_4['igst'] + nums[0],
                            'cgst': row_4_4['cgst'] + nums[1],
                            'sgst': row_4_4['sgst'] + nums[2]}

        # Row 4(A)(5) – All other ITC
        # Use ^ so '9(5)' in 6.1(B) label and '17(5)' in rule refs don't match
        if re.search(r'^\(5\)|^\(v\)|all\s+other\s+itc', label):
            nums = _collect_nums_from_rows(all_rows, i, 2)
            if len(nums) >= 3:
                row_4_5 = {'igst': row_4_5['igst'] + nums[0],
                            'cgst': row_4_5['cgst'] + nums[1],
                            'sgst': row_4_5['sgst'] + nums[2]}

        # Row 6.1(A) – "(A) Other than reverse charge" label row.
        # The NEXT 3 rows are the IGST / CGST / SGST data rows respectively.
        # Each data row layout (11 cols):
        #   0:TaxType  1:TaxPayable  2:Adj  3:NetTaxPayable  4:ITC-IGST  5:ITC-CGST
        #   6:ITC-SGST  7:ITC-Cess  8:CashPaid  9:InterestCash  10:LateFee
        if re.search(r'other\s+than\s+reverse\s+charge', label):
            def _gc(r, col):
                try: return _parse_num(r[col])
                except IndexError: return 0.0
            igst_r = all_rows[i + 1] if i + 1 < n_rows else []
            cgst_r = all_rows[i + 2] if i + 2 < n_rows else []
            sgst_r = all_rows[i + 3] if i + 3 < n_rows else []
            igst_tax  = _gc(igst_r, 1)
            cgst_tax  = _gc(cgst_r, 1)
            sgst_tax  = _gc(sgst_r, 1)
            igst_cash = _gc(igst_r, 8)
            cgst_cash = _gc(cgst_r, 8)
            sgst_cash = _gc(sgst_r, 8)
            result['row_61a_tax']  = {'igst': igst_tax,  'cgst': cgst_tax,  'sgst': sgst_tax}
            result['row_61a_itc']  = {'igst': max(0.0, igst_tax - igst_cash),
                                       'cgst': max(0.0, cgst_tax - cgst_cash),
                                       'sgst': max(0.0, sgst_tax - sgst_cash)}
            result['row_61a_cash'] = {'igst': igst_cash, 'cgst': cgst_cash, 'sgst': sgst_cash}
            # ITC credit-ledger type breakdown: sum each ITC column (4=IGST,5=CGST,6=SGST)
            # across all 3 output rows to get how much of each credit ledger was consumed.
            result['row_61a_credit_type'] = {
                'igst': _gc(igst_r, 4) + _gc(cgst_r, 4) + _gc(sgst_r, 4),
                'cgst': _gc(igst_r, 5) + _gc(cgst_r, 5) + _gc(sgst_r, 5),
                'sgst': _gc(igst_r, 6) + _gc(cgst_r, 6) + _gc(sgst_r, 6),
            }

        # Row 6.1(B) – "(B) Reverse charge …" label row. Same column structure.
        # RCM: Tax Payable (col 1) = Cash paid (col 8); ITC = 0 for RCM.
        if re.search(r'reverse\s+charge', label) and 'other than' not in label:
            def _gc2(r, col):
                try: return _parse_num(r[col])
                except IndexError: return 0.0
            igst_r2 = all_rows[i + 1] if i + 1 < n_rows else []
            cgst_r2 = all_rows[i + 2] if i + 2 < n_rows else []
            sgst_r2 = all_rows[i + 3] if i + 3 < n_rows else []
            result['row_61b'] = {
                'igst': _gc2(igst_r2, 1),
                'cgst': _gc2(cgst_r2, 1),
                'sgst': _gc2(sgst_r2, 1),
            }

    # Fallback: if table extraction gave nothing for 4(C), try raw text
    if not any(result['row_4c'].values()):
        result['row_4c'] = _extract_row_from_text(full_text, r'net\s+itc\s+available')

    result['row_4_45'] = {
        'igst': row_4_4['igst'] + row_4_5['igst'],
        'cgst': row_4_4['cgst'] + row_4_5['cgst'],
        'sgst': row_4_4['sgst'] + row_4_5['sgst'],
    }

    return result


def _extract_row_from_text(text: str, pattern: str) -> dict:
    """Regex fallback: find a line matching pattern, return first 3 nums as igst/cgst/sgst."""
    for line in text.split('\n'):
        if re.search(pattern, line, re.IGNORECASE):
            nums = [n for n in _nums_from_text(line) if n >= 0]
            if len(nums) >= 3:
                return {'igst': nums[0], 'cgst': nums[1], 'sgst': nums[2]}
    return _empty_tax()


# ---------------------------------------------------------------------------
# Excel Parser
# ---------------------------------------------------------------------------

def _parse_gstr3b_excel(data: bytes, filename: str) -> dict:
    suffix = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''
    if data[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1' or suffix == 'xls':
        data = _ensure_xlsx(data)

    result = {
        'gstin': '', 'state': '', 'state_short': '', 'period': '', 'filing_date': '',
        'row_4c':       _empty_tax(),
        'row_4_45':     _empty_tax(),
        'row_61a_tax':  _empty_tax(),
        'row_61a_itc':  _empty_tax(),
        'row_61a_cash': _empty_tax(),
        'row_61b':      _empty_tax(),
    }

    sheets = pd.read_excel(BytesIO(data), sheet_name=None, dtype=str, header=None)

    all_text_parts: list[str] = []
    for _sname, df in sheets.items():
        df = df.fillna('')
        for _, row in df.iterrows():
            for cell in row:
                s = str(cell).strip()
                if s:
                    all_text_parts.append(s)

    all_text = ' '.join(all_text_parts)

    for part in all_text_parts:
        upper = part.upper()
        if _GSTIN_RE.match(upper):
            result['gstin'] = upper
            result['state'], result['state_short'] = _state_info(upper)
            break

    pm = _PERIOD_RE.search(all_text)
    if pm:
        result['period'] = pm.group().title()
    else:
        mm = _PERIOD_MM_YYYY.search(all_text)
        if mm:
            result['period'] = f"{_MONTHS[int(mm.group(1)) - 1]} {mm.group(2)}"

    # ---- Filing Date (DD/MM/YYYY in Excel cell values) ----
    dates_found = _DATE_DDMMYYYY.findall(all_text)
    if dates_found:
        try:
            d, m, y = dates_found[0].split('/')
            month_abbr = _MONTHS[int(m) - 1][:3].capitalize()
            result['filing_date'] = f"{d}-{month_abbr}-{y}"
        except Exception:
            pass

    row_4_4 = _empty_tax()
    row_4_5 = _empty_tax()

    for _sname, df in sheets.items():
        df = df.fillna('')
        rows_raw = df.values.tolist()

        # Detect IGST/CGST/SGST column indices from the header row
        igst_col = cgst_col = sgst_col = None
        for raw_row in rows_raw:
            cells_lower = [str(c).strip().lower() for c in raw_row]
            if any('igst' in c or 'integrated' in c for c in cells_lower):
                for ci, c in enumerate(cells_lower):
                    if ('igst' in c or 'integrated' in c) and igst_col is None:
                        igst_col = ci
                    elif ('cgst' in c or 'central tax' in c) and cgst_col is None:
                        cgst_col = ci
                    elif ('sgst' in c or 'utgst' in c or ('state' in c and 'tax' in c)) and sgst_col is None:
                        sgst_col = ci
                break

        igst_col = igst_col if igst_col is not None else 1
        cgst_col = cgst_col if cgst_col is not None else 2
        sgst_col = sgst_col if sgst_col is not None else 3

        def get(row_data, col):
            try:
                return _parse_num(row_data[col])
            except IndexError:
                return 0.0

        n_rows = len(rows_raw)

        for ri, raw_row in enumerate(rows_raw):
            label = ' '.join(str(c) for c in raw_row).lower()

            def _row_nums(start_ri, look=2):
                """Collect strictly-numeric values from rows[start_ri : start_ri+look]."""
                out = []
                for lr in rows_raw[start_ri:start_ri + look]:
                    out.extend(_parse_num(c) for c in lr if _is_numeric(c))
                return out

            # Row 4(C) — use _is_numeric so text cells ("Net ITC Available") are skipped
            if re.search(r'net\s+itc\s+available|4\s*\(c\)', label):
                nums = _row_nums(ri)
                if len(nums) >= 3 and nums[0] + nums[1] + nums[2] > 0:
                    result['row_4c'] = {'igst': nums[0], 'cgst': nums[1], 'sgst': nums[2]}

            # Row 4(A)(4) – ISD
            if re.search(r'\(4\)|\(iv\)|inward.*isd|from\s+isd', label):
                nums = _row_nums(ri)
                if len(nums) >= 3:
                    row_4_4['igst'] += nums[0]
                    row_4_4['cgst'] += nums[1]
                    row_4_4['sgst'] += nums[2]

            # Row 4(A)(5) – All other ITC
            if re.search(r'\(5\)|\(v\)|all\s+other\s+itc', label):
                nums = _row_nums(ri)
                if len(nums) >= 3:
                    row_4_5['igst'] += nums[0]
                    row_4_5['cgst'] += nums[1]
                    row_4_5['sgst'] += nums[2]

            # Row 6.1(A) – Other than reverse charge
            if re.search(r'other\s+than\s+reverse\s+charge|6\.1.*\(a\)', label):
                nums = []
                for look in rows_raw[ri:ri+4]:
                    nums.extend(_parse_num(c) for c in look if _is_numeric(c))
                tax, itc, cash = _extract_61_section(nums)
                result['row_61a_tax']  = tax
                result['row_61a_itc']  = itc
                result['row_61a_cash'] = cash

            # Row 6.1(B) – Reverse charge (RCM)
            if (re.search(r'6\.1.*\(b\)', label) or
                    (re.search(r'reverse\s+charge', label) and 'other than' not in label)):
                nums = []
                for look in rows_raw[ri:ri+4]:
                    nums.extend(_parse_num(c) for c in look if _is_numeric(c))
                tax, _itc, _cash = _extract_61_section(nums)
                result['row_61b'] = tax

    result['row_4_45'] = {
        'igst': row_4_4['igst'] + row_4_5['igst'],
        'cgst': row_4_4['cgst'] + row_4_5['cgst'],
        'sgst': row_4_4['sgst'] + row_4_5['sgst'],
    }

    return result


# ---------------------------------------------------------------------------
# Main dispatcher
# ---------------------------------------------------------------------------

def parse_gstr3b(filename: str, data: bytes) -> dict:
    """Parse a GSTR-3B file (PDF or Excel) and return structured extracted data."""
    suffix = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''
    if suffix == 'pdf':
        return _parse_gstr3b_pdf(data)
    return _parse_gstr3b_excel(data, filename)


# ---------------------------------------------------------------------------
# Tally entry row builder
# ---------------------------------------------------------------------------

def build_tally_entries(
    parsed: dict,
    coa: dict | None = None,
    vt_master: dict | None = None,
) -> list[dict]:
    """
    Generate Tally journal entries. When coa dict is provided, ledger names are matched
    to the brand's Chart of Accounts. When vt_master dict is provided, voucher types are
    matched to the brand's Voucher Type master (e.g., 'Journal' → 'Journal UP').

    Journal 1 — ITC Availed:
      Dr Credit Ledger IGST/CGST/SGST  ← Row 4(C)
      Cr Input IGST/CGST/SGST          ← Row 4(4+5)
      Cr RCM IGST/CGST/SGST            ← Row 6.1(B) if any

    Journal 2 — ITC Utilized:
      Dr Output IGST/CGST/SGST         ← Row 6.1(A) tax amount
      Cr Credit Ledger IGST/CGST/SGST  ← ITC used
      Cr Electronic Cash Ledger        ← cash paid

    Journal 3 — RCM Tax Payment (only if row_61b > 0):
      Dr Output IGST/CGST/SGST         ← Row 6.1(B)
      Cr Electronic Cash Ledger        ← total RCM cash
    """
    st               = parsed.get('state_short', parsed.get('state', ''))
    period           = parsed.get('period', '')
    voucher_date     = parsed.get('filing_date') or _derive_voucher_date(period)
    r4c              = parsed['row_4c']
    r4_45            = parsed['row_4_45']
    r61a_tax         = parsed.get('row_61a_tax', _empty_tax())
    r61a_itc         = parsed['row_61a_itc']
    r61a_cash        = parsed['row_61a_cash']
    r61a_credit_type = parsed.get('row_61a_credit_type', _empty_tax())
    r61b             = parsed['row_61b']

    entries: list[dict] = []

    def _name(raw: str) -> str:
        return _match_coa_ledger(raw, coa, st) if coa else raw

    def _vt(base: str) -> str:
        return _match_voucher_type(base, vt_master, st) if vt_master else base

    def section(title: str):
        entries.append({'_type': 'section', 'sno': '', 'particulars': title,
                        'debit': '', 'credit': '', 'date': '', 'voucher_type': ''})

    def col_header():
        entries.append({'_type': 'header', 'sno': 'S.No', 'particulars': 'Particulars',
                        'debit': 'Debit (₹)', 'credit': 'Credit (₹)',
                        'date': 'Date', 'voucher_type': 'Voucher Type'})

    def row(sno: str, particular: str, debit='', credit=''):
        entries.append({'_type': 'data', 'sno': sno, 'particulars': _name(particular),
                        'debit': debit, 'credit': credit,
                        'date': voucher_date, 'voucher_type': _vt('Journal')})

    def blank():
        entries.append({'_type': 'blank', 'sno': '', 'particulars': '',
                        'debit': '', 'credit': '', 'date': '', 'voucher_type': ''})

    TAX = [('IGST', 'igst'), ('CGST', 'cgst'), ('SGST', 'sgst')]
    has_rcm = any(r61b[k] for k in ('igst', 'cgst', 'sgst'))

    # ── Journal 1: ITC Availed ─────────────────────────────────────────────
    section('Journal 1 — ITC Availed (Narration: ITC Availed)')
    col_header()
    sno = 1
    for tax_name, tax_key in TAX:
        row(str(sno), f'Credit Ledger {tax_name} {st}', debit=r4c[tax_key])
        sno += 1
    for tax_name, tax_key in TAX:
        row('', f'Input {tax_name} {st}', credit=r4_45[tax_key])
    if has_rcm:
        for tax_name, tax_key in TAX:
            amt = r61b[tax_key]
            if amt:
                row('', f'RCM {tax_name} {st}', credit=amt)

    blank()

    # ── Journal 2: ITC Utilized ────────────────────────────────────────────
    def tax_amt(key):
        t = r61a_tax.get(key, 0.0)
        return t if t else r61a_itc.get(key, 0.0) + r61a_cash.get(key, 0.0)

    total_cash = sum(r61a_cash[k] for k in ('igst', 'cgst', 'sgst'))
    _credit_total = sum(r61a_credit_type[k] for k in ('igst', 'cgst', 'sgst'))
    total_itc = sum(r61a_itc[k] for k in ('igst', 'cgst', 'sgst'))
    _j2_has_data = any(tax_amt(k) != 0.0 for k in ('igst', 'cgst', 'sgst')) or total_cash > 0

    if _j2_has_data:
        section('Journal 2 — ITC Utilized (Narration: ITC Utilization)')
        col_header()
        sno = 1
        for tax_name, tax_key in TAX:
            amt = tax_amt(tax_key)
            if amt != 0.0:
                row(str(sno), f'Output {tax_name} {st}', debit=amt)
                sno += 1
        if _credit_total > 0:
            for _itc_name, _itc_key in TAX:
                _amt = r61a_credit_type.get(_itc_key, 0.0)
                if _amt:
                    row('', f'Credit Ledger {_itc_name} {st}', credit=_amt)
        elif total_itc > 0:
            row('', f'Credit Ledger IGST {st}', credit=total_itc)
        if total_cash:
            row('', f'Electronic Cash Ledger {st}', credit=total_cash)
        blank()

    # ── Journal 3: RCM Tax Payment ────────────────────────────────────────
    if has_rcm:
        section('Journal 3 — RCM Tax Payment (Narration: RCM)')
        col_header()
        sno = 1
        for tax_name, tax_key in TAX:
            amt = r61b[tax_key]
            if amt:
                row(str(sno), f'Output {tax_name} {st}', debit=amt)
                sno += 1
        total_rcm = sum(r61b[k] for k in ('igst', 'cgst', 'sgst'))
        row('', f'Electronic Cash Ledger {st}', credit=total_rcm)

    return entries


# ---------------------------------------------------------------------------
# Multi-file processor
# ---------------------------------------------------------------------------

def process_multi(
    files_list: list,
    coa: dict | None = None,
    vt_master: dict | None = None,
) -> tuple[list, list, list]:
    """
    Process 1–15 GSTR-3B files in FIFO order.

    files_list: list of {"filename": str, "content": bytes}
    coa:        pre-built {lower: original} dict from _load_coa / _load_coa_from_list, or None
    vt_master:  pre-built {lower: original} dict from _load_voucher_types / _load_voucher_types_from_list, or None
    Returns:    (all_entries, monthly_data, state_summary)
    """
    all_entries: list[dict] = []
    monthly_data: list[dict] = []
    state_totals: dict[str, dict] = {}

    for idx, finfo in enumerate(files_list):
        parsed  = parse_gstr3b(finfo['filename'], finfo['content'])
        entries = build_tally_entries(parsed, coa=coa, vt_master=vt_master)

        period       = parsed.get('period', f'Month {idx + 1}')
        gstin        = parsed.get('gstin', '')
        state        = parsed.get('state', '')
        state_short  = parsed.get('state_short', '')
        voucher_date = parsed.get('filing_date') or _derive_voucher_date(period)

        total_debit  = sum(e['debit']  for e in entries
                           if e['_type'] == 'data' and isinstance(e.get('debit'),  float))
        total_credit = sum(e['credit'] for e in entries
                           if e['_type'] == 'data' and isinstance(e.get('credit'), float))

        j_debit: dict[int, float] = {}
        j_idx = 0
        for e in entries:
            if e['_type'] == 'section':
                j_idx += 1
                j_debit[j_idx] = 0.0
            elif e['_type'] == 'data' and j_idx > 0:
                d = e.get('debit', '')
                if isinstance(d, float):
                    j_debit[j_idx] = j_debit.get(j_idx, 0.0) + d

        monthly_data.append({
            'period':       period,
            'gstin':        gstin,
            'state':        state,
            'state_short':  state_short,
            'voucher_date': voucher_date,
            'entries':      entries,
            'total_debit':  total_debit,
            'total_credit': total_credit,
            'j1_debit':     j_debit.get(1, 0.0),
            'j2_debit':     j_debit.get(2, 0.0),
            'j3_debit':     j_debit.get(3, 0.0),
        })

        if idx > 0:
            _sep = {'_type': 'blank', 'sno': '', 'particulars': '',
                    'debit': '', 'credit': '', 'date': '', 'voucher_type': ''}
            all_entries.append(_sep.copy())
            all_entries.append(_sep.copy())
        all_entries.extend(entries)

        if state_short not in state_totals:
            state_totals[state_short] = {
                'state': state, 'months': 0,
                'total_debit': 0.0, 'total_credit': 0.0,
            }
        state_totals[state_short]['months']        += 1
        state_totals[state_short]['total_debit']   += total_debit
        state_totals[state_short]['total_credit']  += total_credit

    state_summary = [{'state_short': k, **v} for k, v in state_totals.items()]
    return all_entries, monthly_data, state_summary


# ---------------------------------------------------------------------------
# Excel Workbook Builder
# ---------------------------------------------------------------------------

_NAV_FILL   = PatternFill('solid', fgColor='123C69')
_SECT_FILL  = PatternFill('solid', fgColor='FF6600')
_HDR_FILL   = PatternFill('solid', fgColor='D6E4F7')
_WHITE_FONT = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
_SECT_FONT  = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
_HDR_FONT   = Font(bold=True, color='1E3A5F', name='Calibri', size=10)
_DATA_FONT  = Font(name='Calibri', size=10)
_NUM_FMT    = '#,##0.00'

_THIN   = Side(style='thin', color='CCCCCC')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _write_entries_to_sheet(ws, monthly_data: list, start_row: int = 1) -> None:
    """Write all months' entries into ws starting at start_row."""
    COLS = 6  # A–F

    ws.column_dimensions['A'].width = 7
    ws.column_dimensions['B'].width = 44
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 14

    current_row = start_row

    for m_idx, month_info in enumerate(monthly_data):
        gstin        = month_info.get('gstin', 'N/A')
        state_full   = month_info.get('state', 'N/A')
        state_short  = month_info.get('state_short', '')
        period       = month_info.get('period', 'N/A')
        voucher_date = month_info.get('voucher_date', '')

        # Month banner
        ws.merge_cells(f'A{current_row}:F{current_row}')
        c = ws.cell(row=current_row, column=1)
        c.value = (
            f'GSTIN: {gstin}    |    State: {state_full} ({state_short})'
            f'    |    Period: {period}    |    Date: {voucher_date}    |    Voucher Type: Journal'
        )
        c.font      = _WHITE_FONT
        c.fill      = _NAV_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 28
        current_row += 1

        # Blank spacer after banner
        ws.row_dimensions[current_row].height = 6
        current_row += 1

        for entry in month_info['entries']:
            rtype = entry.get('_type', 'data')

            if rtype == 'blank':
                ws.row_dimensions[current_row].height = 8
                current_row += 1
                continue

            if rtype == 'section':
                ws.merge_cells(f'A{current_row}:F{current_row}')
                c = ws.cell(row=current_row, column=1, value=entry['particulars'])
                c.font      = _SECT_FONT
                c.fill      = _SECT_FILL
                c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
                ws.row_dimensions[current_row].height = 22
                current_row += 1
                continue

            if rtype == 'header':
                spec = [
                    ('sno',          Alignment(horizontal='center', vertical='center')),
                    ('particulars',  Alignment(horizontal='left',   vertical='center')),
                    ('debit',        Alignment(horizontal='center', vertical='center')),
                    ('credit',       Alignment(horizontal='center', vertical='center')),
                    ('date',         Alignment(horizontal='center', vertical='center')),
                    ('voucher_type', Alignment(horizontal='center', vertical='center')),
                ]
                for col, (key, align) in enumerate(spec, start=1):
                    c = ws.cell(row=current_row, column=col, value=entry.get(key, ''))
                    c.font      = _HDR_FONT
                    c.fill      = _HDR_FILL
                    c.border    = _BORDER
                    c.alignment = align
                ws.row_dimensions[current_row].height = 18
                current_row += 1
                continue

            # Data row
            debit_raw  = entry.get('debit', '')
            credit_raw = entry.get('credit', '')
            debit_num  = debit_raw  if isinstance(debit_raw,  (int, float)) else None
            credit_num = credit_raw if isinstance(credit_raw, (int, float)) else None
            font = _DATA_FONT

            cells_spec = [
                (1, entry.get('sno', ''),          Alignment(horizontal='center', vertical='center')),
                (2, entry.get('particulars', ''),  Alignment(horizontal='left',   vertical='center', indent=1)),
                (3, debit_num  if debit_num  is not None else '', Alignment(horizontal='right', vertical='center')),
                (4, credit_num if credit_num is not None else '', Alignment(horizontal='right', vertical='center')),
                (5, entry.get('date', ''),          Alignment(horizontal='center', vertical='center')),
                (6, entry.get('voucher_type', ''), Alignment(horizontal='center', vertical='center')),
            ]
            for col, val, align in cells_spec:
                c = ws.cell(row=current_row, column=col, value=val)
                c.font      = font
                c.border    = _BORDER
                c.alignment = align
                if col in (3, 4) and isinstance(val, (int, float)):
                    c.number_format = _NUM_FMT
            ws.row_dimensions[current_row].height = 16
            current_row += 1

        # 2 blank rows between months
        if m_idx < len(monthly_data) - 1:
            ws.row_dimensions[current_row].height = 8
            current_row += 1
            ws.row_dimensions[current_row].height = 8
            current_row += 1

    ws.freeze_panes = 'A2'


def _write_flat_journal_sheet(
    ws,
    monthly_data: list,
    journal_keyword: str,
    narration: str,
) -> None:
    """
    Write a flat (one-row-per-entry) table for one journal type.
    journal_keyword: 'ITC Availed' or 'ITC Utilized' — matched against section labels.
    """
    HEADERS = ['Period', 'Date', 'GSTIN', 'Voucher Type', 'State', 'Ledger',
               'Debit (₹)', 'Credit (₹)', 'Narration']
    col_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
    col_widths  = [18,  16,  30,  14,  22,  40,  18,  18,  18]

    for letter, width in zip(col_letters, col_widths):
        ws.column_dimensions[letter].width = width

    for col, val in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=val)
        c.font      = _HDR_FONT
        c.fill      = _HDR_FILL
        c.border    = _BORDER
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 18

    row_num = 2
    for month in monthly_data:
        active = False
        for entry in month['entries']:
            if entry['_type'] == 'section':
                active = journal_keyword in entry['particulars']
            elif active and entry['_type'] == 'data':
                debit  = entry['debit']  if isinstance(entry['debit'],  (int, float)) else ''
                credit = entry['credit'] if isinstance(entry['credit'], (int, float)) else ''
                vals = [
                    month['period'],
                    month['voucher_date'],
                    month['gstin'],
                    entry.get('voucher_type', 'Journal'),
                    month['state'],
                    entry['particulars'],
                    debit,
                    credit,
                    narration,
                ]
                for col, val in enumerate(vals, start=1):
                    c = ws.cell(row=row_num, column=col, value=val)
                    c.font      = _DATA_FONT
                    c.border    = _BORDER
                    c.alignment = Alignment(
                        horizontal='right'  if col in (7, 8) else
                        ('left'             if col in (6, 9) else 'center'),
                        vertical='center',
                    )
                    if col in (7, 8) and isinstance(val, (int, float)):
                        c.number_format = _NUM_FMT
                ws.row_dimensions[row_num].height = 16
                row_num += 1

    ws.freeze_panes = 'A2'


def build_gstr3b_tally_workbook(
    monthly_data: list,
    all_entries: list,
    state_summary: list | None = None,
) -> Workbook:
    """
    Build Excel workbook from multi-month data.
    Sheet 1: Tally Entries (all months stacked)
    Sheet 2: Month-wise Summary
    Sheet 3: State-wise Summary
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Tally Entries'
    _write_entries_to_sheet(ws, monthly_data)

    # ── Sheet 2: ITC Availed (flat per-entry table — Journal 1) ────────────
    ws_av = wb.create_sheet('ITC Availed')
    _write_flat_journal_sheet(ws_av, monthly_data, 'ITC Availed', 'ITC Availed')

    # ── Sheet 3: ITC Utilized (flat per-entry table — Journal 2) ───────────
    ws_ut = wb.create_sheet('ITC Utilized')
    _write_flat_journal_sheet(ws_ut, monthly_data, 'ITC Utilized', 'ITC Utilization')

    # ── Sheet 4: Month-wise Summary ─────────────────────────────────────────
    ws_m = wb.create_sheet('Month-wise Summary')
    ws_m.column_dimensions['A'].width = 18
    ws_m.column_dimensions['B'].width = 32
    ws_m.column_dimensions['C'].width = 22
    ws_m.column_dimensions['D'].width = 20
    ws_m.column_dimensions['E'].width = 20
    ws_m.column_dimensions['F'].width = 22
    ws_m.column_dimensions['G'].width = 22
    ws_m.column_dimensions['H'].width = 16

    m_hdrs = ['Period', 'GSTIN', 'State', 'Total Debit (₹)', 'Total Credit (₹)',
              'ITC Availed Dr (₹)', 'ITC Utilized Dr (₹)', 'RCM Dr (₹)']
    for col, val in enumerate(m_hdrs, start=1):
        c = ws_m.cell(row=1, column=col, value=val)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.border = _BORDER
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws_m.row_dimensions[1].height = 18

    for ri, m in enumerate(monthly_data, start=2):
        vals = [
            m.get('period', ''), m.get('gstin', ''), m.get('state', ''),
            m.get('total_debit', 0.0), m.get('total_credit', 0.0),
            m.get('j1_debit', 0.0), m.get('j2_debit', 0.0), m.get('j3_debit', 0.0),
        ]
        for col, val in enumerate(vals, start=1):
            c = ws_m.cell(row=ri, column=col, value=val)
            c.font   = _DATA_FONT
            c.border = _BORDER
            c.alignment = Alignment(
                horizontal='left' if col == 2 else 'center', vertical='center'
            )
            if col >= 4 and isinstance(val, float):
                c.number_format = _NUM_FMT
        ws_m.row_dimensions[ri].height = 16

    # ── Sheet 3: State-wise Summary ─────────────────────────────────────────
    ws_s = wb.create_sheet('State-wise Summary')
    ws_s.column_dimensions['A'].width = 22
    ws_s.column_dimensions['B'].width = 14
    ws_s.column_dimensions['C'].width = 10
    ws_s.column_dimensions['D'].width = 20
    ws_s.column_dimensions['E'].width = 20

    s_hdrs = ['State', 'State Code', 'Months', 'Total Debit (₹)', 'Total Credit (₹)']
    for col, val in enumerate(s_hdrs, start=1):
        c = ws_s.cell(row=1, column=col, value=val)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.border = _BORDER
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws_s.row_dimensions[1].height = 18

    ss = state_summary or []
    for ri, s in enumerate(ss, start=2):
        vals = [
            s.get('state', ''), s.get('state_short', ''), s.get('months', 0),
            s.get('total_debit', 0.0), s.get('total_credit', 0.0),
        ]
        for col, val in enumerate(vals, start=1):
            c = ws_s.cell(row=ri, column=col, value=val)
            c.font   = _DATA_FONT
            c.border = _BORDER
            c.alignment = Alignment(
                horizontal='left' if col == 1 else 'center', vertical='center'
            )
            if col >= 4 and isinstance(val, float):
                c.number_format = _NUM_FMT
        ws_s.row_dimensions[ri].height = 16

    return wb
