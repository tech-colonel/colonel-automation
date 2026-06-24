"""
Multi-state GSTR-2B vs Books reconciliation.

Wraps the single-state engine (gstr_2b_books.py) with:
  1. Multi-file input: N GSTR-2B files, N Purchase Register files, N Debit Note files.
     Each record is tagged with _file_idx in its raw dict during parsing.
  2. Phase 2 cross-state detection: for entries that remain unmatched after main
     reconciliation, searches across other uploaded files to explain cross-state
     booking errors. Populates suggested_action_3 on matching rows.

All core matching logic (Passes 1–5 in reconcile_by_invoice_no) is UNCHANGED.
"""
from __future__ import annotations

from typing import Any

from .core import NormalizedInvoice, MatchResult
from .gstr_2b_books import (
    parse_gstr2b,
    parse_books,
    reconcile_by_invoice_no,
    build_gstr2b_books_workbook,
    GST_STATE_CODES,
    ALLOWED_SHEETS,
    _name_sim,
    _get_val,
    _extract_gstin,
    _ensure_xlsx,
    _append_rcm_rows,
    _is_octa_format,
)


def reconcile_gstr2b_vs_books_multistate(
    gstr2b_files: list[bytes],
    purchase_files: list[bytes],
    debit_files: list[bytes],
    tolerance: float = 1.0,
    entity_gstins: list[str] | None = None,
) -> tuple[list[NormalizedInvoice], list[NormalizedInvoice], list[MatchResult]]:
    """
    Phase 1 — parse all files, tag each record with _file_idx, merge into single pools,
               run existing reconcile_by_invoice_no() UNCHANGED.
    Phase 2 — for rows with suggested_action in {"Showing in 2B but Not in Books",
               "Showing in Books but Not in 2B"}, check whether the same invoice (or a
               close amount+date+name match) exists in a DIFFERENT file index.
               If found, set suggested_action_3 with a cross-state explanation.
    """

    # ── Phase 1: parse, tag, merge ────────────────────────────────────────────
    all_gstr2b: list[NormalizedInvoice] = []
    all_books: list[NormalizedInvoice] = []

    for idx, data in enumerate(gstr2b_files or []):
        if not data:
            continue
        recs = parse_gstr2b(data)
        entity_gstin = (entity_gstins[idx] if entity_gstins and idx < len(entity_gstins) else "")
        for r in recs:
            r.raw["_file_idx"] = idx
            if entity_gstin:
                r.raw["_entity_gstin"] = entity_gstin
        all_gstr2b.extend(recs)

    # Pad shorter list so zip covers all indices
    max_len = max(len(purchase_files or []), len(debit_files or []), 1)
    purchase_list = list(purchase_files or []) + [b""] * (max_len - len(purchase_files or []))
    debit_list    = list(debit_files    or []) + [b""] * (max_len - len(debit_files    or []))

    for idx, (pr_data, dn_data) in enumerate(zip(purchase_list, debit_list)):
        if not pr_data and not dn_data:
            continue
        recs = parse_books(pr_data or b"", dn_data or b"")
        entity_gstin = (entity_gstins[idx] if entity_gstins and idx < len(entity_gstins) else "")
        for r in recs:
            r.raw["_file_idx"] = idx
            if entity_gstin:
                r.raw["_entity_gstin"] = entity_gstin
        all_books.extend(recs)

    # Enrich Books records with GSTIN from GSTR-2B (same as single-state engine)
    name_to_gstin: dict[str, str] = {}
    for g in all_gstr2b:
        name_clean = str(g.supplier_name or "").strip().lower()
        if name_clean and g.supplier_gstin:
            name_to_gstin[name_clean] = g.supplier_gstin
    for b in all_books:
        if not b.supplier_gstin:
            name_clean = str(b.supplier_name or "").strip().lower()
            b.supplier_gstin = name_to_gstin.get(name_clean, "")

    # Run existing reconciliation logic UNCHANGED
    results = reconcile_by_invoice_no(all_gstr2b, all_books, tolerance)

    # ── Phase 2: cross-state Remark 3 ────────────────────────────────────────
    # Build fast lookups: normalized_doc_no → list[NormalizedInvoice]
    g_by_doc: dict[str, list[NormalizedInvoice]] = {}
    b_by_doc: dict[str, list[NormalizedInvoice]] = {}
    for rec in all_gstr2b:
        if rec.normalized_doc_no:
            g_by_doc.setdefault(rec.normalized_doc_no, []).append(rec)
    for rec in all_books:
        if rec.normalized_doc_no:
            b_by_doc.setdefault(rec.normalized_doc_no, []).append(rec)

    UNMATCHED = {"Showing in 2B but Not in Books", "Showing in Books but Not in 2B"}

    for r in results:
        g = _get_val(r, "gstr2b", None)
        b = _get_val(r, "purchase", None)
        status = _get_val(r, "suggested_action", "")

        # Matched but from different state files — cross-state booking that happened to match
        if g and b and status not in UNMATCHED:
            g_idx = g.raw.get("_file_idx")
            b_idx = b.raw.get("_file_idx")
            if g_idx is not None and b_idx is not None and g_idx != b_idx:
                g_state = GST_STATE_CODES.get((g.supplier_gstin or "")[:2], "Unknown State")
                r.suggested_action_3 = (
                    f"Cross-state entry: GSTR-2B from file {g_idx + 1} "
                    f"(GSTIN {g.supplier_gstin} – {g_state}), "
                    f"booked in Books file {b_idx + 1}. Verify Tally state allocation."
                )
            continue

        if status not in UNMATCHED:
            continue

        src = g or b
        if src is None:
            continue
        src_idx = src.raw.get("_file_idx")

        if g and not b:
            # In 2B but not in Books — search other Books files
            found = False
            if g.normalized_doc_no:
                # Tier 1: exact invoice number match across other state's Books file
                for c in b_by_doc.get(g.normalized_doc_no, []):
                    c_idx = c.raw.get("_file_idx")
                    if c_idx != src_idx and abs(c.taxable_value - g.taxable_value) <= tolerance:
                        state = GST_STATE_CODES.get((g.supplier_gstin or "")[:2], "Unknown State")
                        r.suggested_action_3 = (
                            f"Cross-state: Invoice {g.doc_no} found in Books file {(c_idx or 0) + 1}. "
                            f"Vendor GSTIN {g.supplier_gstin} ({state}). "
                            f"Likely booked under wrong state."
                        )
                        found = True
                        break
            if not found:
                # Tier 2: amount + YYYY-MM + supplier name similarity (for entries without invoice numbers)
                g_month = (g.doc_date or "")[:7]
                for c in all_books:
                    c_idx = c.raw.get("_file_idx")
                    if (c_idx != src_idx and g_month
                            and abs(c.taxable_value - g.taxable_value) <= tolerance
                            and (c.doc_date or "")[:7] == g_month
                            and _name_sim(g.supplier_name, c.supplier_name) >= 0.5):
                        state = GST_STATE_CODES.get((g.supplier_gstin or "")[:2], "Unknown State")
                        r.suggested_action_3 = (
                            f"Possible cross-state: ₹{g.taxable_value:,.2f} on {g_month} "
                            f"matches Books file {(c_idx or 0) + 1} "
                            f"(GSTIN {g.supplier_gstin} – {state}). "
                            f"No invoice number — verify manually."
                        )
                        break

        elif b and not g:
            # In Books but not in 2B — search other GSTR-2B files
            found = False
            if b.normalized_doc_no:
                # Tier 1: exact invoice number match across other state's GSTR-2B file
                for c in g_by_doc.get(b.normalized_doc_no, []):
                    c_idx = c.raw.get("_file_idx")
                    if c_idx != src_idx and abs(c.taxable_value - b.taxable_value) <= tolerance:
                        state = GST_STATE_CODES.get((c.supplier_gstin or "")[:2], "Unknown State")
                        r.suggested_action_3 = (
                            f"Cross-state: Invoice {b.doc_no} found in GSTR-2B file {(c_idx or 0) + 1} "
                            f"under GSTIN {c.supplier_gstin} ({state}). "
                            f"Check if wrong state booked in Books."
                        )
                        found = True
                        break
            if not found:
                # Tier 2: amount + YYYY-MM + supplier name similarity
                b_month = (b.doc_date or "")[:7]
                for c in all_gstr2b:
                    c_idx = c.raw.get("_file_idx")
                    if (c_idx != src_idx and b_month
                            and abs(c.taxable_value - b.taxable_value) <= tolerance
                            and (c.doc_date or "")[:7] == b_month
                            and _name_sim(b.supplier_name, c.supplier_name) >= 0.5):
                        state = GST_STATE_CODES.get((c.supplier_gstin or "")[:2], "Unknown State")
                        r.suggested_action_3 = (
                            f"Possible cross-state: ₹{b.taxable_value:,.2f} on {b_month} "
                            f"matches GSTR-2B file {(c_idx or 0) + 1} "
                            f"under GSTIN {c.supplier_gstin} ({state}). "
                            f"No invoice number — verify manually."
                        )
                        break

    return all_gstr2b, all_books, results


def _find_data_start_row(src_ws, max_scan: int = 15) -> int:
    """Return 1-based row index of first DATA row (row after column headers).

    Uses the LAST header row found, not the first.  GSTR-2B portal exports have
    two header rows (group headers at row 5, sub-headers at row 6).  The group
    header row contains 'GSTIN of Supplier' which would trigger a first-match
    return of row 6 — including the sub-header row as a garbage data row.
    Using the last match correctly identifies row 6 as the final header row and
    returns 7 as data start.  For single-header Tally files the behaviour is
    unchanged."""
    HEADER_KEYWORDS = {
        'date', 'gstin', 'gstin of supplier', 'invoice number',
        'particulars', 'voucher no', 'voucher type', 'narration',
    }
    last_header_row = 0
    for row_idx, row in enumerate(src_ws.iter_rows(max_row=max_scan, values_only=True), start=1):
        for cell in row:
            if cell and str(cell).strip().lower() in HEADER_KEYWORDS:
                last_header_row = row_idx
                break  # found keyword in this row; keep scanning for a later header row
    return last_header_row + 1 if last_header_row else 2


def _merge_state_into_sheets(wb, file_bytes: bytes, label_prefix: str) -> None:
    """Append data rows from an additional state's file into the matching output sheets.
    label_prefix is "2B", "PR", or "DN" — mirrors _copy_workbook_sheets() naming convention."""
    import openpyxl
    from io import BytesIO
    _allowed_2b = {s.upper() for s in ALLOWED_SHEETS}
    pfx = f"{label_prefix} - "
    try:
        xlsx_bytes = _ensure_xlsx(file_bytes)
        # OCTA 2B exports are a single flat sheet — copy it as-is, don't filter to B2B tabs.
        is_octa_2b = label_prefix == "2B" and _is_octa_format(xlsx_bytes)
        src_wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)

        for src_sheet_name in src_wb.sheetnames:
            # For 2B only copy the same sheets the base engine copies
            if label_prefix == "2B" and not is_octa_2b and src_sheet_name.upper().strip() not in _allowed_2b:
                continue

            # Match to existing output sheet
            target_name = (pfx + src_sheet_name)[:31]
            if target_name not in wb.sheetnames:
                # Fallback: find any wb sheet that starts with the same prefix
                candidates = [n for n in wb.sheetnames if n.startswith(pfx)]
                if candidates:
                    # Prefer partial-name match; fall back to first candidate so state 2-N
                    # data always lands in the output tab even with non-standard sheet names.
                    refined = [n for n in candidates
                               if src_sheet_name[:8].upper() in n.upper()
                               or n[len(pfx):][:8].upper() in src_sheet_name.upper()]
                    target_name = refined[0] if refined else candidates[0]
                else:
                    target_name = None
            if not target_name:
                continue

            dst_ws = wb[target_name]
            src_ws = src_wb[src_sheet_name]
            data_start = _find_data_start_row(src_ws)

            for row_vals in src_ws.iter_rows(min_row=data_start, values_only=True):
                if any(v is not None for v in row_vals):  # skip fully empty rows
                    dst_ws.append(list(row_vals))

        src_wb.close()
    except Exception:
        pass  # If a state file can't be read, skip silently


def build_gstr2b_books_multistate_workbook(
    results: list[Any],
    payload: dict | None = None,
) -> Any:
    """
    Build Excel workbook using the base single-state engine, then:
    - Add per-state source sheets for states 2-N (2B / PR / DN tabs)
    - Inject Remark 3 as column W (23) on 'Reco 2B vs Books'
    - Inject Remark 3 as column U (21) on 'Vendor Summary'
    Remark 3 cells use amber-orange font (FF6600).
    """
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    remark3_font = Font(color="FF6600")
    header_font  = Font(bold=True, size=10)
    thin         = Side(style="thin")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    center       = Alignment(horizontal="center", vertical="center")

    # Build base_payload for source sheet copying — use only the first file of each type
    base_payload: dict | None = None
    if payload:
        base_payload = {k: v for k, v in payload.items()
                        if k not in ("_gstr2b_b64", "_purchase_b64", "_debit_b64")}
        for field in ("_gstr2b_b64", "_purchase_b64", "_debit_b64"):
            # Accept both "field_0" (multi-state numbering) and bare "field" (legacy single-state)
            for candidate in (field, f"{field}_0"):
                if candidate in payload:
                    base_payload[field] = payload[candidate]
                    break

    wb = build_gstr2b_books_workbook(results, base_payload)

    # ── Add per-state source sheets for states 2-N ───────────────────────────
    import base64 as _b64
    all_gstr2b_b64   = (payload or {}).get("_all_gstr2b_b64",   [])
    all_purchase_b64 = (payload or {}).get("_all_purchase_b64", [])
    all_debit_b64    = (payload or {}).get("_all_debit_b64",    [])

    for idx in range(1, len(all_gstr2b_b64)):      # skip state 1 (already in base wb)
        _merge_state_into_sheets(wb, _b64.b64decode(all_gstr2b_b64[idx]),   "2B")
    for idx in range(1, len(all_purchase_b64)):
        _merge_state_into_sheets(wb, _b64.b64decode(all_purchase_b64[idx]), "PR")
    for idx in range(1, len(all_debit_b64)):
        _merge_state_into_sheets(wb, _b64.b64decode(all_debit_b64[idx]),    "DN")

    # ── Append RCM rows from states 2-N into the RCM sheet ───────────────────
    # State 1 RCM was already written by build_gstr2b_books_workbook.
    # We append data rows (no headers) from each additional state's GSTR-2B file.
    for idx in range(1, len(all_gstr2b_b64)):
        _append_rcm_rows(_b64.b64decode(all_gstr2b_b64[idx]), wb)

    # ── Reco 2B vs Books: add Remark 3 as column AC (29) ─────────────────────
    # +6 columns vs single-state: GSTIN+Source+State for both 2B and Books sections
    REMARK3_COL = 29
    if "Reco 2B vs Books" in wb.sheetnames:
        ws = wb["Reco 2B vs Books"]

        hdr = ws.cell(row=3, column=REMARK3_COL, value="Remark 3")
        hdr.font      = header_font
        hdr.alignment = center
        hdr.border    = border
        ws.column_dimensions[get_column_letter(REMARK3_COL)].width = 55

        # Data rows in _populate_workbook start at row 4 (rows 1-3 are headers)
        for row_idx, r in enumerate(results, start=4):
            val = getattr(r, "suggested_action_3", "") or ""
            c   = ws.cell(row=row_idx, column=REMARK3_COL, value=val)
            c.border = border
            if val:
                c.font = remark3_font

    # ── Vendor Summary: add Remark 3 as column U (21) ────────────────────────
    REMARK3_VS_COL = 21
    if "Vendor Summary" in wb.sheetnames:
        ws_vs = wb["Vendor Summary"]

        # Build GSTIN → list[remark3] from results for fast lookup
        gstin_to_r3: dict[str, list[str]] = {}
        for r in results:
            val = getattr(r, "suggested_action_3", "") or ""
            if not val:
                continue
            g = _get_val(r, "gstr2b", None)
            b = _get_val(r, "purchase", None)
            key = _extract_gstin(g) if g else ""
            if not key and b:
                key = _extract_gstin(b)
            if not key:
                name = str(_get_val(g or b, "supplier_name", "") or "").strip().upper()
                key  = f"__name__{name}" if name else "__unknown__"
            gstin_to_r3.setdefault(key, []).append(val)

        hdr = ws_vs.cell(row=3, column=REMARK3_VS_COL, value="Remark 3")
        hdr.font      = header_font
        hdr.alignment = center
        hdr.border    = border
        ws_vs.column_dimensions[get_column_letter(REMARK3_VS_COL)].width = 55

        # Data rows in _build_vendor_summary_sheet start at row 4
        for row_idx in range(4, ws_vs.max_row + 1):
            gstin_val = str(ws_vs.cell(row=row_idx, column=2).value or "").strip().upper()
            r3_list: list[str] = []
            if gstin_val:
                r3_list = gstin_to_r3.get(gstin_val, [])
            if not r3_list:
                # Fallback: try matching by common name (column A)
                name_val = str(ws_vs.cell(row=row_idx, column=1).value or "").strip().upper()
                if name_val:
                    r3_list = gstin_to_r3.get(f"__name__{name_val}", [])
            if r3_list:
                display = (r3_list[0] if len(r3_list) == 1
                           else f"Cross-state entries ({len(r3_list)}): {r3_list[0]}")
                c = ws_vs.cell(row=row_idx, column=REMARK3_VS_COL, value=display)
                c.border = border
                c.font   = remark3_font
            else:
                ws_vs.cell(row=row_idx, column=REMARK3_VS_COL).border = border

    return wb
